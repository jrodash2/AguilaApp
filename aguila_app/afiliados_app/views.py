from datetime import datetime, timezone
import logging
from venv import logger
from django.forms import IntegerField
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login as auth_login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import Group, User
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse
import openpyxl
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
import requests
import unicodedata
from .form import AfiliadoForm, CentroVotacionForm, ComisionForm, ComunidadForm, PerfilForm, SectorForm, UserCreateForm, UserEditForm, InstitucionForm, OrganizacionIntegranteForm, CoordinadorOrganizacionForm, LiderComunitarioOrganizacionForm, EstructuraOrganizativaForm, ResponsableTerritorialForm, ReunionTerritorialForm, IncidenciaTerritorialForm, JuventudIntegranteForm, CoordinadorJuventudForm, LiderJuvenilForm, EstructuraJuventudForm, ResponsableJuventudForm, ReunionJuventudForm, IncidenciaJuventudForm, MujeresIntegranteForm, CoordinadoraMujeresForm, LiderMujeresForm, EstructuraMujeresForm, ResponsableMujeresForm, ReunionMujeresForm, IncidenciaMujeresForm, LogisticaIntegranteForm, CoordinadorLogisticaForm, LiderLogisticaForm, EstructuraLogisticaForm, ResponsableLogisticaForm, ReunionLogisticaForm, IncidenciaLogisticaForm, RecursoLogisticoForm, AsignacionLogisticaForm, SolicitudLogisticaForm, EntregaLogisticaForm, AgendaLogisticaForm, ComunicacionIntegranteForm, CoordinadorComunicacionForm, LiderComunicacionForm, EstructuraComunicacionForm, ResponsableComunicacionForm, ReunionComunicacionForm, IncidenciaComunicacionForm, AgendaPublicacionForm, SolicitudContenidoForm, CoberturaActividadForm, BancoMediosForm, CampanaComunicacionForm, MonitoreoRedForm, PlanHormigaIntegranteForm, CoordinadorPlanHormigaForm, EnlaceTerritorialPlanHormigaForm, EstructuraPlanHormigaForm, ResponsableZonaPlanHormigaForm, ReunionPlanHormigaForm, IncidenciaPlanHormigaForm, VisitaPlanHormigaForm, SeguimientoContactoPlanHormigaForm, CompromisoTerritorialPlanHormigaForm, PuntoVisitadoPlanHormigaForm, ActivacionTerritorialPlanHormigaForm, CoberturaVisitaPlanHormigaForm
from .models import Afiliado, CentroVotacion, Comision, Comunidad, Eleccion2023, Perfil, Institucion, Sector, PadronElectoral, OrganizacionIntegrante, CoordinadorOrganizacion, LiderComunitarioOrganizacion, EstructuraOrganizativa, ResponsableTerritorial, ReunionTerritorial, IncidenciaTerritorial, EstructuraIntegrante, EstadoRegistro, JuventudIntegrante, CoordinadorJuventud, LiderJuvenil, EstructuraJuventud, ResponsableJuventud, ReunionJuventud, IncidenciaJuventud, EstructuraIntegranteJuventud, MujeresIntegrante, CoordinadoraMujeres, LiderMujeres, EstructuraMujeres, ResponsableMujeres, ReunionMujeres, IncidenciaMujeres, EstructuraIntegranteMujeres, LogisticaIntegrante, CoordinadorLogistica, LiderLogistica, EstructuraLogistica, ResponsableLogistica, ReunionLogistica, IncidenciaLogistica, EstructuraIntegranteLogistica, RecursoLogistico, AsignacionLogistica, SolicitudLogistica, EntregaLogistica, AgendaLogistica, ComunicacionIntegrante, CoordinadorComunicacion, LiderComunicacion, EstructuraComunicacion, ResponsableComunicacion, ReunionComunicacion, IncidenciaComunicacion, EstructuraIntegranteComunicacion, AgendaPublicacion, SolicitudContenido, CoberturaActividad, BancoMedios, CampanaComunicacion, MonitoreoRed, PlanHormigaIntegrante, CoordinadorPlanHormiga, EnlaceTerritorialPlanHormiga, EstructuraPlanHormiga, ResponsableZonaPlanHormiga, ReunionPlanHormiga, IncidenciaPlanHormiga, EstructuraIntegrantePlanHormiga, VisitaPlanHormiga, SeguimientoContactoPlanHormiga, CompromisoTerritorialPlanHormiga, PuntoVisitadoPlanHormiga, ActivacionTerritorialPlanHormiga, CoberturaVisitaPlanHormiga
from django.views.generic import CreateView
from django.views.generic import ListView
from django.urls import reverse_lazy
from django.http import Http404, HttpResponseNotAllowed, JsonResponse, HttpResponseForbidden
from django.core.exceptions import ValidationError
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.db import models
from django.db.models import Sum, F, Value, Count, Q, Case, When, OuterRef, Subquery, IntegerField, Prefetch
from django.contrib.auth.decorators import login_required, user_passes_test
from collections import defaultdict
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
import json
from django.contrib.auth.models import Group
from django.views.decorators.http import require_GET
from django.db.models.functions import Coalesce
from django.db import transaction, IntegrityError, connection
from django.db.models.deletion import ProtectedError
from django.db.models import Sum
from django.shortcuts import render
from django.template.loader import render_to_string
from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from weasyprint import HTML
from django.db.models.functions import Cast
from django.utils import timezone
from datetime import timedelta
from reportlab.lib.pagesizes import landscape, letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import datetime
from django.core.mail import send_mail
from django.conf import settings
from django.utils.html import strip_tags
from decimal import Decimal, InvalidOperation
from datetime import datetime
from datetime import datetime as DateTime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import re
import uuid
from urllib.parse import urlencode
from django.views.generic.detail import DetailView
from django.core.mail import BadHeaderError
from smtplib import SMTPException
from django.core.serializers.json import DjangoJSONEncoder

from .forms import PadronUploadForm

logger = logging.getLogger(__name__)


def _agrupar_fechas_por_mes_seguro(queryset, campo='fecha_creacion'):
    """Agrupa fechas en Python sin funciones de zona horaria de la base de datos."""
    conteos = defaultdict(int)
    modelo = getattr(getattr(queryset, 'model', None), '__name__', 'desconocido')

    for fecha in queryset.values_list(campo, flat=True):
        if fecha is None or not hasattr(fecha, 'year') or not hasattr(fecha, 'month'):
            logger.warning(
                "Fecha inválida detectada en crecimiento mensual. modelo=%s valor=%r",
                modelo,
                fecha,
            )
            continue

        # Los DateTimeField llegan de MySQL como valores UTC; la conversión se hace
        # aquí para conservar el mes de TIME_ZONE sin depender de CONVERT_TZ.
        if isinstance(fecha, DateTime) and timezone.is_aware(fecha):
            fecha = timezone.localtime(fecha)

        anio = fecha.year
        mes = fecha.month
        if not isinstance(anio, int) or not isinstance(mes, int) or not 1 <= mes <= 12:
            logger.warning(
                "Fecha inválida detectada en crecimiento mensual. modelo=%s valor=%r",
                modelo,
                fecha,
            )
            continue

        conteos['{:04d}-{:02d}'.format(anio, mes)] += 1

    return [
        {'mes': mes, 'total': conteos[mes]}
        for mes in sorted(conteos)
    ]


def _construir_series_crecimiento(querysets, campo='fecha_creacion'):
    """Construye el contrato común de datos crudos y series de los panoramas."""
    crecimiento = {
        nombre: _agrupar_fechas_por_mes_seguro(queryset, campo)
        for nombre, queryset in querysets.items()
    }
    meses = sorted({
        item['mes']
        for datos in crecimiento.values()
        for item in datos
    })
    mapas = {
        nombre: {item['mes']: item['total'] for item in datos}
        for nombre, datos in crecimiento.items()
    }
    series = {'labels': meses}
    for nombre in querysets:
        series[nombre] = [mapas[nombre].get(mes, 0) for mes in meses]
    return crecimiento, series


def safe_context(request, extra=None):
    user = getattr(request, "user", None)
    is_authenticated = bool(user and user.is_authenticated)

    base = {
        "view_key": "",
        "view_label": "",
        "next_url": "",
        "user_groups": [g.name for g in user.groups.all()] if is_authenticated else [],
        "es_admin": user.groups.filter(name="Administrador").exists() if is_authenticated else False,
    }

    if extra:
        base.update(extra)

    return base


def safe_render(request, template_name, context=None, *args, **kwargs):
    return render(request, template_name, safe_context(request, context), *args, **kwargs)


from django.db.models import Sum
from django.shortcuts import render
from .models import TrepCentroResultado
import json

from django.db.models import Sum
from django.shortcuts import render
from .models import TrepCentroResultado
import json

def elecciones2023(request):
    raise Http404()


    
    
@login_required
def editar_institucion(request):
    institucion = Institucion.objects.first()  # Solo debería haber una

    if request.method == 'POST':
        form = InstitucionForm(request.POST, request.FILES, instance=institucion)
        if form.is_valid():
            form.save()
            messages.success(request, "Datos institucionales actualizados correctamente.")
            return redirect('afiliados:editar_institucion')  # Reemplaza con la URL real
    else:
        form = InstitucionForm(instance=institucion)

    return safe_render(request, 'afiliados/editar_institucion.html', {'form': form})



@login_required
def user_create(request):
    if request.method == 'POST':
        form = UserCreateForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            password = form.cleaned_data.get('new_password')
            user.set_password(password)
            user.save()

            group = form.cleaned_data.get('group')
            user.groups.add(group)

            # ✅ Espera a que la señal cree el perfil automáticamente
            foto = form.cleaned_data.get('foto')
            try:
                perfil = user.perfil  # accede al perfil creado por la señal
                if foto:
                    perfil.foto = foto
                    perfil.save()
            except Perfil.DoesNotExist:
                # Fallback solo si la señal falló (raro)
                Perfil.objects.create(user=user, foto=foto)

            messages.success(request, 'Usuario creado correctamente.')
            return redirect('afiliados:user_create')
    else:
        form = UserCreateForm()

    users = User.objects.all()
    return safe_render(request, 'afiliados/user_form_create.html', {'form': form, 'users': users})

@login_required
def user_edit(request, user_id):
    user = get_object_or_404(User, pk=user_id)

    if request.method == 'POST':
        form = UserEditForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario editado correctamente.')
            return redirect('afiliados:user_edit', user_id=user.id)
    else:
        form = UserEditForm(instance=user)

    context = {
        'form': form,
        'user': user,
        'users': User.objects.all(),
    }
    return safe_render(request, 'afiliados/user_form_edit.html', context)




@login_required
def user_delete(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.delete()
        return redirect('afiliados:user_create')  # Redirige a la misma página para mostrar la lista actualizada
    return safe_render(request, 'afiliados/user_confirm_delete.html', {'user': user})


def home(request):
    return safe_render(request, 'afiliados/login.html')

from datetime import timedelta
from django.utils import timezone
from django.db.models import Count, Q, Sum
import json

from django.db.models.functions import Coalesce
from django.db.models import Value, Count


@login_required
def dahsboard(request):

    # === Totales ===
    total_afiliados = Afiliado.objects.count()
    total_lideres = Afiliado.objects.filter(es_lider_comunitario=True).count()
    total_comunidades = Comunidad.objects.count()
    total_empadronados = Afiliado.objects.filter(empadronado=True).count()
    if total_afiliados == 0:
        logger.info("Dashboard sin afiliados registrados.")

    # === Últimos afiliados ===
    ultimos_afiliados = (
        Afiliado.objects.select_related('comunidad')
        .order_by('-id')[:10]
    )

    # === Afiliados por comunidad ===
    # Nota: si comunidad es NULL, el render anterior enviaba "None" (Python) al JS,
    # lo cual rompía los gráficos. Coalesce + JSON válido evita el fallo.
    afiliados_por_comunidad = (
        Afiliado.objects
        .annotate(
            comunidad_nombre=Coalesce('comunidad__nombre', Value('SIN COMUNIDAD'))
        )
        .values('comunidad_nombre')
        .annotate(total=Count('id'))
        .order_by('comunidad_nombre')
    )
    if not afiliados_por_comunidad:
        logger.info("Dashboard sin afiliados por comunidad.")

    # === Líderes por comunidad ===
    lideres_por_comunidad = (
        Afiliado.objects
        .filter(es_lider_comunitario=True)
        .annotate(
            comunidad_nombre=Coalesce('comunidad__nombre', Value('SIN COMUNIDAD'))
        )
        .values('comunidad_nombre')
        .annotate(total=Count('id'))
        .order_by('comunidad_nombre')
    )
    if not lideres_por_comunidad:
        logger.info("Dashboard sin líderes por comunidad.")

    # === Afiliados por rango edad ===
    hoy = timezone.now().date()
    rangos = {"18-25": 0, "26-35": 0, "36-45": 0, "46-60": 0, "60+": 0}

    afiliados = Afiliado.objects.exclude(fecha_nacimiento__isnull=True)
    for a in afiliados:
        fn = a.fecha_nacimiento
        edad = hoy.year - fn.year - ((hoy.month, hoy.day) < (fn.month, fn.day))

        if edad <= 25: rangos["18-25"] += 1
        elif edad <= 35: rangos["26-35"] += 1
        elif edad <= 45: rangos["36-45"] += 1
        elif edad <= 60: rangos["46-60"] += 1
        else: rangos["60+"] += 1

    rangos_edad = [{"rango": r, "total": rangos[r]} for r in rangos]

    total_no_empadronados = Afiliado.objects.filter(empadronado=False).count()

# ==============================================================
# ⭐ 1. NUEVA GRAFICA: Afiliados creados por mes
# ==============================================================
    afiliados_por_mes = _agrupar_fechas_por_mes_seguro(
        Afiliado.objects.all()
    )

# Convertir mes número → nombre (para ApexCharts)
    meses_nombre = [
        "Enero","Febrero","Marzo","Abril","Mayo","Junio",
        "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"
    ]

    afiliados_por_mes_list = []
    for item in afiliados_por_mes:
        anio, mes_num = (int(parte) for parte in item["mes"].split("-"))
        afiliados_por_mes_list.append({
            "mes": meses_nombre[mes_num - 1],
            "anio": anio,
            "total": item.get("total", 0)
        })
    if not afiliados_por_mes_list:
        logger.info("Dashboard sin afiliados registrados por mes.")

    # ==============================================================
    # ⭐ 2. NUEVA GRAFICA: Comunidades por Sector
    # ==============================================================

    comunidades_por_sector = (
    Comunidad.objects
        .values(nombre_sector=Coalesce('sector__nombre', Value('Sin Sector')))
        .annotate(total=Count('id'))
        .order_by('nombre_sector')
    )
    if not comunidades_por_sector:
        logger.info("Dashboard sin comunidades por sector.")

    # === CONTEXT ===
    context = {
        'total_afiliados': total_afiliados,
        'total_lideres': total_lideres,
        'total_comunidades': total_comunidades,
        'total_empadronados': total_empadronados,
        'ultimos_afiliados': ultimos_afiliados,
        'afiliados_por_comunidad': json.dumps(
            list(afiliados_por_comunidad), ensure_ascii=False
        ),
        'lideres_por_comunidad': json.dumps(
            list(lideres_por_comunidad), ensure_ascii=False
        ),
        'rangos_edad': json.dumps(rangos_edad, ensure_ascii=False),

        # Nuevos datos
        'afiliados_por_mes': json.dumps(afiliados_por_mes_list, ensure_ascii=False),
        'comunidades_por_sector': json.dumps(
            list(comunidades_por_sector), ensure_ascii=False
        ),

        'empadronados': json.dumps({
            'empadronados': total_empadronados,
            'no_empadronados': total_no_empadronados,
        }, ensure_ascii=False),
    }

    return safe_render(request, 'afiliados/dahsboard.html', context)





def signout(request):
    logout(request)
    return redirect('afiliados:signin')


def signin(request):  
    institucion = Institucion.objects.first()
    if request.method == 'GET':
        # Deberías instanciar el AuthenticationForm correctamente
        return safe_render(request, 'afiliados/login.html', {
            'form': AuthenticationForm(),
            'institucion': institucion,
        })
    else:
        # Se instancia AuthenticationForm con los datos del POST para mantener el estado
        form = AuthenticationForm(request, data=request.POST)
        
        if form.is_valid():
            # El método authenticate devuelve el usuario si es válido
            user = form.get_user()
            
            # Si el usuario es encontrado, se inicia sesión
            auth_login(request, user)
            
            # Ahora verificamos los grupos
            for g in user.groups.all():
                print(g.name)
                if g.name == 'Administrador':
                    return redirect('afiliados:dahsboard')
                elif g.name == 'Gestor':
                    return redirect('afiliados:dahsboard')
                elif g.name == 'afiliados':
                    return redirect('afiliados:dahsboard')
                elif g.name == 'Organizacion':
                    return redirect('afiliados:dashboard_organizacion')
                elif g.name == 'Jovenes':
                    return redirect('afiliados:dashboard_juventud')
                elif g.name == 'Mujeres':
                    return redirect('afiliados:dashboard_mujeres')
                elif g.name == 'Logistica':
                    return redirect('afiliados:dashboard_logistica')
                elif g.name == 'Comunicacion':
                    return redirect('afiliados:dashboard_comunicacion')
                elif g.name == 'PlanHormiga':
                    return redirect('afiliados:dashboard_plan_hormiga')
            # Si no se encuentra el grupo adecuado, se redirige a una página por defecto
            return redirect('afiliados:dahsboard')
        else:
            # Si el formulario no es válido, se retorna con el error
            return safe_render(request, 'afiliados/login.html', {
                'form': form,  # Pasamos el formulario con los errores
                'error': 'Usuario o contraseña incorrectos',
                'institucion': institucion,
            })



def elecciones_dashboard(request):
    TOP_PARTIDOS_GLOBAL = 12

    tipo = (request.GET.get("tipo") or "ALCALDE").strip().upper()
    if tipo not in {"ALCALDE", "DIPUTADOS"}:
        tipo = "ALCALDE"

    qs = TrepCentroResultado.objects.filter(tipo=tipo)
    qs_count = qs.count()

    totales = {}
    for row in qs.iterator():
        for partido, votos in (row.partidos or {}).items():
            partido_key = str(partido or "").strip().upper()
            totales[partido_key] = totales.get(partido_key, 0) + int(votos or 0)

    # Fuente única para tabla + gráfica global.
    ranking = sorted(totales.items(), key=lambda x: x[1], reverse=True)
    total_votos_global = int(sum(v for _, v in ranking))

    ranking_top = ranking[:TOP_PARTIDOS_GLOBAL]
    resto_global = int(sum(v for _, v in ranking[TOP_PARTIDOS_GLOBAL:]))
    if resto_global > 0:
        ranking_top.append(("OTROS", resto_global))

    tabla_global = []
    totales_chart = {}
    for partido, votos in ranking_top:
        votos_int = int(votos)
        porcentaje = round((votos_int / total_votos_global) * 100, 2) if total_votos_global else 0
        tabla_global.append({"partido": partido, "votos": votos_int, "porcentaje": porcentaje})
        totales_chart[partido] = votos_int

    centros = list(
        qs.exclude(centro_codigo__isnull=True)
        .exclude(centro_codigo="")
        .values("centro_codigo", "centro_nombre")
        .distinct()
        .order_by("centro_nombre", "centro_codigo")
    )

    for c in centros:
        if not c.get("centro_nombre"):
            c["centro_nombre"] = c.get("centro_codigo")

    centro_default = centros[0]["centro_codigo"] if centros else ""

    if settings.DEBUG:
        logger.info(
            "Elecciones dashboard -> tipo=%s qs_count=%s total_votos=%s len_totales=%s centros=%s",
            tipo,
            qs_count,
            total_votos_global,
            len(totales),
            len(centros),
        )

    global_payload = {"totales": totales_chart, "total": total_votos_global}
    context = {
        "tipo": tipo,
        "tabla_global": tabla_global,
        "global_json": json.dumps(global_payload, cls=DjangoJSONEncoder),
        "centros": centros,
        "centro_default": centro_default,
    }

    return safe_render(request, "afiliados/elecciones_dashboard.html", context)


def elecciones_centro_datos(request, centro_codigo):
    tipo = (request.GET.get("tipo") or "ALCALDE").strip().upper()
    if tipo not in {"ALCALDE", "DIPUTADOS"}:
        tipo = "ALCALDE"

    qs = TrepCentroResultado.objects.filter(tipo=tipo, centro_codigo=centro_codigo)

    totales = {}
    mesas = []
    centro_nombre = ""

    for row in qs.iterator():
        centro_nombre = row.centro_nombre or row.centro_codigo or centro_codigo
        partidos_row = row.partidos or {}

        total_mesa = 0
        top_partido = "-"
        top_partido_votos = -1

        for p, v in partidos_row.items():
            key = str(p or "").strip().upper()
            votos = int(v or 0)
            totales[key] = totales.get(key, 0) + votos
            total_mesa += votos
            if votos > top_partido_votos:
                top_partido_votos = votos
                top_partido = key

        mesas.append({
            "mesa": str(row.mesa or "-"),
            "total": int(total_mesa),
            "top_partido": top_partido,
        })

    total_votos = int(sum(totales.values()))
    ranking = [
        {
            "partido": k,
            "votos": int(v),
            "pct": round((int(v) / total_votos) * 100, 2) if total_votos else 0,
        }
        for k, v in sorted(totales.items(), key=lambda x: x[1], reverse=True)
    ]

    mesas.sort(key=lambda x: x["total"], reverse=True)

    return JsonResponse(
        {
            "centro_codigo": centro_codigo,
            "centro_nombre": centro_nombre or centro_codigo,
            "total_votos": total_votos,
            "totales_partidos": totales,
            "ranking": ranking,
            "mesas": mesas,
        },
        safe=True,
    )

def dashboard_elecciones(request):

    centros = Eleccion2023.objects.values_list(
        "centro_votacion", flat=True
    ).distinct().order_by("centro_votacion")

    partidos = [
        "todos","cambio","morena","vamos","pin","renovador",
        "valor","azul","une","fcn_nacion","podemos","uc"
    ]

    # 🔥 GENERAR TOTALES GLOBALES
    totales_globales = {
        p: Eleccion2023.objects.aggregate(total=Sum(p))["total"] or 0
        for p in partidos
    }

    # 🔥 ORDENAR DE MAYOR A MENOR Y AGREGAR RANKING
    ranking_global = sorted(
        totales_globales.items(),
        key=lambda x: x[1],
        reverse=True
    )

    return safe_render(request, "afiliados/elecciones2023.html", {
        "centros": centros,
        "ranking_global": ranking_global,  # ⬅ nueva variable
    })



from django.http import JsonResponse
from django.db.models import Sum

def datos_centro(request):
    centro = request.GET.get("centro", "").strip()
    mesa = request.GET.get("mesa", "").strip()

    if not centro:
        return JsonResponse({"error": "Centro no enviado"})

    queryset = Eleccion2023.objects.filter(centro_votacion=centro)

    if not queryset.exists():
        return JsonResponse({"error": "Centro no encontrado"})

    # 🔥 SOLO filtrar mesa si no es "0"
    if mesa and mesa != "0":
        queryset = queryset.filter(mesa=mesa)

    partidos = [
        "todos","cambio","morena","vamos","pin","renovador",
        "valor","azul","une","fcn_nacion","podemos","uc"
    ]

    data_partidos = {
        p: queryset.aggregate(total=Sum(p))["total"] or 0
        for p in partidos
    }

    resumen = {
        "votos_blanco": queryset.aggregate(Sum("votos_blanco"))["votos_blanco__sum"] or 0,
        "votos_nulos": queryset.aggregate(Sum("votos_nulos"))["votos_nulos__sum"] or 0,
        "votos_invalidos": queryset.aggregate(Sum("votos_invalidos"))["votos_invalidos__sum"] or 0,
        "total": queryset.aggregate(Sum("total"))["total__sum"] or 0,
    }

    mesas = list(
        Eleccion2023.objects.filter(centro_votacion=centro)
        .exclude(mesa__isnull=True)
        .values_list("mesa", flat=True)
        .order_by("mesa")
    )

    return JsonResponse({
        "partidos": data_partidos,
        "resumen": resumen,
        "mesas": mesas,
    })



@login_required
@require_GET
def consultar_padron_local(request):
    payload, status_code = _resultado_padron_local(request.GET.get("dpi", ""))
    return JsonResponse(payload, status=status_code)




@login_required
@require_GET
def verificar_empadronamiento(request):
    """
    Endpoint estable para frontend (local/cPanel) reutilizando la lógica vigente.
    Siempre intenta responder JSON para facilitar diagnóstico en producción.
    """
    try:
        return consultar_padron_local(request)
    except Exception as exc:
        logger.exception('Error en verificar_empadronamiento')
        return JsonResponse({
            'ok': False,
            'error': f'Error interno al verificar empadronamiento: {exc.__class__.__name__}',
        }, status=500)


def _resultado_padron_local(dpi_raw):
    dpi_limpio = re.sub(r"\D", "", dpi_raw or "")

    if len(dpi_limpio) != 13:
        return ({
            "ok": False,
            "found": False,
            "empadronado": False,
            "error": "DPI inválido. Debe contener exactamente 13 dígitos.",
        }, 400)

    persona = PadronElectoral.objects.filter(identificacion=dpi_limpio).first()

    if not persona:
        return ({
            "ok": True,
            "found": False,
            "empadronado": False,
            "message": "No aparece en padrón local: podría estar vecindado en otro municipio o no empadronado. Consulte TSE.",
        }, 200)

    return ({
        "ok": True,
        "found": True,
        "empadronado": True,
        "data": {
            "dpi": persona.identificacion,
            "nombre_completo": persona.nombre,
            "comunidad": persona.comunidad,
            "departamento": persona.departamento,
            "municipio": persona.municipio,
            "edad": persona.edad,
        },
        "message": "Encontrado en padrón local",
    }, 200)

def _normalizar_columna(nombre_columna):
    texto = str(nombre_columna or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return texto


def _leer_padron_dataframe(archivo_subido):
    extension = archivo_subido.name.lower().strip()
    if extension.endswith(".csv"):
        return pd.read_csv(archivo_subido, dtype=str)
    if extension.endswith(".xlsx"):
        return pd.read_excel(archivo_subido, dtype=str)
    raise ValueError("Formato inválido. Solo se permiten archivos .xlsx o .csv.")


def _valor_nulo(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    if texto == "" or texto.lower() == "null" or texto.lower() == "nan":
        return None
    return texto


def _normalizar_identificacion(valor):
    valor = _valor_nulo(valor)
    if not valor:
        return None

    texto = str(valor).strip()
    if "e+" in texto.lower() or "e-" in texto.lower():
        try:
            texto = format(Decimal(texto), "f")
        except (InvalidOperation, ValueError):
            pass

    if texto.endswith(".0"):
        texto = texto[:-2]

    return texto.strip()


def _normalizar_edad(valor):
    valor = _valor_nulo(valor)
    if valor is None:
        return None

    try:
        numero = int(float(str(valor).strip()))
        if numero < 0:
            raise ValueError
        return numero
    except (ValueError, TypeError):
        raise ValueError(f"Edad inválida: {valor}")


def _importar_padron_desde_dataframe(df, replace=False):
    columnas_requeridas = {"nombre", "edad", "identificacion", "comunidad", "departamento", "municipio"}

    columnas_mapeadas = {_normalizar_columna(col): col for col in df.columns}
    faltantes = sorted(columnas_requeridas - set(columnas_mapeadas.keys()))
    if faltantes:
        recibidas = ", ".join(str(c) for c in df.columns)
        raise ValueError(
            "Encabezado inválido. Faltan columnas requeridas: "
            f"{', '.join(faltantes)}. Encabezados recibidos: {recibidas}"
        )

    df = df.rename(columns={original: normalizada for normalizada, original in columnas_mapeadas.items()})

    resumen = {"creados": 0, "actualizados": 0, "omitidos": 0, "errores": []}

    with transaction.atomic():
        if replace:
            PadronElectoral.objects.all().delete()

        for indice, fila in df.iterrows():
            fila_excel = indice + 2
            try:
                identificacion = _normalizar_identificacion(fila.get("identificacion"))
                if not identificacion:
                    resumen["omitidos"] += 1
                    resumen["errores"].append(f"Fila {fila_excel}: identificacion vacía")
                    continue

                nombre = _valor_nulo(fila.get("nombre"))
                if not nombre:
                    resumen["omitidos"] += 1
                    resumen["errores"].append(f"Fila {fila_excel}: nombre vacío")
                    continue

                edad = _normalizar_edad(fila.get("edad"))
                defaults = {
                    "nombre": nombre,
                    "edad": edad,
                    "comunidad": _valor_nulo(fila.get("comunidad")),
                    "departamento": _valor_nulo(fila.get("departamento")),
                    "municipio": _valor_nulo(fila.get("municipio")),
                }
                _, creado = PadronElectoral.objects.update_or_create(
                    identificacion=identificacion,
                    defaults=defaults,
                )
                if creado:
                    resumen["creados"] += 1
                else:
                    resumen["actualizados"] += 1
            except (ValueError, IntegrityError) as exc:
                resumen["errores"].append(f"Fila {fila_excel}: {exc}")

    return resumen


@login_required
def padron_cargar(request):
    if request.method == 'POST':
        if 'archivo' not in request.FILES:
            messages.error(request, 'No llegó ningún archivo en la solicitud. Verifique el formulario e intente nuevamente.')
            return redirect('afiliados:padron_cargar')

        form = PadronUploadForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data['archivo']
            replace = form.cleaned_data.get('replace', False)

            try:
                resumen = _importar_padron_desde_dataframe(_leer_padron_dataframe(archivo), replace=replace)
                mensaje = (
                    f"Importación completada. Creados: {resumen['creados']}, "
                    f"actualizados: {resumen['actualizados']}, omitidos: {resumen['omitidos']}, "
                    f"errores: {len(resumen['errores'])}."
                )
                if replace:
                    messages.warning(request, 'Modo reemplazo activo: se recargó el padrón desde el archivo enviado.')
                messages.success(request, mensaje)
                if resumen['errores']:
                    messages.warning(request, 'Primeros errores: ' + ' | '.join(resumen['errores'][:10]))
                return redirect('afiliados:padron_cargar')
            except ValueError as exc:
                logger.exception('Error de validación en importación de padrón')
                messages.error(request, str(exc))
            except IntegrityError as exc:
                logger.exception('Error de integridad en importación de padrón')
                messages.error(request, f'Error de integridad al guardar padrón: {exc}')
            except Exception as exc:
                logger.exception('Error inesperado al importar padrón electoral')
                messages.error(request, f'Error inesperado al procesar archivo: {exc.__class__.__name__}')
        else:
            if 'archivo' in form.errors:
                messages.error(request, f"Archivo inválido: {' '.join(form.errors['archivo'])}")
            else:
                messages.error(request, 'No se pudo validar el formulario de carga.')
    else:
        form = PadronUploadForm()

    return safe_render(request, 'afiliados/padron/cargar_padron.html', {'form': form})



def _construir_filtros_afiliados(request):
    modo = request.GET.get('modo', 'afiliados')
    comunidad_id = request.GET.get('comunidad_id')
    lider_id = request.GET.get('lider_id')
    sector_id = request.GET.get('sector_id')
    centro_votacion_id = request.GET.get('centro_votacion_id')
    comision_id = request.GET.get('comision_id')

    base_qs = (
        Afiliado.objects
        .select_related('comunidad__sector', 'centro_votacion', 'lider_vinculado')
        .prefetch_related(
            Prefetch(
                'comunidad__sector__centros',
                queryset=CentroVotacion.objects.only('id', 'nombre').order_by('nombre')
            )
        )
        .order_by('nombre_completo')
    )

    if comunidad_id:
        base_qs = base_qs.filter(comunidad_id=comunidad_id)
    if sector_id:
        base_qs = base_qs.filter(comunidad__sector_id=sector_id)
    if centro_votacion_id:
        base_qs = base_qs.filter(centro_votacion_id=centro_votacion_id)
    if comision_id:
        base_qs = base_qs.filter(comisiones__id=comision_id).distinct()

    resultados = base_qs
    dependientes = base_qs.none()
    referidos = base_qs.none()
    mensaje = ''
    lider_principal = None

    if modo == 'referidos':
        resultados = base_qs.filter(lider_vinculado_id=lider_id) if lider_id else base_qs.none()
    elif modo == 'lideres_dependientes':
        resultados = base_qs.filter(lider_vinculado_id=lider_id, es_lider_comunitario=True) if lider_id else base_qs.none()
    elif modo == 'dependientes_y_referidos':
        if lider_id:
            lider_principal = Afiliado.objects.filter(pk=lider_id, es_lider_comunitario=True).first()
            dependientes = base_qs.filter(lider_vinculado_id=lider_id, es_lider_comunitario=True).distinct()
            ids_dependientes = list(dependientes.values_list('id', flat=True))
            referidos = base_qs.filter(lider_vinculado_id__in=ids_dependientes).exclude(id__in=ids_dependientes).distinct()
        else:
            mensaje = 'Seleccione un líder para ver dependientes y referidos'
    else:
        if lider_id:
            resultados = resultados.filter(lider_vinculado_id=lider_id)

    return {
        'modo': modo,
        'comunidad_id': comunidad_id or '',
        'lider_id': lider_id or '',
        'sector_id': sector_id or '',
        'centro_votacion_id': centro_votacion_id or '',
        'comision_id': comision_id or '',
        'resultados': resultados,
        'dependientes': dependientes,
        'referidos': referidos,
        'mensaje': mensaje,
        'lider_principal': lider_principal,
        'total_dependientes': dependientes.count(),
        'total_referidos': referidos.count(),
    }


def _centro_votacion_desde_afiliado(afiliado):
    """
    Resuelve el centro de votación priorizando el centro enlazado al sector de la comunidad.
    Si no existe, usa el centro asignado directamente al afiliado.
    """
    comunidad = getattr(afiliado, 'comunidad', None)
    sector = getattr(comunidad, 'sector', None) if comunidad else None
    if sector:
        centros_sector = list(sector.centros.all())
        if centros_sector:
            return centros_sector[0]
    return getattr(afiliado, 'centro_votacion', None)


def _adjuntar_centro_votacion_resuelto(items):
    for item in items:
        item.centro_votacion_resuelto = _centro_votacion_desde_afiliado(item)


@login_required
def dashboard_filtros(request):
    filtros = _construir_filtros_afiliados(request)
    page_obj = filtros['resultados']
    if filtros['modo'] != 'dependientes_y_referidos':
        paginator = Paginator(filtros['resultados'], 25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

    if filtros['modo'] == 'dependientes_y_referidos':
        _adjuntar_centro_votacion_resuelto(filtros['dependientes'])
        _adjuntar_centro_votacion_resuelto(filtros['referidos'])
    else:
        _adjuntar_centro_votacion_resuelto(page_obj)

    context = {
        'comunidades': Comunidad.objects.all().order_by('nombre'),
        'lideres': Afiliado.objects.filter(es_lider_comunitario=True).order_by('nombre_completo'),
        'sectores': Sector.objects.all().order_by('nombre'),
        'centros_votacion': CentroVotacion.objects.all().order_by('nombre'),
        'comisiones': Comision.objects.all().order_by('nombre'),
        'modo': filtros['modo'],
        'comunidad_id': filtros['comunidad_id'],
        'lider_id': filtros['lider_id'],
        'sector_id': filtros['sector_id'],
        'centro_votacion_id': filtros['centro_votacion_id'],
        'comision_id': filtros['comision_id'],
        'resultados': page_obj,
        'dependientes': filtros['dependientes'],
        'referidos': filtros['referidos'],
        'mensaje': filtros['mensaje'],
        'lider_principal': filtros['lider_principal'],
        'total_dependientes': filtros['total_dependientes'],
        'total_referidos': filtros['total_referidos'],
    }

    return safe_render(request, 'afiliados/filtros.html', context)


def _rows_reporte_filtros(resultados):
    rows = []
    for afiliado in resultados:
        centro_resuelto = _centro_votacion_desde_afiliado(afiliado)
        rows.append([
            afiliado.dpi,
            afiliado.nombre_completo,
            afiliado.comunidad.nombre if afiliado.comunidad else '',
            afiliado.sector.nombre if afiliado.sector else '',
            centro_resuelto.nombre if centro_resuelto else '',
            afiliado.lider_vinculado.nombre_completo if afiliado.lider_vinculado else '',
            'Sí' if afiliado.es_lider_comunitario else 'No',
        ])
    return rows


@login_required
def exportar_filtros_excel(request):
    filtros = _construir_filtros_afiliados(request)
    rows = _rows_reporte_filtros(filtros['resultados'])

    wb = Workbook()
    headers = ['DPI', 'Nombre', 'Comunidad', 'Sector', 'Centro de Votación', 'Líder Referente', 'Es Líder']

    def _write_sheet(ws, sheet_rows):
        ws.append(headers)
        for row in sheet_rows:
            ws.append(row)
        for col_idx, col_name in enumerate(headers, start=1):
            max_len = len(col_name)
            for row_idx in range(2, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    if filtros['modo'] == 'dependientes_y_referidos':
        ws_dep = wb.active
        ws_dep.title = 'LideresDependientes'
        _write_sheet(ws_dep, _rows_reporte_filtros(filtros['dependientes']))

        ws_ref = wb.create_sheet('AfiliadosDependientes')
        _write_sheet(ws_ref, _rows_reporte_filtros(filtros['referidos']))
    else:
        ws = wb.active
        ws.title = 'Filtros'
        _write_sheet(ws, rows)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="reporte_filtros_{filtros["modo"]}_{fecha}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_filtros_pdf(request):
    filtros = _construir_filtros_afiliados(request)
    rows = _rows_reporte_filtros(filtros['resultados'])

    response = HttpResponse(content_type='application/pdf')
    fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="reporte_filtros_{filtros["modo"]}_{fecha}.pdf"'

    p = canvas.Canvas(response, pagesize=landscape(letter))
    width, height = landscape(letter)

    p.setFont('Helvetica-Bold', 14)
    p.drawString(30, height - 35, 'Reporte de Filtros')

    p.setFont('Helvetica', 9)
    filtros_txt = (
        f"Modo: {filtros['modo']} | Comunidad: {filtros['comunidad_id'] or '-'} | "
        f"Líder: {filtros['lider_id'] or '-'} | Sector: {filtros['sector_id'] or '-'} | "
        f"Centro: {filtros['centro_votacion_id'] or '-'} | Comisión: {filtros['comision_id'] or '-'}"
    )
    p.drawString(30, height - 52, filtros_txt)

    headers = ['DPI', 'Nombre', 'Comunidad', 'Sector', 'Centro', 'Líder Ref.', 'Es Líder']
    x_positions = [30, 150, 330, 450, 540, 690, 800]

    def _draw_section(title, section_rows, y_start):
        y = y_start
        p.setFont('Helvetica-Bold', 10)
        p.drawString(30, y, title)
        y -= 14

        p.setFont('Helvetica-Bold', 8)
        for i, header in enumerate(headers):
            p.drawString(x_positions[i], y, header)
        y -= 14

        p.setFont('Helvetica', 8)
        for row in section_rows:
            if y < 30:
                p.showPage()
                y = height - 35
                p.setFont('Helvetica-Bold', 8)
                for i, header in enumerate(headers):
                    p.drawString(x_positions[i], y, header)
                y -= 14
                p.setFont('Helvetica', 8)
            for i, val in enumerate(row):
                p.drawString(x_positions[i], y, str(val)[:28])
            y -= 12
        return y - 10

    if filtros['modo'] == 'dependientes_y_referidos':
        y = _draw_section('Líderes dependientes', _rows_reporte_filtros(filtros['dependientes']), height - 75)
        _draw_section('Afiliados referidos por líderes dependientes', _rows_reporte_filtros(filtros['referidos']), y)
    else:
        _draw_section('Resultados', rows, height - 75)

    p.save()
    return response


@login_required
def afiliado_lista(request):
    prefill = None
    prefill_token = (request.GET.get("prefill_token") or "").strip()
    if prefill_token:
        session_key = f"afiliacion_prefill:{prefill_token}"
        prefill = request.session.pop(session_key, None)

    afiliados = (
        Afiliado.objects
        .select_related('comunidad__sector', 'centro_votacion')
        .prefetch_related(
            Prefetch(
                'comunidad__sector__centros',
                queryset=CentroVotacion.objects.only('id', 'nombre').order_by('nombre')
            )
        )
    )
    _adjuntar_centro_votacion_resuelto(afiliados)
    form_initial = {}
    prefill_dpi = ""
    abrir_modal_prefill = False

    if prefill:
        form_initial = {
            "nombre_completo": prefill.get("nombre", ""),
            "telefono": prefill.get("telefono", ""),
            "direccion": prefill.get("direccion", ""),
            "comunidad": prefill.get("comunidad_id") or None,
        }
        prefill_dpi = prefill.get("dpi", "")
        abrir_modal_prefill = True

    form = AfiliadoForm(initial=form_initial)

    if request.method == 'POST':
        form = AfiliadoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('afiliados:afiliado_lista')

    return safe_render(request, 'afiliados/lista.html', {
        'afiliados': afiliados,
        'form': form,  # Pasamos el formulario al template
        'TSE_CONSULTA_URL': settings.TSE_CONSULTA_URL,
        'prefill_data': prefill or {},
        'prefill_dpi': prefill_dpi,
        'abrir_modal_prefill': abrir_modal_prefill,
    })


def _es_usuario_afiliacion(user):
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=['Administrador', 'Gestor', 'afiliados']).exists()


def _normalizar_dpi(dpi_raw):
    return re.sub(r"\D", "", (dpi_raw or ""))


def _coalescer_texto(*values):
    for value in values:
        txt = (value or "").strip() if isinstance(value, str) else value
        if txt:
            return txt
    return ""


SECRETARIAS_AFILIACION_FUENTES = [
    ("Organización", "coordinador_organizacion", CoordinadorOrganizacion, "fecha_creacion"),
    ("Organización", "lider_organizacion", LiderComunitarioOrganizacion, "fecha_creacion"),
    ("Organización", "responsable_organizacion", ResponsableTerritorial, "fecha_creacion"),
    ("Organización", "estructura_integrante_organizacion", EstructuraIntegrante, "fecha_registro"),
    ("Juventud", "coordinador_juventud", CoordinadorJuventud, "fecha_creacion"),
    ("Juventud", "lider_juventud", LiderJuvenil, "fecha_creacion"),
    ("Juventud", "responsable_juventud", ResponsableJuventud, "fecha_creacion"),
    ("Juventud", "estructura_integrante_juventud", EstructuraIntegranteJuventud, "fecha_registro"),
    ("Mujeres", "coordinadora_mujeres", CoordinadoraMujeres, "fecha_creacion"),
    ("Mujeres", "lider_mujeres", LiderMujeres, "fecha_creacion"),
    ("Mujeres", "responsable_mujeres", ResponsableMujeres, "fecha_creacion"),
    ("Mujeres", "estructura_integrante_mujeres", EstructuraIntegranteMujeres, "fecha_registro"),
]


SECRETARIAS_AFILIACION_MODELOS = {
    clave: modelo for _secretaria, clave, modelo, _campo_fecha in SECRETARIAS_AFILIACION_FUENTES
}


def _extraer_prefill_desde_fuente(persona):
    comunidad_obj = getattr(persona, 'comunidad', None)
    comunidad_nombre = getattr(comunidad_obj, 'nombre', '') or ''
    comunidad_id = getattr(comunidad_obj, 'id', '')
    return {
        'nombre': _coalescer_texto(getattr(persona, 'nombre_completo', ''), ''),
        'telefono': _coalescer_texto(getattr(persona, 'telefono', ''), ''),
        'direccion': _coalescer_texto(getattr(persona, 'direccion', ''), ''),
        'comunidad': comunidad_nombre,
        'comunidad_id': str(comunidad_id) if comunidad_id else '',
    }


def _fuente_por_referencia(fuente_clave, fuente_id, dpi):
    modelo = SECRETARIAS_AFILIACION_MODELOS.get(fuente_clave)
    if not modelo:
        return None
    if not fuente_id:
        return None
    try:
        persona = modelo.objects.select_related('comunidad').get(pk=fuente_id)
    except modelo.DoesNotExist:
        return None
    if _normalizar_dpi(getattr(persona, 'dpi', '')) != dpi:
        return None
    return persona


def _buscar_afiliado_por_dpi_normalizado(dpi):
    if not dpi:
        return None
    for afiliado in Afiliado.objects.only("id", "dpi"):
        if _normalizar_dpi(afiliado.dpi) == dpi:
            return afiliado
    return None


def _obtener_pendientes_afiliacion_secretarias():
    afiliados_dpi = {_normalizar_dpi(v) for v in Afiliado.objects.values_list('dpi', flat=True) if _normalizar_dpi(v)}
    pendientes_por_dpi = {}

    for secretaria, fuente_clave, modelo, campo_fecha in SECRETARIAS_AFILIACION_FUENTES:
        queryset = modelo.objects.select_related('comunidad').all()
        for persona in queryset:
            dpi = _normalizar_dpi(getattr(persona, 'dpi', ''))
            if len(dpi) != 13:
                continue
            if dpi in afiliados_dpi:
                continue

            info = pendientes_por_dpi.setdefault(
                dpi,
                {
                    'dpi': dpi,
                    'nombre_completo': _coalescer_texto(getattr(persona, 'nombre_completo', ''), f"Registro {dpi}"),
                    'comunidad': getattr(getattr(persona, 'comunidad', None), 'nombre', '') or "N/A",
                    'estado': _coalescer_texto(
                        getattr(persona, 'get_estado_display', lambda: '')(),
                        getattr(persona, 'estado', ''),
                        'N/A',
                    ),
                    'fecha_registro': getattr(persona, campo_fecha, None),
                    'secretarias': set(),
                    'fuentes': [],
                }
            )

            if not info['nombre_completo'] and getattr(persona, 'nombre_completo', None):
                info['nombre_completo'] = persona.nombre_completo
            if info['comunidad'] == "N/A":
                info['comunidad'] = getattr(getattr(persona, 'comunidad', None), 'nombre', '') or "N/A"
            if info['estado'] == 'N/A':
                info['estado'] = _coalescer_texto(
                    getattr(persona, 'get_estado_display', lambda: '')(),
                    getattr(persona, 'estado', ''),
                    'N/A',
                )
            if not info['fecha_registro'] and getattr(persona, campo_fecha, None):
                info['fecha_registro'] = getattr(persona, campo_fecha, None)

            info['secretarias'].add(secretaria)
            info['fuentes'].append({
                'secretaria': secretaria,
                'fuente_clave': fuente_clave,
                'fuente_id': persona.pk,
                'prefill': _extraer_prefill_desde_fuente(persona),
            })

    filas = []
    total_por_secretaria = defaultdict(int)

    for item in pendientes_por_dpi.values():
        secretarias = sorted(item['secretarias'])
        fuente_preferida = sorted(
            item['fuentes'],
            key=lambda fuente: (
                1 if fuente['prefill'].get('telefono') else 0,
                1 if fuente['prefill'].get('comunidad_id') else 0,
                1 if fuente['prefill'].get('nombre') else 0,
            ),
            reverse=True,
        )[0]
        prefill = fuente_preferida['prefill']

        for sec in secretarias:
            total_por_secretaria[sec] += 1
        filas.append({
            'dpi': item['dpi'],
            'nombre_completo': item['nombre_completo'],
            'secretarias': ", ".join(secretarias),
            'comunidad': item['comunidad'],
            'telefono': prefill.get('telefono') or '',
            'estado': item['estado'],
            'fecha_registro': item['fecha_registro'],
            'fuente_clave': fuente_preferida['fuente_clave'],
            'fuente_id': fuente_preferida['fuente_id'],
            'afiliar_url': (
                f"{reverse('afiliados:afiliar_desde_secretaria')}?"
                f"{urlencode({'dpi': item['dpi'], 'fuente': fuente_preferida['fuente_clave'], 'fuente_id': fuente_preferida['fuente_id'], 'nombre': prefill.get('nombre') or item['nombre_completo'], 'comunidad': prefill.get('comunidad') or (item['comunidad'] if item['comunidad'] != 'N/A' else ''), 'comunidad_id': prefill.get('comunidad_id', ''), 'telefono': prefill.get('telefono', ''), 'direccion': prefill.get('direccion', '')})}"
            ),
        })

    filas.sort(key=lambda x: x['nombre_completo'].lower())
    return filas, total_por_secretaria


@login_required
def pendientes_afiliacion_secretarias(request):
    if not _es_usuario_afiliacion(request.user):
        raise PermissionDenied("No tiene permisos para acceder a esta vista.")

    filas, total_por_secretaria = _obtener_pendientes_afiliacion_secretarias()

    context = {
        'pendientes': filas,
        'total_pendientes': len(filas),
        'total_organizacion': total_por_secretaria.get('Organización', 0),
        'total_juventud': total_por_secretaria.get('Juventud', 0),
        'total_mujeres': total_por_secretaria.get('Mujeres', 0),
        'form': AfiliadoForm(),
    }
    return safe_render(request, 'afiliados/pendientes_afiliacion_secretarias.html', context)


@login_required
def exportar_pendientes_afiliacion_secretarias_excel(request):
    if not _es_usuario_afiliacion(request.user):
        raise PermissionDenied("No tiene permisos para exportar pendientes de afiliación.")

    filas, _totales = _obtener_pendientes_afiliacion_secretarias()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes afiliación"
    headers = ["DPI", "Nombre completo", "Secretaría", "Comunidad", "Teléfono", "Estado", "Fecha registro"]
    ws.append(headers)

    for row in filas:
        fecha = row.get('fecha_registro')
        fecha_txt = fecha.strftime("%Y-%m-%d %H:%M:%S") if fecha else ""
        ws.append([
            row.get('dpi', ''),
            row.get('nombre_completo', ''),
            row.get('secretarias', ''),
            row.get('comunidad', ''),
            row.get('telefono', ''),
            row.get('estado', ''),
            fecha_txt,
        ])

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(16, len(header) + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pendientes_afiliacion_secretarias.xlsx"'
    wb.save(response)
    return response


@login_required
def prefill_afiliacion_desde_secretaria(request):
    if not _es_usuario_afiliacion(request.user):
        raise PermissionDenied("No tiene permisos para ejecutar esta acción.")

    dpi = _normalizar_dpi(request.GET.get('dpi', ''))
    fuente_clave = (request.GET.get('fuente') or '').strip()
    fuente_id = (request.GET.get('fuente_id') or '').strip()

    if len(dpi) != 13:
        return JsonResponse({'ok': False, 'message': 'DPI inválido.'}, status=400)

    afiliado_existente = _buscar_afiliado_por_dpi_normalizado(dpi)
    if afiliado_existente:
        return JsonResponse({
            'ok': False,
            'exists': True,
            'message': f'El DPI {dpi} ya está afiliado.',
            'detalle_url': reverse('afiliados:afiliado_detalle', args=[afiliado_existente.pk]),
        }, status=409)

    persona_fuente = _fuente_por_referencia(fuente_clave, fuente_id, dpi) if fuente_clave else None
    if not persona_fuente:
        return JsonResponse({
            'ok': False,
            'message': 'No se encontró un registro fuente válido para ese DPI y secretaría de origen.',
        }, status=404)

    prefill = _extraer_prefill_desde_fuente(persona_fuente)
    nombre = prefill.get('nombre') or ''
    if not nombre:
        return JsonResponse({
            'ok': False,
            'message': 'No se puede afiliar: el registro fuente no tiene nombre completo.',
        }, status=422)

    return JsonResponse({
        'ok': True,
        'prefill': {
            'dpi': dpi,
            'nombre': nombre,
            'comunidad': prefill.get('comunidad') or '',
            'comunidad_id': prefill.get('comunidad_id') or '',
            'telefono': prefill.get('telefono') or '',
            'direccion': prefill.get('direccion') or '',
            'fuente': fuente_clave,
            'fuente_id': str(fuente_id),
        }
    })


@login_required
def modal_afiliar_desde_secretaria_form(request):
    if not _es_usuario_afiliacion(request.user):
        raise PermissionDenied("No tiene permisos para ejecutar esta acción.")

    dpi = _normalizar_dpi(request.GET.get('dpi', ''))
    fuente_clave = (request.GET.get('fuente') or '').strip()
    fuente_id = (request.GET.get('fuente_id') or '').strip()

    if len(dpi) != 13:
        return JsonResponse({'ok': False, 'message': 'DPI inválido.'}, status=400)

    afiliado_existente = _buscar_afiliado_por_dpi_normalizado(dpi)
    if afiliado_existente:
        return JsonResponse({
            'ok': False,
            'exists': True,
            'message': f'El DPI {dpi} ya está afiliado.',
            'detalle_url': reverse('afiliados:afiliado_detalle', args=[afiliado_existente.pk]),
        }, status=409)

    persona_fuente = _fuente_por_referencia(fuente_clave, fuente_id, dpi) if fuente_clave else None
    if not persona_fuente:
        return JsonResponse({
            'ok': False,
            'message': 'No se encontró un registro fuente válido para ese DPI y secretaría de origen.',
        }, status=404)

    prefill = _extraer_prefill_desde_fuente(persona_fuente)
    form = AfiliadoForm(initial={
        'dpi': dpi,
        'nombre_completo': prefill.get('nombre', ''),
        'telefono': prefill.get('telefono', ''),
        'direccion': prefill.get('direccion', ''),
        'comunidad': prefill.get('comunidad_id') or None,
        'empadronado': True,
    })
    form_html = render_to_string(
        'afiliados/partials/modal_afiliar_desde_secretaria_form.html',
        {
            'form': form,
            'dpi': dpi,
            'fuente': fuente_clave,
            'fuente_id': fuente_id,
        },
        request=request,
    )
    return JsonResponse({'ok': True, 'form_html': form_html})


@login_required
def modal_afiliar_desde_secretaria_guardar(request):
    if not _es_usuario_afiliacion(request.user):
        raise PermissionDenied("No tiene permisos para ejecutar esta acción.")

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'message': 'Método no permitido.'}, status=405)

    dpi = _normalizar_dpi(request.POST.get('dpi', ''))
    fuente_clave = (request.POST.get('fuente') or '').strip()
    fuente_id = (request.POST.get('fuente_id') or '').strip()
    persona_fuente = _fuente_por_referencia(fuente_clave, fuente_id, dpi) if fuente_clave else None
    if not persona_fuente:
        return JsonResponse({
            'ok': False,
            'message': 'No se encontró un registro fuente válido para guardar la afiliación.',
        }, status=404)

    afiliado_existente = _buscar_afiliado_por_dpi_normalizado(dpi)
    if afiliado_existente:
        return JsonResponse({
            'ok': False,
            'exists': True,
            'message': f'El DPI {dpi} ya está afiliado.',
            'detalle_url': reverse('afiliados:afiliado_detalle', args=[afiliado_existente.pk]),
        }, status=409)

    form = AfiliadoForm(request.POST)
    if form.is_valid():
        afiliado = form.save()
        return JsonResponse({'ok': True, 'message': f"Afiliado '{afiliado.nombre_completo}' guardado con éxito."})

    form_html = render_to_string(
        'afiliados/partials/modal_afiliar_desde_secretaria_form.html',
        {
            'form': form,
            'dpi': request.POST.get('dpi', ''),
            'fuente': fuente_clave,
            'fuente_id': fuente_id,
        },
        request=request,
    )
    return JsonResponse({'ok': False, 'message': 'Error de validación.', 'form_html': form_html}, status=400)


@login_required
def afiliar_desde_secretaria(request):
    if not _es_usuario_afiliacion(request.user):
        raise PermissionDenied("No tiene permisos para ejecutar esta acción.")

    dpi = _normalizar_dpi(request.GET.get('dpi', ''))
    fuente_clave = (request.GET.get('fuente') or '').strip()
    fuente_id = (request.GET.get('fuente_id') or '').strip()

    if len(dpi) != 13:
        messages.error(request, "No se puede afiliar: DPI inválido.")
        return redirect('afiliados:pendientes_afiliacion_secretarias')

    afiliado_existente = _buscar_afiliado_por_dpi_normalizado(dpi)
    if afiliado_existente:
        messages.info(request, f"El DPI {dpi} ya está afiliado. Se abrió su detalle.")
        return redirect('afiliados:afiliado_detalle', pk=afiliado_existente.pk)

    persona_fuente = _fuente_por_referencia(fuente_clave, fuente_id, dpi) if fuente_clave else None
    if not persona_fuente:
        messages.error(
            request,
            "No se encontró un registro fuente válido para ese DPI y secretaría de origen. "
            "Actualice la lista e intente nuevamente."
        )
        return redirect('afiliados:pendientes_afiliacion_secretarias')

    prefill = _extraer_prefill_desde_fuente(persona_fuente)
    nombre = prefill.get('nombre') or (request.GET.get('nombre') or '').strip()
    comunidad = prefill.get('comunidad') or (request.GET.get('comunidad') or '').strip()
    comunidad_id = prefill.get('comunidad_id') or (request.GET.get('comunidad_id') or '').strip()
    telefono = prefill.get('telefono') or (request.GET.get('telefono') or '').strip()
    direccion = prefill.get('direccion') or (request.GET.get('direccion') or '').strip()

    if not nombre:
        messages.error(request, "No se puede abrir la afiliación: el registro fuente no tiene nombre completo.")
        return redirect('afiliados:pendientes_afiliacion_secretarias')

    prefill_token = uuid.uuid4().hex
    request.session[f"afiliacion_prefill:{prefill_token}"] = {
        "dpi": dpi,
        "nombre": nombre,
        "comunidad": comunidad,
        "comunidad_id": comunidad_id,
        "telefono": telefono,
        "direccion": direccion,
        "fuente": fuente_clave,
        "fuente_id": str(fuente_id),
    }
    query = urlencode({'prefill_token': prefill_token})
    messages.info(request, "Se abrió el formulario de afiliación con datos precargados.")
    return redirect(f"{reverse('afiliados:afiliado_lista')}?{query}")


@login_required
@csrf_exempt # Solo si estás usando AJAX sin manejar CSRF token en el header correctamente
def afiliado_nuevo(request):
    institucion = Institucion.objects.first()
    
    if request.method == 'POST':
        # 1. Usar el AfiliadoForm actualizado
        form = AfiliadoForm(request.POST) 
        
        # Si la solicitud es AJAX (desde el modal)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            
            if form.is_valid():
                afiliado = form.save()
                
                # 🛠️ Verificación Adicional (Opcional pero Recomendada)
                # Si el afiliado es guardado con lider_vinculado, nos aseguramos que ese 'lider_vinculado'
                # realmente exista y sea un líder (aunque el form ya lo limite, esto es una capa de seguridad).
                if afiliado.lider_vinculado and not afiliado.lider_vinculado.es_lider_comunitario:
                    # Esto no debería pasar si limit_choices_to funciona, pero es buena práctica.
                    # Podrías registrar un error o limpiar el campo si es necesario.
                    pass 
                
                # Preparamos la respuesta para AJAX
                response_data = {
                    'exito': True,
                    'mensaje': f"Afiliado '{afiliado.nombre_completo}' guardado con éxito.",
                    'afiliado': {
                        'id': afiliado.id,
                        'nombre_completo': afiliado.nombre_completo,
                        # ... puedes añadir más campos para actualizar la tabla si es necesario
                    }
                }
                return JsonResponse(response_data, status=200)
            
            else:
                # Manejar errores de validación para AJAX
                error_html = render_to_string('afiliados/partials/form_errors.html', {'form': form})
                return JsonResponse({
                    'exito': False,
                    'mensaje': "Error de validación. Revise los campos.",
                    'errores': form.errors.as_json()
                }, status=400)
                
        # Si la solicitud es POST normal (no AJAX)
        else:
            if form.is_valid():
                form.save()
                messages.success(request, 'Afiliado guardado con éxito.')
                return redirect('afiliados:afiliados_lista') # Asegúrate que el nombre de la URL sea correcto
            else:
                messages.error(request, 'Hubo un error al guardar el afiliado.')
                # Si no es AJAX, podrías renderizar la página completa con el formulario con errores
                
    else:
        # Petición GET: Creamos el formulario vacío
        form = AfiliadoForm()

    # Renderizamos la lista principal o el contexto para el modal (si no es AJAX)
    # Asumiendo que esta vista es la que renderiza la lista principal y el modal
    context = {
        'afiliados': Afiliado.objects.all(),
        'form': form, # El formulario vacío o con errores
        'institucion': institucion
    }
    return safe_render(request, 'afiliados/afiliados_lista.html', context)

# -------------------------------
# EDITAR AFILIADO
# -------------------------------
@login_required
def afiliado_editar(request, pk):
    afiliado = get_object_or_404(Afiliado, pk=pk)
    if request.method == 'POST':
        form = AfiliadoForm(request.POST, instance=afiliado)
        if form.is_valid():
            form.save()
            messages.success(request, "Afiliado actualizado correctamente.")
            return redirect('afiliados:afiliado_lista')
    else:
        form = AfiliadoForm(instance=afiliado)
    return safe_render(request, 'afiliados/form.html', {'form': form, 'afiliado': afiliado})

@login_required
def lider_editar(request, pk):
    afiliado = get_object_or_404(Afiliado, pk=pk)
    
    # OPCIONAL: Si el objeto no es un líder, redirigir a la edición general
    # if not afiliado.es_lider_comunitario:
    #     return redirect('afiliados:afiliado_editar', pk=pk)

    if request.method == 'POST':
        form = AfiliadoForm(request.POST, instance=afiliado)
        if form.is_valid():
            form.save()
            messages.success(request, f"Líder {afiliado.nombre_completo} actualizado correctamente.")
            return redirect('afiliados:lideres_lista') # 🚀 Redirección CLAVE
    else:
        form = AfiliadoForm(instance=afiliado)
    
    # Cambiamos el contexto para indicar que es una edición de líder (opcional)
    context = {
        'form': form, 
        'afiliado': afiliado,
        'es_lider_edicion': True # Usar esta variable en el template para cambiar el título
    }
    return safe_render(request, 'afiliados/form.html', context)

@login_required
def lideres_lista(request):
    """Muestra solo a los afiliados que son Líderes Comunitarios."""
    
    # Filtramos la lista de Afiliados donde es_lider_comunitario es True
    lideres = Afiliado.objects.filter(es_lider_comunitario=True).select_related('comunidad', 'centro_votacion').order_by('nombre_completo')
    
    # Necesitas el formulario para el modal 'Nuevo Afiliado'
    form = AfiliadoForm()
    institucion = Institucion.objects.first()

    context = {
        'lideres': lideres, # Renombramos la variable a 'lideres' para evitar confusiones
        'form': form,
        'institucion': institucion,
        'TSE_CONSULTA_URL': settings.TSE_CONSULTA_URL,
    }
    
    return safe_render(request, 'afiliados/lideres_lista.html', context)

# -------------------------------
# ELIMINAR AFILIADO
# -------------------------------
@login_required
def afiliado_eliminar(request, pk):
    afiliado = get_object_or_404(Afiliado, pk=pk)

    if request.method != 'POST':
        return JsonResponse({
            'ok': False,
            'exito': False,
            'mensaje': 'Método no permitido. Use POST para eliminar.'
        }, status=405)

    referidos_count = afiliado.referidos.count()
    if referidos_count > 0:
        return JsonResponse({
            'ok': False,
            'exito': False,
            'mensaje': f'No se puede eliminar porque tiene {referidos_count} referido(s) asociado(s).',
            'referidos': referidos_count,
            'puede_desactivar': False,
        }, status=400)

    try:
        afiliado.delete()
    except ProtectedError as e:
        relacionados = len(getattr(e, 'protected_objects', []))
        return JsonResponse({
            'ok': False,
            'exito': False,
            'mensaje': f'No se puede eliminar porque tiene {relacionados} dependencia(s) protegida(s).'
        }, status=400)
    except IntegrityError:
        logger.exception('IntegrityError al eliminar afiliado id=%s', pk)
        return JsonResponse({
            'ok': False,
            'exito': False,
            'mensaje': 'No se puede eliminar el afiliado por restricciones de integridad en la base de datos.'
        }, status=400)
    except PermissionDenied:
        return JsonResponse({
            'ok': False,
            'exito': False,
            'mensaje': 'No tiene permisos para eliminar este afiliado.'
        }, status=403)
    except Exception:
        logger.exception('Error inesperado al eliminar afiliado id=%s', pk)
        return JsonResponse({
            'ok': False,
            'exito': False,
            'mensaje': 'Error inesperado al intentar eliminar el afiliado.'
        }, status=500)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'ok': True, 'exito': True, 'mensaje': 'Afiliado eliminado exitosamente.'})
    return redirect('afiliados:afiliado_lista')

# En tu archivo afiliados_app/views.py

from .form import AfiliadoForm # Asegúrate de importar el formulario

@login_required
def afiliado_detalle(request, pk):
    afiliado = get_object_or_404(Afiliado, pk=pk)
    institucion = Institucion.objects.first()
    
    referidos = afiliado.referidos.all() # Ya no necesitamos el if, si no es líder estará vacío
    
    # 🆕 IMPORTANTE: Creamos el formulario para el modal
    form = AfiliadoForm() 
    
    context = {
        'afiliado': afiliado,
        'institucion': institucion,
        'referidos': referidos,
        'form': form, # <-- Pasamos el formulario al template
        'TSE_CONSULTA_URL': settings.TSE_CONSULTA_URL,
    }
    
    return safe_render(request, 'afiliados/afiliado_detalle.html', context)


# -----------------------------------------
# CRUD - COMUNIDAD
# -----------------------------------------

@login_required
def comunidad_lista(request):
    comunidades = Comunidad.objects.all().order_by('nombre')
    form = ComunidadForm()  # Nuevo formulario vacío 

    return safe_render(request, 'afiliados/comunidad_lista.html', {
        'form': form,
        'comunidades': comunidades
    })


@login_required
def comunidad_nueva(request):
    if request.method == 'POST':
        form = ComunidadForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Comunidad creada correctamente.")
            return redirect('afiliados:comunidad_lista')
    else:
        form = ComunidadForm()

    comunidades = Comunidad.objects.all().order_by('nombre')

    return safe_render(request, 'afiliados/comunidad_lista.html', {
        'form': form,
        'comunidades': comunidades
    })


@login_required
def comunidad_editar(request, pk):
    comunidad = get_object_or_404(Comunidad, pk=pk)

    if request.method == "POST":
        form = ComunidadForm(request.POST, instance=comunidad)
        if form.is_valid():
            form.save()
            messages.success(request, "Comunidad actualizada correctamente.")
            return redirect('afiliados:comunidad_lista')
    else:
        form = ComunidadForm(instance=comunidad)

    comunidades = Comunidad.objects.all().order_by('nombre')

    return safe_render(request, 'afiliados/comunidad_lista.html', {
        'form': form,
        'comunidad': comunidad,
        'comunidades': comunidades,
    })


@login_required
def comunidad_eliminar(request, pk):
    comunidad = get_object_or_404(Comunidad, pk=pk)

    if request.method == 'POST':
        comunidad.delete()
        messages.success(request, "Comunidad eliminada.")
        return redirect('afiliados:comunidad_lista')

    # ❌ Ya NO renderizamos nada
    return redirect('afiliados:comunidad_lista')

# ------------------------
# CRUD DE SECTOR
# ------------------------

@login_required
def sector_nueva(request):
    if request.method == 'POST':
        form = SectorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Sector creado correctamente.")
            return redirect('afiliados:sector_lista')
    else:
        form = SectorForm()

    sectores = Sector.objects.all().order_by('nombre')

    return safe_render(request, 'afiliados/sector_lista.html', {
        'form': form,
        'sectores': sectores
    })


@login_required
def sector_editar(request, pk):
    sector = get_object_or_404(Sector, pk=pk)

    if request.method == 'POST':
        form = SectorForm(request.POST, instance=sector)
        if form.is_valid():
            form.save()
            messages.success(request, "Sector actualizado correctamente.")
            return redirect('afiliados:sector_lista')
    else:
        form = SectorForm(instance=sector)

    sectores = Sector.objects.all().order_by('nombre')

    return safe_render(request, 'afiliados/sector_lista.html', {
        'form': form,
        'sector': sector,
        'sectores': sectores
    })


@login_required
def sector_eliminar(request, pk):
    sector = get_object_or_404(Sector, pk=pk)

    if request.method == 'POST':
        sector.delete()
        messages.success(request, "Sector eliminado.")
        return redirect('afiliados:sector_lista')

    return redirect('afiliados:sector_lista')



# -----------------------------------------
# CRUD - CENTRO DE VOTACION
# -----------------------------------------
# -----------------------------------------
# CRUD - CENTRO DE VOTACIÓN (MisMO estilo)
# -----------------------------------------

@login_required
def centro_lista(request):
    centros = CentroVotacion.objects.all().order_by('nombre')
    form = CentroVotacionForm()  # formulario vacío para creación

    return safe_render(request, 'afiliados/centro_lista.html', {
        'form': form,
        'centros': centros
    })


@login_required
def centro_nuevo(request):
    if request.method == 'POST':
        form = CentroVotacionForm(request.POST)
        if form.is_valid():
            centro = form.save()
           
            messages.success(request, "Centro de Votación creado correctamente.")
            return redirect('afiliados:centro_lista')
    else:
        form = CentroVotacionForm()

    centros = CentroVotacion.objects.all().order_by('nombre')

    return safe_render(request, 'afiliados/centro_lista.html', {
        'form': form,
        'centros': centros
    })


@login_required
def centro_editar(request, pk):
    centro = get_object_or_404(CentroVotacion, pk=pk)

    if request.method == "POST":
        form = CentroVotacionForm(request.POST, instance=centro)
        if form.is_valid():
            centro = form.save()
           
            messages.success(request, "Centro de Votación actualizado correctamente.")
            return redirect('afiliados:centro_lista')
    else:
        form = CentroVotacionForm(instance=centro)

    centros = CentroVotacion.objects.all().order_by('nombre')

    return safe_render(request, 'afiliados/centro_lista.html', {
        'form': form,
        'centro': centro,
        'centros': centros,
    })


@login_required
def centro_eliminar(request, pk):
    centro = get_object_or_404(CentroVotacion, pk=pk)

    if request.method == 'POST':
        centro.delete()
        messages.success(request, "Centro de votación eliminado.")
        return redirect('afiliados:centro_lista')

    return redirect('afiliados:centro_lista')

# -----------------------------------------
# CRUD - COMISION
# -----------------------------------------
# -----------------------------------------
# CRUD - COMISIÓN (Mismo estilo)
# -----------------------------------------

@login_required
def comision_lista(request):
    comisiones = Comision.objects.all().order_by('nombre')
    form = ComisionForm()

    return safe_render(request, 'afiliados/comision_lista.html', {
        'form': form,
        'comisiones': comisiones,
    })


@login_required
def comision_nueva(request):
    if request.method == 'POST':
        form = ComisionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Comisión creada correctamente.")
            return redirect('afiliados:comision_lista')

    comisiones = Comision.objects.all().order_by('nombre')
    form = ComisionForm()

    return safe_render(request, 'afiliados/comision_lista.html', {
        'form': form,
        'comisiones': comisiones,
    })


@login_required
def comision_editar(request, pk):
    comision = get_object_or_404(Comision, pk=pk)

    if request.method == 'POST':
        form = ComisionForm(request.POST, instance=comision)
        if form.is_valid():
            form.save()
            messages.success(request, "Comisión actualizada correctamente.")
            return redirect('afiliados:comision_lista')
    else:
        form = ComisionForm(instance=comision)

    comisiones = Comision.objects.all().order_by('nombre')

    return safe_render(request, 'afiliados/comision_lista.html', {
        'form': form,
        'comision': comision,
        'comisiones': comisiones,
    })


@login_required
def comision_eliminar(request, pk):
    comision = get_object_or_404(Comision, pk=pk)

    if request.method == 'POST':
        comision.delete()
        messages.success(request, "Comisión eliminada correctamente.")
        return redirect('afiliados:comision_lista')

    return redirect('afiliados:comision_lista')


def _es_usuario_organizacion(user):
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=['Organizacion', 'Administrador']).exists()


def _require_organizacion(request):
    if not _es_usuario_organizacion(request.user):
        return HttpResponseForbidden("No tiene permisos para acceder a Secretaría de Organización.")
    return None


def _afiliado_por_dpi(dpi_raw):
    dpi_limpio = re.sub(r"\D", "", dpi_raw or "")
    if len(dpi_limpio) != 13:
        return None, dpi_limpio, "DPI inválido. Debe contener exactamente 13 dígitos."
    return Afiliado.objects.filter(dpi=dpi_limpio).first(), dpi_limpio, None


@login_required
def dashboard_organizacion(request):
    denied = _require_organizacion(request)
    if denied:
        return denied

    registros = OrganizacionIntegrante.objects.select_related('afiliado')
    context = {
        'total_integrantes': registros.count(),
        'total_pendientes': registros.filter(estado=OrganizacionIntegrante.Estado.PENDIENTE).count(),
        'total_revisados': registros.filter(estado=OrganizacionIntegrante.Estado.REVISADO).count(),
        'total_aprobados': registros.filter(estado=OrganizacionIntegrante.Estado.APROBADO).count(),
        'ultimos_registros': registros.order_by('-fecha_creacion')[:10],
    }
    return safe_render(request, 'afiliados/organizacion/dashboard.html', context)


@login_required
def lista_integrantes_organizacion(request):
    denied = _require_organizacion(request)
    if denied:
        return denied

    integrantes = OrganizacionIntegrante.objects.select_related('afiliado', 'usuario_registro')
    return safe_render(request, 'afiliados/organizacion/lista.html', {'integrantes': integrantes})


@login_required
def crear_integrante_organizacion(request):
    denied = _require_organizacion(request)
    if denied:
        return denied

    form = OrganizacionIntegranteForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])

        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif OrganizacionIntegrante.objects.filter(afiliado=afiliado).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de Organización.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_registro = request.user
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante agregado a Secretaría de Organización correctamente.')
            return redirect('afiliados:lista_integrantes_organizacion')

    return safe_render(request, 'afiliados/organizacion/form.html', {'form': form, 'es_edicion': False})


@login_required
def detalle_integrante_organizacion(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied

    integrante = get_object_or_404(
        OrganizacionIntegrante.objects.select_related('afiliado', 'usuario_registro', 'usuario_modificacion'),
        pk=pk,
    )
    return safe_render(request, 'afiliados/organizacion/detalle.html', {'integrante': integrante})


@login_required
def editar_integrante_organizacion(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied

    integrante = get_object_or_404(OrganizacionIntegrante, pk=pk)
    form = OrganizacionIntegranteForm(request.POST or None, instance=integrante)
    form.fields['dpi'].initial = integrante.afiliado.dpi

    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])
        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif OrganizacionIntegrante.objects.filter(afiliado=afiliado).exclude(pk=integrante.pk).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de Organización.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante de Secretaría de Organización actualizado correctamente.')
            return redirect('afiliados:detalle_integrante_organizacion', pk=integrante.pk)

    return safe_render(request, 'afiliados/organizacion/form.html', {'form': form, 'es_edicion': True, 'integrante': integrante})


@login_required
@require_GET
def buscar_por_dpi_organizacion(request):
    denied = _require_organizacion(request)
    if denied:
        return denied

    afiliado, dpi_limpio, error_dpi = _afiliado_por_dpi(request.GET.get('dpi', ''))
    if error_dpi:
        return JsonResponse({'ok': False, 'error': error_dpi}, status=400)

    if not afiliado:
        return JsonResponse({'ok': True, 'exists': False, 'message': 'No existe afiliado con ese DPI.'})

    ya_registrado = OrganizacionIntegrante.objects.filter(afiliado=afiliado).exists()
    return JsonResponse({
        'ok': True,
        'exists': True,
        'already_registered': ya_registrado,
        'data': {
            'dpi': dpi_limpio,
            'nombre_completo': afiliado.nombre_completo,
            'telefono': afiliado.telefono,
            'direccion': afiliado.direccion,
            'comunidad': afiliado.comunidad.nombre if afiliado.comunidad else None,
            'empadronado': afiliado.empadronado,
        }
    })


@login_required
@require_GET
def verificar_empadronamiento_organizacion(request):
    denied = _require_organizacion(request)
    if denied:
        return denied
    payload, status_code = _resultado_padron_local(request.GET.get("dpi", ""))
    return JsonResponse(payload, status=status_code)


@login_required
@require_POST
def cambiar_estado_organizacion(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied

    integrante = get_object_or_404(OrganizacionIntegrante, pk=pk)
    nuevo_estado = request.POST.get('estado', '').strip().upper()
    estados_validos = {choice[0] for choice in OrganizacionIntegrante.Estado.choices}

    if nuevo_estado not in estados_validos:
        messages.error(request, 'Estado inválido.')
        return redirect('afiliados:detalle_integrante_organizacion', pk=pk)

    integrante.estado = nuevo_estado
    integrante.usuario_modificacion = request.user
    integrante.save(update_fields=['estado', 'usuario_modificacion', 'fecha_actualizacion'])
    messages.success(request, 'Estado actualizado correctamente.')
    return redirect('afiliados:detalle_integrante_organizacion', pk=pk)


@login_required
def panorama_municipal_organizacion(request):
    denied = _require_organizacion(request)
    if denied:
        return denied

    comunidad_qs = Comunidad.objects.select_related('sector')
    sector_qs = Sector.objects.all()
    centro_qs = CentroVotacion.objects.all()
    coordinadores = CoordinadorOrganizacion.objects.all()
    lideres = LiderComunitarioOrganizacion.objects.all()
    estructuras = EstructuraOrganizativa.objects.all()
    reuniones = ReunionTerritorial.objects.all()
    incidencias = IncidenciaTerritorial.objects.filter(estado=EstadoRegistro.ACTIVO)
    responsables = ResponsableTerritorial.objects.filter(estado=EstadoRegistro.ACTIVO)

    comunidad_ids_cubiertas = set(
        CoordinadorOrganizacion.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)
    ) | set(
        LiderComunitarioOrganizacion.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)
    ) | set(
        EstructuraOrganizativa.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)
    ) | set(
        ResponsableTerritorial.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)
    )

    sector_ids_cubiertos = set(
        CoordinadorOrganizacion.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True)
    ) | set(
        EstructuraOrganizativa.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True)
    ) | set(
        ResponsableTerritorial.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True)
    )

    centro_ids_cubiertos = set(
        CoordinadorOrganizacion.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True)
    ) | set(
        EstructuraOrganizativa.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True)
    ) | set(
        ResponsableTerritorial.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True)
    )

    crecimiento, crecimiento_series = _construir_series_crecimiento({
        'coordinadores': coordinadores,
        'lideres': lideres,
        'estructuras': estructuras,
        'reuniones': reuniones,
    })

    resumen_estructuras_por_comunidad = comunidad_qs.annotate(
        estructuras_total=Count('estructuras_organizacion', distinct=True),
    ).order_by('nombre')

    resumen_miembros_por_estructura = EstructuraOrganizativa.objects.select_related('comunidad').annotate(
        miembros_total=Count('integrantes', distinct=True),
    ).order_by('nombre')

    context = {
        'total_comunidades': comunidad_qs.count(),
        'comunidades_cubiertas': len(comunidad_ids_cubiertas),
        'comunidades_no_cubiertas': max(comunidad_qs.count() - len(comunidad_ids_cubiertas), 0),
        'total_sectores': sector_qs.count(),
        'sectores_con_responsable': responsables.exclude(sector__isnull=True).values('sector_id').distinct().count(),
        'sectores_sin_responsable': max(sector_qs.count() - responsables.exclude(sector__isnull=True).values('sector_id').distinct().count(), 0),
        'total_centros': centro_qs.count(),
        'centros_con_cobertura': len(centro_ids_cubiertos),
        'centros_sin_cobertura': max(centro_qs.count() - len(centro_ids_cubiertos), 0),
        'coordinadores_activos': coordinadores.filter(estado=EstadoRegistro.ACTIVO).count(),
        'coordinadores_inactivos': coordinadores.filter(estado=EstadoRegistro.INACTIVO).count(),
        'total_lideres': lideres.count(),
        'estructuras_activas': estructuras.filter(estado=EstadoRegistro.ACTIVO).count(),
        'reuniones_mes': reuniones.filter(fecha__month=timezone.now().month, fecha__year=timezone.now().year).count(),
        'incidencias_abiertas': incidencias.count(),
        'resumen_comunidad': comunidad_qs.annotate(
            total_sectores=Count('sector', distinct=True),
            reuniones_total=Count('reuniones_organizacion', distinct=True),
            incidencias_abiertas=Count('incidencias_organizacion', filter=Q(incidencias_organizacion__estado=EstadoRegistro.ACTIVO), distinct=True),
        ),
        'resumen_sector': sector_qs.annotate(
            lideres_total=Count('lideres_organizacion', distinct=True),
            estructuras_total=Count('estructuras_organizacion', distinct=True),
            incidencias_total=Count('incidencias_organizacion', distinct=True),
        ),
        'resumen_centro': centro_qs.annotate(
            incidencias_total=Count('incidencias_organizacion', distinct=True),
            reuniones_total=Count('reuniones_organizacion', distinct=True),
        ),
        'resumen_estructuras_por_comunidad': resumen_estructuras_por_comunidad,
        'resumen_miembros_por_estructura': resumen_miembros_por_estructura,
        'crecimiento': json.dumps(crecimiento, cls=DjangoJSONEncoder),
        'charts_data': json.dumps({
            'cobertura': {
                'comunidades': [len(comunidad_ids_cubiertas), max(comunidad_qs.count() - len(comunidad_ids_cubiertas), 0)],
                'sectores': [len(sector_ids_cubiertos), max(sector_qs.count() - len(sector_ids_cubiertos), 0)],
                'centros': [len(centro_ids_cubiertos), max(centro_qs.count() - len(centro_ids_cubiertos), 0)],
            },
            'estructura': {
                'coordinadores': [coordinadores.filter(estado=EstadoRegistro.ACTIVO).count(), coordinadores.filter(estado=EstadoRegistro.INACTIVO).count()],
                'lideres_total': lideres.count(),
                'estructuras_activas': estructuras.filter(estado=EstadoRegistro.ACTIVO).count(),
            },
            'crecimiento': crecimiento_series,
        }, cls=DjangoJSONEncoder),
        'sector_ids_cubiertos': sector_ids_cubiertos,
        'centro_ids_cubiertos': centro_ids_cubiertos,
        'comunidad_ids_cubiertas': comunidad_ids_cubiertas,
    }
    return safe_render(request, 'afiliados/organizacion/panorama.html', context)


def _org_crud(request, model, form_class, template, list_name, create_name, edit_name, pk=None, extra_context=None):
    denied = _require_organizacion(request)
    if denied:
        return denied
    instance = get_object_or_404(model, pk=pk) if pk else None
    form = form_class(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if not getattr(obj, 'usuario_creador_id', None):
            obj.usuario_creador = request.user
        obj.usuario_modificador = request.user
        obj.save()
        messages.success(request, 'Registro guardado correctamente.')
        return redirect(list_name)
    descripciones = {
        'coordinadores': 'Aquí registras coordinadores territoriales por comunidad. El sistema deriva sector y centro automáticamente.',
        'lideres': 'Aquí registras líderes comunitarios vinculados a comunidad, con control de estado y seguimiento.',
        'estructuras': 'Aquí registras comités, equipos, células o comisiones comunitarias activas del municipio.',
        'responsables': 'Aquí asignas responsables por comunidad para mantener cobertura y orden territorial.',
        'reuniones': 'Aquí registras reuniones territoriales, acuerdos y estado de seguimiento.',
        'incidencias': 'Aquí reportas incidencias territoriales para su gestión y resolución oportuna.',
    }
    detail_names = {
        'coordinadores': 'afiliados:detalle_coordinador',
        'lideres': 'afiliados:detalle_lider',
        'estructuras': 'afiliados:detalle_estructura',
        'responsables': 'afiliados:detalle_responsable',
        'reuniones': 'afiliados:detalle_reunion',
        'incidencias': 'afiliados:detalle_incidencia',
    }

    items_qs = model.objects.all().order_by('-id')
    if template == 'reuniones':
        items_qs = items_qs.select_related('comunidad', 'sector', 'centro_votacion')

    context = {
        'form': form,
        'items': items_qs[:200],
        'entity_label': template,
        'create_name': create_name,
        'edit_name': edit_name,
        'detail_name': detail_names.get(template),
        'descripcion_vista': descripciones.get(template, ''),
        'comunidad_lookup_url': reverse('afiliados:organizacion_comunidad_lookup'),
        'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_organizacion'),
    }

    if template == 'reuniones':
        responsables_modal = []

        for responsable in ResponsableTerritorial.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({
                'tipo': 'RESPONSABLE',
                'tipo_label': 'Responsable',
                'id': responsable.id,
                'nombre': responsable.nombre_completo,
                'dpi': responsable.dpi,
                'comunidad': responsable.comunidad.nombre if responsable.comunidad else '',
            })

        for coordinador in CoordinadorOrganizacion.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({
                'tipo': 'COORDINADOR',
                'tipo_label': 'Coordinador',
                'id': coordinador.id,
                'nombre': coordinador.nombre_completo,
                'dpi': coordinador.dpi,
                'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else '',
            })

        for lider in LiderComunitarioOrganizacion.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({
                'tipo': 'LIDER',
                'tipo_label': 'Líder',
                'id': lider.id,
                'nombre': lider.nombre_completo,
                'dpi': lider.dpi,
                'comunidad': lider.comunidad.nombre if lider.comunidad else '',
            })

        context['responsables_modal'] = responsables_modal
    if template == 'estructuras':
        coordinadores_modal = []
        for coordinador in CoordinadorOrganizacion.objects.select_related('comunidad').order_by('nombre_completo'):
            coordinadores_modal.append({
                'id': coordinador.id,
                'nombre': coordinador.nombre_completo,
                'dpi': coordinador.dpi,
                'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else '',
                'estado': coordinador.get_estado_display(),
            })

        coordinador_responsable_display = ''
        coordinador_id = form['coordinador_responsable'].value()
        if coordinador_id:
            coordinador_actual = CoordinadorOrganizacion.objects.filter(pk=coordinador_id).first()
            if coordinador_actual:
                coordinador_responsable_display = coordinador_actual.nombre_completo

        context['coordinadores_modal'] = coordinadores_modal
        context['coordinador_responsable_display'] = coordinador_responsable_display
    if extra_context:
        context.update(extra_context)
    return safe_render(request, f'afiliados/organizacion/crud_{template}.html', context)


@login_required
def lista_estructura_territorial(request):
    denied = _require_organizacion(request)
    if denied:
        return denied
    comunidades = Comunidad.objects.select_related('sector').order_by('nombre')
    sectores = Sector.objects.order_by('nombre')
    centros = CentroVotacion.objects.order_by('nombre')
    return safe_render(request, 'afiliados/organizacion/estructura_territorial.html', {
        'comunidades': comunidades,
        'sectores': sectores,
        'centros': centros,
        'total_comunidades': comunidades.count(),
        'total_sectores': sectores.count(),
        'total_centros': centros.count(),
    })


@login_required
def lista_comunidades(request):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return comunidad_lista(request)


@login_required
def crear_comunidad(request):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return comunidad_nueva(request)


@login_required
def editar_comunidad(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return comunidad_editar(request, pk)


@login_required
def lista_sectores(request):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return sector_nueva(request)


@login_required
def crear_sector(request):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return sector_nueva(request)


@login_required
def editar_sector(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return sector_editar(request, pk)


@login_required
def lista_centros_votacion(request):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return centro_lista(request)


@login_required
def crear_centro_votacion(request):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return centro_nuevo(request)


@login_required
def editar_centro_votacion(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return centro_editar(request, pk)


@login_required
def lista_coordinadores(request):
    return _org_crud(request, CoordinadorOrganizacion, CoordinadorOrganizacionForm, 'coordinadores', 'afiliados:lista_coordinadores', 'afiliados:crear_coordinador', 'afiliados:editar_coordinador')


@login_required
def crear_coordinador(request):
    return _org_crud(request, CoordinadorOrganizacion, CoordinadorOrganizacionForm, 'coordinadores', 'afiliados:lista_coordinadores', 'afiliados:crear_coordinador', 'afiliados:editar_coordinador')


@login_required
def editar_coordinador(request, pk):
    return _org_crud(request, CoordinadorOrganizacion, CoordinadorOrganizacionForm, 'coordinadores', 'afiliados:lista_coordinadores', 'afiliados:crear_coordinador', 'afiliados:editar_coordinador', pk=pk)


@login_required
def detalle_coordinador(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/organizacion/detalle_generico.html', {'obj': get_object_or_404(CoordinadorOrganizacion, pk=pk), 'titulo': 'Coordinador'})


@login_required
def lista_lideres(request):
    return _org_crud(request, LiderComunitarioOrganizacion, LiderComunitarioOrganizacionForm, 'lideres', 'afiliados:lista_lideres', 'afiliados:crear_lider', 'afiliados:editar_lider')


@login_required
def crear_lider(request):
    return _org_crud(request, LiderComunitarioOrganizacion, LiderComunitarioOrganizacionForm, 'lideres', 'afiliados:lista_lideres', 'afiliados:crear_lider', 'afiliados:editar_lider')


@login_required
def editar_lider(request, pk):
    return _org_crud(request, LiderComunitarioOrganizacion, LiderComunitarioOrganizacionForm, 'lideres', 'afiliados:lista_lideres', 'afiliados:crear_lider', 'afiliados:editar_lider', pk=pk)


@login_required
def detalle_lider(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/organizacion/detalle_generico.html', {'obj': get_object_or_404(LiderComunitarioOrganizacion, pk=pk), 'titulo': 'Líder comunitario'})


@login_required
def lista_estructuras(request):
    return _org_crud(request, EstructuraOrganizativa, EstructuraOrganizativaForm, 'estructuras', 'afiliados:lista_estructuras', 'afiliados:crear_estructura', 'afiliados:editar_estructura')


@login_required
def crear_estructura(request):
    return _org_crud(request, EstructuraOrganizativa, EstructuraOrganizativaForm, 'estructuras', 'afiliados:lista_estructuras', 'afiliados:crear_estructura', 'afiliados:editar_estructura')


@login_required
def editar_estructura(request, pk):
    return _org_crud(request, EstructuraOrganizativa, EstructuraOrganizativaForm, 'estructuras', 'afiliados:lista_estructuras', 'afiliados:crear_estructura', 'afiliados:editar_estructura', pk=pk)


@login_required
def detalle_estructura(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied
    estructura = get_object_or_404(EstructuraOrganizativa.objects.select_related('comunidad', 'sector', 'centro_votacion', 'coordinador_responsable'), pk=pk)

    if request.method == 'POST':
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        dpi = re.sub(r"\D", "", request.POST.get('dpi', ''))
        verified_dpi = re.sub(r"\D", "", request.POST.get('verified_dpi', ''))

        if not dpi or verified_dpi != dpi:
            msg = 'Primero debe verificar empadronamiento del DPI antes de confirmar agregado.'
            if is_ajax:
                return JsonResponse({'ok': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('afiliados:detalle_estructura', pk=estructura.pk)

        payload, status_code = _resultado_padron_local(dpi)

        if status_code != 200 or not payload.get('ok') or not payload.get('found'):
            msg = payload.get('error') or payload.get('message') or 'No aparece empadronado en padrón local.'
            if is_ajax:
                return JsonResponse({'ok': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('afiliados:detalle_estructura', pk=estructura.pk)

        persona = payload.get('data') or {}
        if EstructuraIntegrante.objects.filter(estructura=estructura, dpi=dpi).exists():
            msg = 'Este integrante ya está registrado en la estructura.'
            if is_ajax:
                return JsonResponse({'ok': False, 'message': msg}, status=409)
            messages.warning(request, msg)
            return redirect('afiliados:detalle_estructura', pk=estructura.pk)

        comunidad = Comunidad.objects.filter(nombre=persona.get('comunidad')).first() if persona.get('comunidad') else None
        try:
            integrante = EstructuraIntegrante.objects.create(
                estructura=estructura,
                dpi=dpi,
                nombre_completo=persona.get('nombre_completo') or f"Integrante {dpi}",
                comunidad=comunidad,
                usuario_registro=request.user,
            )
        except IntegrityError:
            msg = 'Este integrante ya está registrado en la estructura.'
            if is_ajax:
                return JsonResponse({'ok': False, 'message': msg}, status=409)
            messages.warning(request, msg)
            return redirect('afiliados:detalle_estructura', pk=estructura.pk)

        if is_ajax:
            total = estructura.integrantes.count()
            return JsonResponse({
                'ok': True,
                'message': 'Integrante agregado correctamente a la estructura.',
                'integrante': {
                    'id': integrante.id,
                    'dpi': integrante.dpi,
                    'nombre_completo': integrante.nombre_completo,
                    'estado': integrante.estado,
                    'fecha_registro': timezone.localtime(integrante.fecha_registro).strftime('%d/%m/%Y %H:%M'),
                },
                'total_integrantes': total,
            })

        messages.success(request, 'Integrante agregado correctamente a la estructura.')
        return redirect('afiliados:detalle_estructura', pk=estructura.pk)

    integrantes = estructura.integrantes.select_related('usuario_registro').all()
    return safe_render(request, 'afiliados/organizacion/detalle_estructura.html', {
        'estructura': estructura,
        'integrantes': integrantes,
        'total_integrantes': integrantes.count(),
        'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_organizacion'),
    })


@login_required
@require_POST
def eliminar_integrante_estructura(request, pk, integrante_id):
    denied = _require_organizacion(request)
    if denied:
        return denied

    estructura = get_object_or_404(EstructuraOrganizativa, pk=pk)
    integrante = get_object_or_404(EstructuraIntegrante, pk=integrante_id, estructura=estructura)
    integrante.delete()
    messages.success(request, 'Integrante eliminado correctamente de la estructura.')
    return redirect('afiliados:detalle_estructura', pk=estructura.pk)


@login_required
def lista_responsables(request):
    return _org_crud(request, ResponsableTerritorial, ResponsableTerritorialForm, 'responsables', 'afiliados:lista_responsables', 'afiliados:crear_responsable', 'afiliados:editar_responsable')


@login_required
def crear_responsable(request):
    return _org_crud(request, ResponsableTerritorial, ResponsableTerritorialForm, 'responsables', 'afiliados:lista_responsables', 'afiliados:crear_responsable', 'afiliados:editar_responsable')


@login_required
def editar_responsable(request, pk):
    return _org_crud(request, ResponsableTerritorial, ResponsableTerritorialForm, 'responsables', 'afiliados:lista_responsables', 'afiliados:crear_responsable', 'afiliados:editar_responsable', pk=pk)


@login_required
def detalle_responsable(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/organizacion/detalle_generico.html', {'obj': get_object_or_404(ResponsableTerritorial, pk=pk), 'titulo': 'Responsable territorial'})


@login_required
def lista_reuniones(request):
    return _org_crud(request, ReunionTerritorial, ReunionTerritorialForm, 'reuniones', 'afiliados:lista_reuniones', 'afiliados:crear_reunion', 'afiliados:editar_reunion')


@login_required
def crear_reunion(request):
    return _org_crud(request, ReunionTerritorial, ReunionTerritorialForm, 'reuniones', 'afiliados:lista_reuniones', 'afiliados:crear_reunion', 'afiliados:editar_reunion')


@login_required
def editar_reunion(request, pk):
    return _org_crud(request, ReunionTerritorial, ReunionTerritorialForm, 'reuniones', 'afiliados:lista_reuniones', 'afiliados:crear_reunion', 'afiliados:editar_reunion', pk=pk)


@login_required
def detalle_reunion(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied
    reunion = get_object_or_404(ReunionTerritorial, pk=pk)
    return safe_render(request, 'afiliados/organizacion/detalle_reunion.html', {'reunion': reunion})


@login_required
def lista_incidencias(request):
    return _org_crud(request, IncidenciaTerritorial, IncidenciaTerritorialForm, 'incidencias', 'afiliados:lista_incidencias', 'afiliados:crear_incidencia', 'afiliados:editar_incidencia')


@login_required
def crear_incidencia(request):
    return _org_crud(request, IncidenciaTerritorial, IncidenciaTerritorialForm, 'incidencias', 'afiliados:lista_incidencias', 'afiliados:crear_incidencia', 'afiliados:editar_incidencia')


@login_required
def editar_incidencia(request, pk):
    return _org_crud(request, IncidenciaTerritorial, IncidenciaTerritorialForm, 'incidencias', 'afiliados:lista_incidencias', 'afiliados:crear_incidencia', 'afiliados:editar_incidencia', pk=pk)


@login_required
def detalle_incidencia(request, pk):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/organizacion/detalle_generico.html', {'obj': get_object_or_404(IncidenciaTerritorial, pk=pk), 'titulo': 'Incidencia territorial'})


@login_required
def reporte_cobertura_comunidades(request):
    return panorama_municipal_organizacion(request)


@login_required
def reporte_coordinadores(request):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/organizacion/reporte_coordinadores.html', {
        'coordinadores': CoordinadorOrganizacion.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha_creacion')
    })


@login_required
def reporte_reuniones(request):
    denied = _require_organizacion(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/organizacion/reporte_reuniones.html', {
        'reuniones': ReunionTerritorial.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha')
    })


@login_required
def reporte_crecimiento_mensual(request):
    return panorama_municipal_organizacion(request)


@login_required
@require_GET
def organizacion_comunidad_lookup(request):
    denied = _require_organizacion(request)
    if denied:
        return denied
    comunidad_id = request.GET.get('comunidad_id')
    comunidad = Comunidad.objects.select_related('sector').filter(pk=comunidad_id).first()
    if not comunidad:
        return JsonResponse({'ok': False, 'error': 'Comunidad no encontrada.'}, status=404)

    centro = comunidad.sector.centros.first() if comunidad.sector else None
    return JsonResponse({
        'ok': True,
        'data': {
            'comunidad': comunidad.nombre,
            'sector': comunidad.sector.nombre if comunidad.sector else None,
            'centro_votacion': centro.nombre if centro else None,
        }
    })


def _es_usuario_juventud(user):
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=['Jovenes', 'Administrador']).exists()


def _require_juventud(request):
    if not _es_usuario_juventud(request.user):
        return HttpResponseForbidden("No tiene permisos para acceder a Secretaría de la Juventud.")
    return None


@login_required
def dashboard_juventud(request):
    denied = _require_juventud(request)
    if denied:
        return denied

    registros = JuventudIntegrante.objects.select_related('afiliado')
    context = {
        'total_integrantes': registros.count(),
        'total_pendientes': registros.filter(estado=JuventudIntegrante.Estado.PENDIENTE).count(),
        'total_revisados': registros.filter(estado=JuventudIntegrante.Estado.REVISADO).count(),
        'total_aprobados': registros.filter(estado=JuventudIntegrante.Estado.APROBADO).count(),
        'ultimos_registros': registros.order_by('-fecha_creacion')[:10],
    }
    return safe_render(request, 'afiliados/juventud/dashboard.html', context)


@login_required
def lista_integrantes_juventud(request):
    denied = _require_juventud(request)
    if denied:
        return denied

    integrantes = JuventudIntegrante.objects.select_related('afiliado', 'usuario_registro')
    return safe_render(request, 'afiliados/juventud/lista.html', {'integrantes': integrantes})


@login_required
def crear_integrante_juventud(request):
    denied = _require_juventud(request)
    if denied:
        return denied

    form = JuventudIntegranteForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])

        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif JuventudIntegrante.objects.filter(afiliado=afiliado).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de la Juventud.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_registro = request.user
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante agregado a Secretaría de la Juventud correctamente.')
            return redirect('afiliados:lista_integrantes_juventud')

    return safe_render(request, 'afiliados/juventud/form.html', {'form': form, 'es_edicion': False})


@login_required
def detalle_integrante_juventud(request, pk):
    denied = _require_juventud(request)
    if denied:
        return denied

    integrante = get_object_or_404(
        JuventudIntegrante.objects.select_related('afiliado', 'usuario_registro', 'usuario_modificacion'),
        pk=pk,
    )
    return safe_render(request, 'afiliados/juventud/detalle.html', {'integrante': integrante})


@login_required
def editar_integrante_juventud(request, pk):
    denied = _require_juventud(request)
    if denied:
        return denied

    integrante = get_object_or_404(JuventudIntegrante, pk=pk)
    form = JuventudIntegranteForm(request.POST or None, instance=integrante)
    form.fields['dpi'].initial = integrante.afiliado.dpi

    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])
        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif JuventudIntegrante.objects.filter(afiliado=afiliado).exclude(pk=integrante.pk).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de la Juventud.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante de Secretaría de la Juventud actualizado correctamente.')
            return redirect('afiliados:detalle_integrante_juventud', pk=integrante.pk)

    return safe_render(request, 'afiliados/juventud/form.html', {'form': form, 'es_edicion': True, 'integrante': integrante})


@login_required
@require_GET
def buscar_por_dpi_juventud(request):
    denied = _require_juventud(request)
    if denied:
        return denied

    afiliado, dpi_limpio, error_dpi = _afiliado_por_dpi(request.GET.get('dpi', ''))
    if error_dpi:
        return JsonResponse({'ok': False, 'error': error_dpi}, status=400)

    if not afiliado:
        return JsonResponse({'ok': True, 'exists': False, 'message': 'No existe afiliado con ese DPI.'})

    ya_registrado = JuventudIntegrante.objects.filter(afiliado=afiliado).exists()
    return JsonResponse({
        'ok': True,
        'exists': True,
        'already_registered': ya_registrado,
        'data': {
            'dpi': dpi_limpio,
            'nombre_completo': afiliado.nombre_completo,
            'telefono': afiliado.telefono,
            'direccion': afiliado.direccion,
            'comunidad': afiliado.comunidad.nombre if afiliado.comunidad else None,
            'empadronado': afiliado.empadronado,
        }
    })


@login_required
@require_GET
def verificar_empadronamiento_juventud(request):
    denied = _require_juventud(request)
    if denied:
        return denied
    payload, status_code = _resultado_padron_local(request.GET.get("dpi", ""))
    return JsonResponse(payload, status=status_code)


@login_required
@require_POST
def cambiar_estado_juventud(request, pk):
    denied = _require_juventud(request)
    if denied:
        return denied

    integrante = get_object_or_404(JuventudIntegrante, pk=pk)
    nuevo_estado = request.POST.get('estado', '').strip().upper()
    estados_validos = {choice[0] for choice in JuventudIntegrante.Estado.choices}

    if nuevo_estado not in estados_validos:
        messages.error(request, 'Estado inválido.')
        return redirect('afiliados:detalle_integrante_juventud', pk=pk)

    integrante.estado = nuevo_estado
    integrante.usuario_modificacion = request.user
    integrante.save(update_fields=['estado', 'usuario_modificacion', 'fecha_actualizacion'])
    messages.success(request, 'Estado actualizado correctamente.')
    return redirect('afiliados:detalle_integrante_juventud', pk=pk)


@login_required
def panorama_municipal_juventud(request):
    denied = _require_juventud(request)
    if denied:
        return denied

    comunidad_qs = Comunidad.objects.select_related('sector')
    sector_qs = Sector.objects.all()
    centro_qs = CentroVotacion.objects.all()
    coordinadores = CoordinadorJuventud.objects.all()
    lideres = LiderJuvenil.objects.all()
    estructuras = EstructuraJuventud.objects.all()
    reuniones = ReunionJuventud.objects.all()
    incidencias = IncidenciaJuventud.objects.filter(estado=EstadoRegistro.ACTIVO)
    responsables = ResponsableJuventud.objects.filter(estado=EstadoRegistro.ACTIVO)

    comunidad_ids_cubiertas = set(CoordinadorJuventud.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(LiderJuvenil.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(EstructuraJuventud.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(ResponsableJuventud.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True))

    sector_ids_cubiertos = set(CoordinadorJuventud.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True)) | set(EstructuraJuventud.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True)) | set(ResponsableJuventud.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True))

    centro_ids_cubiertos = set(CoordinadorJuventud.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True)) | set(EstructuraJuventud.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True)) | set(ResponsableJuventud.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True))

    crecimiento, crecimiento_series = _construir_series_crecimiento({
        'coordinadores': coordinadores,
        'lideres': lideres,
        'estructuras': estructuras,
        'reuniones': reuniones,
    })

    resumen_estructuras_por_comunidad = comunidad_qs.annotate(estructuras_total=Count('estructuras_juventud', distinct=True)).order_by('nombre')

    resumen_miembros_por_estructura = EstructuraJuventud.objects.select_related('comunidad').annotate(miembros_total=Count('integrantes', distinct=True)).order_by('nombre')

    context = {
        'total_comunidades': comunidad_qs.count(),
        'comunidades_cubiertas': len(comunidad_ids_cubiertas),
        'comunidades_no_cubiertas': max(comunidad_qs.count() - len(comunidad_ids_cubiertas), 0),
        'total_sectores': sector_qs.count(),
        'sectores_con_responsable': responsables.exclude(sector__isnull=True).values('sector_id').distinct().count(),
        'sectores_sin_responsable': max(sector_qs.count() - responsables.exclude(sector__isnull=True).values('sector_id').distinct().count(), 0),
        'total_centros': centro_qs.count(),
        'centros_con_cobertura': len(centro_ids_cubiertos),
        'centros_sin_cobertura': max(centro_qs.count() - len(centro_ids_cubiertos), 0),
        'coordinadores_activos': coordinadores.filter(estado=EstadoRegistro.ACTIVO).count(),
        'coordinadores_inactivos': coordinadores.filter(estado=EstadoRegistro.INACTIVO).count(),
        'total_lideres': lideres.count(),
        'estructuras_activas': estructuras.filter(estado=EstadoRegistro.ACTIVO).count(),
        'reuniones_mes': reuniones.filter(fecha__month=timezone.now().month, fecha__year=timezone.now().year).count(),
        'incidencias_abiertas': incidencias.count(),
        'resumen_comunidad': comunidad_qs.annotate(total_sectores=Count('sector', distinct=True), reuniones_total=Count('reuniones_juventud', distinct=True), incidencias_abiertas=Count('incidencias_juventud', filter=Q(incidencias_juventud__estado=EstadoRegistro.ACTIVO), distinct=True)),
        'resumen_sector': sector_qs.annotate(lideres_total=Count('lideres_juventud', distinct=True), estructuras_total=Count('estructuras_juventud', distinct=True), incidencias_total=Count('incidencias_juventud', distinct=True)),
        'resumen_centro': centro_qs.annotate(incidencias_total=Count('incidencias_juventud', distinct=True), reuniones_total=Count('reuniones_juventud', distinct=True)),
        'resumen_estructuras_por_comunidad': resumen_estructuras_por_comunidad,
        'resumen_miembros_por_estructura': resumen_miembros_por_estructura,
        'crecimiento': json.dumps(crecimiento, cls=DjangoJSONEncoder),
        'charts_data': json.dumps({'cobertura': {'comunidades': [len(comunidad_ids_cubiertas), max(comunidad_qs.count() - len(comunidad_ids_cubiertas), 0)], 'sectores': [len(sector_ids_cubiertos), max(sector_qs.count() - len(sector_ids_cubiertos), 0)], 'centros': [len(centro_ids_cubiertos), max(centro_qs.count() - len(centro_ids_cubiertos), 0)]}, 'estructura': {'coordinadores': [coordinadores.filter(estado=EstadoRegistro.ACTIVO).count(), coordinadores.filter(estado=EstadoRegistro.INACTIVO).count()], 'lideres_total': lideres.count(), 'estructuras_activas': estructuras.filter(estado=EstadoRegistro.ACTIVO).count()}, 'crecimiento': crecimiento_series}, cls=DjangoJSONEncoder),
        'sector_ids_cubiertos': sector_ids_cubiertos,
        'centro_ids_cubiertos': centro_ids_cubiertos,
        'comunidad_ids_cubiertas': comunidad_ids_cubiertas,
    }
    return safe_render(request, 'afiliados/juventud/panorama.html', context)


def _juv_crud(request, model, form_class, template, list_name, create_name, edit_name, pk=None, extra_context=None):
    denied = _require_juventud(request)
    if denied:
        return denied
    instance = get_object_or_404(model, pk=pk) if pk else None
    form = form_class(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if not getattr(obj, 'usuario_creador_id', None):
            obj.usuario_creador = request.user
        obj.usuario_modificador = request.user
        obj.save()
        messages.success(request, 'Registro guardado correctamente.')
        return redirect(list_name)

    items_qs = model.objects.all().order_by('-id')
    if template == 'reuniones':
        items_qs = items_qs.select_related('comunidad', 'sector', 'centro_votacion')

    context = {
        'form': form,
        'items': items_qs[:200],
        'entity_label': template,
        'create_name': create_name,
        'edit_name': edit_name,
        'detail_name': {
            'coordinadores': 'afiliados:detalle_coordinador_juventud',
            'lideres': 'afiliados:detalle_lider_juvenil',
            'estructuras': 'afiliados:detalle_estructura_juventud',
            'responsables': 'afiliados:detalle_responsable_juventud',
            'reuniones': 'afiliados:detalle_reunion_juventud',
            'incidencias': 'afiliados:detalle_incidencia_juventud',
        }.get(template),
        'comunidad_lookup_url': reverse('afiliados:juventud_comunidad_lookup'),
        'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_juventud'),
    }

    if template == 'reuniones':
        responsables_modal = []
        for responsable in ResponsableJuventud.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'RESPONSABLE', 'tipo_label': 'Responsable', 'id': responsable.id, 'nombre': responsable.nombre_completo, 'dpi': responsable.dpi, 'comunidad': responsable.comunidad.nombre if responsable.comunidad else ''})
        for coordinador in CoordinadorJuventud.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'COORDINADOR', 'tipo_label': 'Coordinador', 'id': coordinador.id, 'nombre': coordinador.nombre_completo, 'dpi': coordinador.dpi, 'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else ''})
        for lider in LiderJuvenil.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'LIDER', 'tipo_label': 'Líder juvenil', 'id': lider.id, 'nombre': lider.nombre_completo, 'dpi': lider.dpi, 'comunidad': lider.comunidad.nombre if lider.comunidad else ''})
        context['responsables_modal'] = responsables_modal

    if template == 'estructuras':
        coordinadores_modal = []
        for coordinador in CoordinadorJuventud.objects.select_related('comunidad').order_by('nombre_completo'):
            coordinadores_modal.append({'id': coordinador.id, 'nombre': coordinador.nombre_completo, 'dpi': coordinador.dpi, 'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else '', 'estado': coordinador.get_estado_display()})
        coordinador_responsable_display = ''
        coordinador_id = form['coordinador_responsable'].value()
        if coordinador_id:
            coordinador_actual = CoordinadorJuventud.objects.filter(pk=coordinador_id).first()
            if coordinador_actual:
                coordinador_responsable_display = coordinador_actual.nombre_completo
        context['coordinadores_modal'] = coordinadores_modal
        context['coordinador_responsable_display'] = coordinador_responsable_display

    if extra_context:
        context.update(extra_context)
    return safe_render(request, f'afiliados/juventud/crud_{template}.html', context)


@login_required
def lista_estructura_territorial_juventud(request):
    denied = _require_juventud(request)
    if denied:
        return denied
    comunidades = Comunidad.objects.select_related('sector').order_by('nombre')
    sectores = Sector.objects.order_by('nombre')
    centros = CentroVotacion.objects.order_by('nombre')
    return safe_render(request, 'afiliados/juventud/estructura_territorial.html', {'comunidades': comunidades, 'sectores': sectores, 'centros': centros, 'total_comunidades': comunidades.count(), 'total_sectores': sectores.count(), 'total_centros': centros.count()})


@login_required
def lista_coordinadores_juventud(request):
    return _juv_crud(request, CoordinadorJuventud, CoordinadorJuventudForm, 'coordinadores', 'afiliados:lista_coordinadores_juventud', 'afiliados:crear_coordinador_juventud', 'afiliados:editar_coordinador_juventud')


@login_required
def crear_coordinador_juventud(request):
    return _juv_crud(request, CoordinadorJuventud, CoordinadorJuventudForm, 'coordinadores', 'afiliados:lista_coordinadores_juventud', 'afiliados:crear_coordinador_juventud', 'afiliados:editar_coordinador_juventud')


@login_required
def editar_coordinador_juventud(request, pk):
    return _juv_crud(request, CoordinadorJuventud, CoordinadorJuventudForm, 'coordinadores', 'afiliados:lista_coordinadores_juventud', 'afiliados:crear_coordinador_juventud', 'afiliados:editar_coordinador_juventud', pk=pk)


@login_required
def detalle_coordinador_juventud(request, pk):
    denied = _require_juventud(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/juventud/detalle_generico.html', {'obj': get_object_or_404(CoordinadorJuventud, pk=pk), 'titulo': 'Coordinador juvenil'})


@login_required
def lista_lideres_juveniles(request):
    return _juv_crud(request, LiderJuvenil, LiderJuvenilForm, 'lideres', 'afiliados:lista_lideres_juveniles', 'afiliados:crear_lider_juvenil', 'afiliados:editar_lider_juvenil')


@login_required
def crear_lider_juvenil(request):
    return _juv_crud(request, LiderJuvenil, LiderJuvenilForm, 'lideres', 'afiliados:lista_lideres_juveniles', 'afiliados:crear_lider_juvenil', 'afiliados:editar_lider_juvenil')


@login_required
def editar_lider_juvenil(request, pk):
    return _juv_crud(request, LiderJuvenil, LiderJuvenilForm, 'lideres', 'afiliados:lista_lideres_juveniles', 'afiliados:crear_lider_juvenil', 'afiliados:editar_lider_juvenil', pk=pk)


@login_required
def detalle_lider_juvenil(request, pk):
    denied = _require_juventud(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/juventud/detalle_generico.html', {'obj': get_object_or_404(LiderJuvenil, pk=pk), 'titulo': 'Líder juvenil'})


@login_required
def lista_estructuras_juventud(request):
    return _juv_crud(request, EstructuraJuventud, EstructuraJuventudForm, 'estructuras', 'afiliados:lista_estructuras_juventud', 'afiliados:crear_estructura_juventud', 'afiliados:editar_estructura_juventud')


@login_required
def crear_estructura_juventud(request):
    return _juv_crud(request, EstructuraJuventud, EstructuraJuventudForm, 'estructuras', 'afiliados:lista_estructuras_juventud', 'afiliados:crear_estructura_juventud', 'afiliados:editar_estructura_juventud')


@login_required
def editar_estructura_juventud(request, pk):
    return _juv_crud(request, EstructuraJuventud, EstructuraJuventudForm, 'estructuras', 'afiliados:lista_estructuras_juventud', 'afiliados:crear_estructura_juventud', 'afiliados:editar_estructura_juventud', pk=pk)


@login_required
def detalle_estructura_juventud(request, pk):
    denied = _require_juventud(request)
    if denied:
        return denied
    estructura = get_object_or_404(EstructuraJuventud.objects.select_related('comunidad', 'sector', 'centro_votacion', 'coordinador_responsable'), pk=pk)

    if request.method == 'POST':
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        dpi = re.sub(r"\D", "", request.POST.get('dpi', ''))
        verified_dpi = re.sub(r"\D", "", request.POST.get('verified_dpi', ''))

        if not dpi or verified_dpi != dpi:
            msg = 'Primero debe verificar empadronamiento del DPI antes de confirmar agregado.'
            if is_ajax:
                return JsonResponse({'ok': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('afiliados:detalle_estructura_juventud', pk=estructura.pk)

        payload, status_code = _resultado_padron_local(dpi)
        if status_code != 200 or not payload.get('ok') or not payload.get('found'):
            msg = payload.get('error') or payload.get('message') or 'No aparece empadronado en padrón local.'
            if is_ajax:
                return JsonResponse({'ok': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('afiliados:detalle_estructura_juventud', pk=estructura.pk)

        persona = payload.get('data') or {}
        if EstructuraIntegranteJuventud.objects.filter(estructura=estructura, dpi=dpi).exists():
            msg = 'Este integrante ya está registrado en la estructura juvenil.'
            if is_ajax:
                return JsonResponse({'ok': False, 'message': msg}, status=409)
            messages.warning(request, msg)
            return redirect('afiliados:detalle_estructura_juventud', pk=estructura.pk)

        comunidad = Comunidad.objects.filter(nombre=persona.get('comunidad')).first() if persona.get('comunidad') else None
        integrante = EstructuraIntegranteJuventud.objects.create(
            estructura=estructura,
            dpi=dpi,
            nombre_completo=persona.get('nombre_completo') or f"Integrante {dpi}",
            comunidad=comunidad,
            usuario_registro=request.user,
        )

        if is_ajax:
            total = estructura.integrantes.count()
            return JsonResponse({'ok': True, 'message': 'Integrante agregado correctamente a la estructura juvenil.', 'integrante': {'id': integrante.id, 'dpi': integrante.dpi, 'nombre_completo': integrante.nombre_completo, 'estado': integrante.estado, 'fecha_registro': timezone.localtime(integrante.fecha_registro).strftime('%d/%m/%Y %H:%M')}, 'total_integrantes': total})

        messages.success(request, 'Integrante agregado correctamente a la estructura juvenil.')
        return redirect('afiliados:detalle_estructura_juventud', pk=estructura.pk)

    integrantes = estructura.integrantes.select_related('usuario_registro').all()
    return safe_render(request, 'afiliados/juventud/detalle_estructura.html', {'estructura': estructura, 'integrantes': integrantes, 'total_integrantes': integrantes.count(), 'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_juventud')})


@login_required
@require_POST
def eliminar_integrante_estructura_juventud(request, pk, integrante_id):
    denied = _require_juventud(request)
    if denied:
        return denied

    estructura = get_object_or_404(EstructuraJuventud, pk=pk)
    integrante = get_object_or_404(EstructuraIntegranteJuventud, pk=integrante_id, estructura=estructura)
    integrante.delete()
    messages.success(request, 'Integrante eliminado correctamente de la estructura juvenil.')
    return redirect('afiliados:detalle_estructura_juventud', pk=estructura.pk)


@login_required
def lista_responsables_juventud(request):
    return _juv_crud(request, ResponsableJuventud, ResponsableJuventudForm, 'responsables', 'afiliados:lista_responsables_juventud', 'afiliados:crear_responsable_juventud', 'afiliados:editar_responsable_juventud')


@login_required
def crear_responsable_juventud(request):
    return _juv_crud(request, ResponsableJuventud, ResponsableJuventudForm, 'responsables', 'afiliados:lista_responsables_juventud', 'afiliados:crear_responsable_juventud', 'afiliados:editar_responsable_juventud')


@login_required
def editar_responsable_juventud(request, pk):
    return _juv_crud(request, ResponsableJuventud, ResponsableJuventudForm, 'responsables', 'afiliados:lista_responsables_juventud', 'afiliados:crear_responsable_juventud', 'afiliados:editar_responsable_juventud', pk=pk)


@login_required
def detalle_responsable_juventud(request, pk):
    denied = _require_juventud(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/juventud/detalle_generico.html', {'obj': get_object_or_404(ResponsableJuventud, pk=pk), 'titulo': 'Responsable juvenil'})


@login_required
def lista_reuniones_juventud(request):
    return _juv_crud(request, ReunionJuventud, ReunionJuventudForm, 'reuniones', 'afiliados:lista_reuniones_juventud', 'afiliados:crear_reunion_juventud', 'afiliados:editar_reunion_juventud')


@login_required
def crear_reunion_juventud(request):
    return _juv_crud(request, ReunionJuventud, ReunionJuventudForm, 'reuniones', 'afiliados:lista_reuniones_juventud', 'afiliados:crear_reunion_juventud', 'afiliados:editar_reunion_juventud')


@login_required
def editar_reunion_juventud(request, pk):
    return _juv_crud(request, ReunionJuventud, ReunionJuventudForm, 'reuniones', 'afiliados:lista_reuniones_juventud', 'afiliados:crear_reunion_juventud', 'afiliados:editar_reunion_juventud', pk=pk)


@login_required
def detalle_reunion_juventud(request, pk):
    denied = _require_juventud(request)
    if denied:
        return denied
    reunion = get_object_or_404(ReunionJuventud, pk=pk)
    return safe_render(request, 'afiliados/juventud/detalle_reunion.html', {'reunion': reunion})


@login_required
def lista_incidencias_juventud(request):
    return _juv_crud(request, IncidenciaJuventud, IncidenciaJuventudForm, 'incidencias', 'afiliados:lista_incidencias_juventud', 'afiliados:crear_incidencia_juventud', 'afiliados:editar_incidencia_juventud')


@login_required
def crear_incidencia_juventud(request):
    return _juv_crud(request, IncidenciaJuventud, IncidenciaJuventudForm, 'incidencias', 'afiliados:lista_incidencias_juventud', 'afiliados:crear_incidencia_juventud', 'afiliados:editar_incidencia_juventud')


@login_required
def editar_incidencia_juventud(request, pk):
    return _juv_crud(request, IncidenciaJuventud, IncidenciaJuventudForm, 'incidencias', 'afiliados:lista_incidencias_juventud', 'afiliados:crear_incidencia_juventud', 'afiliados:editar_incidencia_juventud', pk=pk)


@login_required
def detalle_incidencia_juventud(request, pk):
    denied = _require_juventud(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/juventud/detalle_generico.html', {'obj': get_object_or_404(IncidenciaJuventud, pk=pk), 'titulo': 'Incidencia juvenil'})


@login_required
def reporte_cobertura_juventud(request):
    return panorama_municipal_juventud(request)


@login_required
def reporte_coordinadores_juventud(request):
    denied = _require_juventud(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/juventud/reporte_coordinadores.html', {'coordinadores': CoordinadorJuventud.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha_creacion')})


@login_required
def reporte_reuniones_juventud(request):
    denied = _require_juventud(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/juventud/reporte_reuniones.html', {'reuniones': ReunionJuventud.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha')})


@login_required
def reporte_crecimiento_mensual_juventud(request):
    return panorama_municipal_juventud(request)


@login_required
@require_GET
def juventud_comunidad_lookup(request):
    denied = _require_juventud(request)
    if denied:
        return denied
    comunidad_id = request.GET.get('comunidad_id')
    comunidad = Comunidad.objects.select_related('sector').filter(pk=comunidad_id).first()
    if not comunidad:
        return JsonResponse({'ok': False, 'error': 'Comunidad no encontrada.'}, status=404)

    centro = comunidad.sector.centros.first() if comunidad.sector else None
    return JsonResponse({'ok': True, 'data': {'comunidad': comunidad.nombre, 'sector': comunidad.sector.nombre if comunidad.sector else None, 'centro_votacion': centro.nombre if centro else None}})


def _es_usuario_mujeres(user):
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=['Mujeres', 'Administrador']).exists()


def _require_mujeres(request):
    if not _es_usuario_mujeres(request.user):
        return HttpResponseForbidden("No tiene permisos para acceder a Secretaría de la Mujeres.")
    return None


@login_required
def dashboard_mujeres(request):
    denied = _require_mujeres(request)
    if denied:
        return denied

    registros = MujeresIntegrante.objects.select_related('afiliado')
    context = {
        'total_integrantes': registros.count(),
        'total_pendientes': registros.filter(estado=MujeresIntegrante.Estado.PENDIENTE).count(),
        'total_revisados': registros.filter(estado=MujeresIntegrante.Estado.REVISADO).count(),
        'total_aprobados': registros.filter(estado=MujeresIntegrante.Estado.APROBADO).count(),
        'ultimos_registros': registros.order_by('-fecha_creacion')[:10],
    }
    return safe_render(request, 'afiliados/mujeres/dashboard.html', context)


@login_required
def lista_integrantes_mujeres(request):
    denied = _require_mujeres(request)
    if denied:
        return denied

    integrantes = MujeresIntegrante.objects.select_related('afiliado', 'usuario_registro').order_by('-fecha_creacion')
    context = {
        'integrantes': integrantes,
        'total_integrantes': integrantes.count(),
        'total_pendientes': integrantes.filter(estado=MujeresIntegrante.Estado.PENDIENTE).count(),
        'total_revisados': integrantes.filter(estado=MujeresIntegrante.Estado.REVISADO).count(),
        'total_aprobados': integrantes.filter(estado=MujeresIntegrante.Estado.APROBADO).count(),
    }
    return safe_render(request, 'afiliados/mujeres/lista.html', context)


@login_required
def crear_integrante_mujeres(request):
    denied = _require_mujeres(request)
    if denied:
        return denied

    form = MujeresIntegranteForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])

        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif MujeresIntegrante.objects.filter(afiliado=afiliado).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de la Mujeres.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_registro = request.user
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante agregado a Secretaría de la Mujeres correctamente.')
            return redirect('afiliados:lista_integrantes_mujeres')

    return safe_render(request, 'afiliados/mujeres/form.html', {'form': form, 'es_edicion': False})


@login_required
def detalle_integrante_mujeres(request, pk):
    denied = _require_mujeres(request)
    if denied:
        return denied

    integrante = get_object_or_404(
        MujeresIntegrante.objects.select_related('afiliado', 'usuario_registro', 'usuario_modificacion'),
        pk=pk,
    )
    return safe_render(request, 'afiliados/mujeres/detalle.html', {'integrante': integrante})


@login_required
def editar_integrante_mujeres(request, pk):
    denied = _require_mujeres(request)
    if denied:
        return denied

    integrante = get_object_or_404(MujeresIntegrante, pk=pk)
    form = MujeresIntegranteForm(request.POST or None, instance=integrante)
    form.fields['dpi'].initial = integrante.afiliado.dpi

    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])
        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif MujeresIntegrante.objects.filter(afiliado=afiliado).exclude(pk=integrante.pk).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de la Mujeres.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante de Secretaría de la Mujeres actualizado correctamente.')
            return redirect('afiliados:detalle_integrante_mujeres', pk=integrante.pk)

    return safe_render(request, 'afiliados/mujeres/form.html', {'form': form, 'es_edicion': True, 'integrante': integrante})


@login_required
@require_GET
def buscar_por_dpi_mujeres(request):
    denied = _require_mujeres(request)
    if denied:
        return denied

    afiliado, dpi_limpio, error_dpi = _afiliado_por_dpi(request.GET.get('dpi', ''))
    if error_dpi:
        return JsonResponse({'ok': False, 'error': error_dpi}, status=400)

    if not afiliado:
        return JsonResponse({'ok': True, 'exists': False, 'message': 'No existe afiliado con ese DPI.'})

    ya_registrado = MujeresIntegrante.objects.filter(afiliado=afiliado).exists()
    return JsonResponse({
        'ok': True,
        'exists': True,
        'already_registered': ya_registrado,
        'data': {
            'dpi': dpi_limpio,
            'nombre_completo': afiliado.nombre_completo,
            'telefono': afiliado.telefono,
            'direccion': afiliado.direccion,
            'comunidad': afiliado.comunidad.nombre if afiliado.comunidad else None,
            'empadronado': afiliado.empadronado,
        }
    })


@login_required
@require_GET
def verificar_empadronamiento_mujeres(request):
    denied = _require_mujeres(request)
    if denied:
        return denied
    payload, status_code = _resultado_padron_local(request.GET.get("dpi", ""))
    return JsonResponse(payload, status=status_code)


@login_required
@require_POST
def cambiar_estado_mujeres(request, pk):
    denied = _require_mujeres(request)
    if denied:
        return denied

    integrante = get_object_or_404(MujeresIntegrante, pk=pk)
    nuevo_estado = request.POST.get('estado', '').strip().upper()
    estados_validos = {choice[0] for choice in MujeresIntegrante.Estado.choices}

    if nuevo_estado not in estados_validos:
        messages.error(request, 'Estado inválido.')
        return redirect('afiliados:detalle_integrante_mujeres', pk=pk)

    integrante.estado = nuevo_estado
    integrante.usuario_modificacion = request.user
    integrante.save(update_fields=['estado', 'usuario_modificacion', 'fecha_actualizacion'])
    messages.success(request, 'Estado actualizado correctamente.')
    return redirect('afiliados:detalle_integrante_mujeres', pk=pk)


@login_required
def panorama_municipal_mujeres(request):
    denied = _require_mujeres(request)
    if denied:
        return denied

    comunidad_qs = Comunidad.objects.select_related('sector')
    sector_qs = Sector.objects.all()
    centro_qs = CentroVotacion.objects.all()
    coordinadores = CoordinadoraMujeres.objects.all()
    lideres = LiderMujeres.objects.all()
    estructuras = EstructuraMujeres.objects.all()
    reuniones = ReunionMujeres.objects.all()
    incidencias = IncidenciaMujeres.objects.filter(estado=EstadoRegistro.ACTIVO)
    responsables = ResponsableMujeres.objects.filter(estado=EstadoRegistro.ACTIVO)

    comunidad_ids_cubiertas = set(CoordinadoraMujeres.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(LiderMujeres.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(EstructuraMujeres.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(ResponsableMujeres.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True))

    sector_ids_cubiertos = set(CoordinadoraMujeres.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True)) | set(EstructuraMujeres.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True)) | set(ResponsableMujeres.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True))

    centro_ids_cubiertos = set(CoordinadoraMujeres.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True)) | set(EstructuraMujeres.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True)) | set(ResponsableMujeres.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True))

    crecimiento, crecimiento_series = _construir_series_crecimiento({
        'coordinadores': coordinadores,
        'lideres': lideres,
        'estructuras': estructuras,
        'reuniones': reuniones,
    })

    resumen_estructuras_por_comunidad = comunidad_qs.annotate(estructuras_total=Count('estructuras_mujeres', distinct=True)).order_by('nombre')

    resumen_miembros_por_estructura = EstructuraMujeres.objects.select_related('comunidad').annotate(miembros_total=Count('integrantes', distinct=True)).order_by('nombre')

    context = {
        'total_comunidades': comunidad_qs.count(),
        'comunidades_cubiertas': len(comunidad_ids_cubiertas),
        'comunidades_no_cubiertas': max(comunidad_qs.count() - len(comunidad_ids_cubiertas), 0),
        'total_sectores': sector_qs.count(),
        'sectores_con_responsable': responsables.exclude(sector__isnull=True).values('sector_id').distinct().count(),
        'sectores_sin_responsable': max(sector_qs.count() - responsables.exclude(sector__isnull=True).values('sector_id').distinct().count(), 0),
        'total_centros': centro_qs.count(),
        'centros_con_cobertura': len(centro_ids_cubiertos),
        'centros_sin_cobertura': max(centro_qs.count() - len(centro_ids_cubiertos), 0),
        'coordinadores_activos': coordinadores.filter(estado=EstadoRegistro.ACTIVO).count(),
        'coordinadores_inactivos': coordinadores.filter(estado=EstadoRegistro.INACTIVO).count(),
        'total_lideres': lideres.count(),
        'estructuras_activas': estructuras.filter(estado=EstadoRegistro.ACTIVO).count(),
        'reuniones_mes': reuniones.filter(fecha__month=timezone.now().month, fecha__year=timezone.now().year).count(),
        'incidencias_abiertas': incidencias.count(),
        'resumen_comunidad': comunidad_qs.annotate(total_sectores=Count('sector', distinct=True), reuniones_total=Count('reuniones_mujeres', distinct=True), incidencias_abiertas=Count('incidencias_mujeres', filter=Q(incidencias_mujeres__estado=EstadoRegistro.ACTIVO), distinct=True)),
        'resumen_sector': sector_qs.annotate(lideres_total=Count('lideres_mujeres', distinct=True), estructuras_total=Count('estructuras_mujeres', distinct=True), incidencias_total=Count('incidencias_mujeres', distinct=True)),
        'resumen_centro': centro_qs.annotate(incidencias_total=Count('incidencias_mujeres', distinct=True), reuniones_total=Count('reuniones_mujeres', distinct=True)),
        'resumen_estructuras_por_comunidad': resumen_estructuras_por_comunidad,
        'resumen_miembros_por_estructura': resumen_miembros_por_estructura,
        'crecimiento': json.dumps(crecimiento, cls=DjangoJSONEncoder),
        'charts_data': json.dumps({'cobertura': {'comunidades': [len(comunidad_ids_cubiertas), max(comunidad_qs.count() - len(comunidad_ids_cubiertas), 0)], 'sectores': [len(sector_ids_cubiertos), max(sector_qs.count() - len(sector_ids_cubiertos), 0)], 'centros': [len(centro_ids_cubiertos), max(centro_qs.count() - len(centro_ids_cubiertos), 0)]}, 'estructura': {'coordinadores': [coordinadores.filter(estado=EstadoRegistro.ACTIVO).count(), coordinadores.filter(estado=EstadoRegistro.INACTIVO).count()], 'lideres_total': lideres.count(), 'estructuras_activas': estructuras.filter(estado=EstadoRegistro.ACTIVO).count()}, 'crecimiento': crecimiento_series}, cls=DjangoJSONEncoder),
        'sector_ids_cubiertos': sector_ids_cubiertos,
        'centro_ids_cubiertos': centro_ids_cubiertos,
        'comunidad_ids_cubiertas': comunidad_ids_cubiertas,
    }
    return safe_render(request, 'afiliados/mujeres/panorama.html', context)


def _muj_crud(request, model, form_class, template, list_name, create_name, edit_name, pk=None, extra_context=None):
    denied = _require_mujeres(request)
    if denied:
        return denied
    instance = get_object_or_404(model, pk=pk) if pk else None
    form = form_class(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if not getattr(obj, 'usuario_creador_id', None):
            obj.usuario_creador = request.user
        obj.usuario_modificador = request.user
        obj.save()
        messages.success(request, 'Registro guardado correctamente.')
        return redirect(list_name)

    items_qs = model.objects.all().order_by('-id')
    if template == 'reuniones':
        items_qs = items_qs.select_related('comunidad', 'sector', 'centro_votacion')

    context = {
        'form': form,
        'items': items_qs[:200],
        'entity_label': template,
        'create_name': create_name,
        'edit_name': edit_name,
        'detail_name': {
            'coordinadores': 'afiliados:detalle_coordinador_mujeres',
            'lideres': 'afiliados:detalle_lider_mujeres',
            'estructuras': 'afiliados:detalle_estructura_mujeres',
            'responsables': 'afiliados:detalle_responsable_mujeres',
            'reuniones': 'afiliados:detalle_reunion_mujeres',
            'incidencias': 'afiliados:detalle_incidencia_mujeres',
        }.get(template),
        'comunidad_lookup_url': reverse('afiliados:mujeres_comunidad_lookup'),
        'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_mujeres'),
    }

    if template == 'reuniones':
        responsables_modal = []
        for responsable in ResponsableMujeres.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'RESPONSABLE', 'tipo_label': 'Responsable', 'id': responsable.id, 'nombre': responsable.nombre_completo, 'dpi': responsable.dpi, 'comunidad': responsable.comunidad.nombre if responsable.comunidad else ''})
        for coordinador in CoordinadoraMujeres.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'COORDINADOR', 'tipo_label': 'Coordinador', 'id': coordinador.id, 'nombre': coordinador.nombre_completo, 'dpi': coordinador.dpi, 'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else ''})
        for lider in LiderMujeres.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'LIDER', 'tipo_label': 'Líder mujeres', 'id': lider.id, 'nombre': lider.nombre_completo, 'dpi': lider.dpi, 'comunidad': lider.comunidad.nombre if lider.comunidad else ''})
        context['responsables_modal'] = responsables_modal

    if template == 'estructuras':
        coordinadores_modal = []
        for coordinador in CoordinadoraMujeres.objects.select_related('comunidad').order_by('nombre_completo'):
            coordinadores_modal.append({'id': coordinador.id, 'nombre': coordinador.nombre_completo, 'dpi': coordinador.dpi, 'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else '', 'estado': coordinador.get_estado_display()})
        coordinador_responsable_display = ''
        coordinador_id = form['coordinador_responsable'].value()
        if coordinador_id:
            coordinador_actual = CoordinadoraMujeres.objects.filter(pk=coordinador_id).first()
            if coordinador_actual:
                coordinador_responsable_display = coordinador_actual.nombre_completo
        context['coordinadores_modal'] = coordinadores_modal
        context['coordinador_responsable_display'] = coordinador_responsable_display

    if extra_context:
        context.update(extra_context)
    return safe_render(request, f'afiliados/mujeres/crud_{template}.html', context)


@login_required
def lista_estructura_territorial_mujeres(request):
    denied = _require_mujeres(request)
    if denied:
        return denied
    comunidades = Comunidad.objects.select_related('sector').order_by('nombre')
    sectores = Sector.objects.order_by('nombre')
    centros = CentroVotacion.objects.order_by('nombre')
    return safe_render(request, 'afiliados/mujeres/estructura_territorial.html', {'comunidades': comunidades, 'sectores': sectores, 'centros': centros, 'total_comunidades': comunidades.count(), 'total_sectores': sectores.count(), 'total_centros': centros.count()})


@login_required
def lista_coordinadores_mujeres(request):
    return _muj_crud(request, CoordinadoraMujeres, CoordinadoraMujeresForm, 'coordinadores', 'afiliados:lista_coordinadores_mujeres', 'afiliados:crear_coordinador_mujeres', 'afiliados:editar_coordinador_mujeres')


@login_required
def crear_coordinador_mujeres(request):
    return _muj_crud(request, CoordinadoraMujeres, CoordinadoraMujeresForm, 'coordinadores', 'afiliados:lista_coordinadores_mujeres', 'afiliados:crear_coordinador_mujeres', 'afiliados:editar_coordinador_mujeres')


@login_required
def editar_coordinador_mujeres(request, pk):
    return _muj_crud(request, CoordinadoraMujeres, CoordinadoraMujeresForm, 'coordinadores', 'afiliados:lista_coordinadores_mujeres', 'afiliados:crear_coordinador_mujeres', 'afiliados:editar_coordinador_mujeres', pk=pk)


@login_required
def detalle_coordinador_mujeres(request, pk):
    denied = _require_mujeres(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/mujeres/detalle_generico.html', {'obj': get_object_or_404(CoordinadoraMujeres, pk=pk), 'titulo': 'Coordinador mujeres'})


@login_required
def lista_lideres_mujeres(request):
    return _muj_crud(request, LiderMujeres, LiderMujeresForm, 'lideres', 'afiliados:lista_lideres_mujeres', 'afiliados:crear_lider_mujeres', 'afiliados:editar_lider_mujeres')


@login_required
def crear_lider_mujeres(request):
    return _muj_crud(request, LiderMujeres, LiderMujeresForm, 'lideres', 'afiliados:lista_lideres_mujeres', 'afiliados:crear_lider_mujeres', 'afiliados:editar_lider_mujeres')


@login_required
def editar_lider_mujeres(request, pk):
    return _muj_crud(request, LiderMujeres, LiderMujeresForm, 'lideres', 'afiliados:lista_lideres_mujeres', 'afiliados:crear_lider_mujeres', 'afiliados:editar_lider_mujeres', pk=pk)


@login_required
def detalle_lider_mujeres(request, pk):
    denied = _require_mujeres(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/mujeres/detalle_generico.html', {'obj': get_object_or_404(LiderMujeres, pk=pk), 'titulo': 'Líder mujeres'})


@login_required
def lista_estructuras_mujeres(request):
    return _muj_crud(request, EstructuraMujeres, EstructuraMujeresForm, 'estructuras', 'afiliados:lista_estructuras_mujeres', 'afiliados:crear_estructura_mujeres', 'afiliados:editar_estructura_mujeres')


@login_required
def crear_estructura_mujeres(request):
    return _muj_crud(request, EstructuraMujeres, EstructuraMujeresForm, 'estructuras', 'afiliados:lista_estructuras_mujeres', 'afiliados:crear_estructura_mujeres', 'afiliados:editar_estructura_mujeres')


@login_required
def editar_estructura_mujeres(request, pk):
    return _muj_crud(request, EstructuraMujeres, EstructuraMujeresForm, 'estructuras', 'afiliados:lista_estructuras_mujeres', 'afiliados:crear_estructura_mujeres', 'afiliados:editar_estructura_mujeres', pk=pk)


@login_required
def detalle_estructura_mujeres(request, pk):
    denied = _require_mujeres(request)
    if denied:
        return denied
    estructura = get_object_or_404(EstructuraMujeres.objects.select_related('comunidad', 'sector', 'centro_votacion', 'coordinador_responsable'), pk=pk)

    if request.method == 'POST':
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        dpi = re.sub(r"\D", "", request.POST.get('dpi', ''))
        verified_dpi = re.sub(r"\D", "", request.POST.get('verified_dpi', ''))

        if not dpi or verified_dpi != dpi:
            msg = 'Primero debe verificar empadronamiento del DPI antes de confirmar agregado.'
            if is_ajax:
                return JsonResponse({'ok': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('afiliados:detalle_estructura_mujeres', pk=estructura.pk)

        payload, status_code = _resultado_padron_local(dpi)
        if status_code != 200 or not payload.get('ok') or not payload.get('found'):
            msg = payload.get('error') or payload.get('message') or 'No aparece empadronado en padrón local.'
            if is_ajax:
                return JsonResponse({'ok': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('afiliados:detalle_estructura_mujeres', pk=estructura.pk)

        persona = payload.get('data') or {}
        if EstructuraIntegranteMujeres.objects.filter(estructura=estructura, dpi=dpi).exists():
            msg = 'Este integrante ya está registrado en la estructura mujeres.'
            if is_ajax:
                return JsonResponse({'ok': False, 'message': msg}, status=409)
            messages.warning(request, msg)
            return redirect('afiliados:detalle_estructura_mujeres', pk=estructura.pk)

        comunidad = Comunidad.objects.filter(nombre=persona.get('comunidad')).first() if persona.get('comunidad') else None
        integrante = EstructuraIntegranteMujeres.objects.create(
            estructura=estructura,
            dpi=dpi,
            nombre_completo=persona.get('nombre_completo') or f"Integrante {dpi}",
            comunidad=comunidad,
            usuario_registro=request.user,
        )

        if is_ajax:
            total = estructura.integrantes.count()
            return JsonResponse({'ok': True, 'message': 'Integrante agregado correctamente a la estructura mujeres.', 'integrante': {'id': integrante.id, 'dpi': integrante.dpi, 'nombre_completo': integrante.nombre_completo, 'estado': integrante.estado, 'fecha_registro': timezone.localtime(integrante.fecha_registro).strftime('%d/%m/%Y %H:%M')}, 'total_integrantes': total})

        messages.success(request, 'Integrante agregado correctamente a la estructura mujeres.')
        return redirect('afiliados:detalle_estructura_mujeres', pk=estructura.pk)

    integrantes = estructura.integrantes.select_related('usuario_registro').all()
    return safe_render(request, 'afiliados/mujeres/detalle_estructura.html', {'estructura': estructura, 'integrantes': integrantes, 'total_integrantes': integrantes.count(), 'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_mujeres')})


@login_required
@require_POST
def eliminar_integrante_estructura_mujeres(request, pk, integrante_id):
    denied = _require_mujeres(request)
    if denied:
        return denied

    estructura = get_object_or_404(EstructuraMujeres, pk=pk)
    integrante = get_object_or_404(EstructuraIntegranteMujeres, pk=integrante_id, estructura=estructura)
    integrante.delete()
    messages.success(request, 'Integrante eliminado correctamente de la estructura mujeres.')
    return redirect('afiliados:detalle_estructura_mujeres', pk=estructura.pk)


@login_required
def lista_responsables_mujeres(request):
    return _muj_crud(request, ResponsableMujeres, ResponsableMujeresForm, 'responsables', 'afiliados:lista_responsables_mujeres', 'afiliados:crear_responsable_mujeres', 'afiliados:editar_responsable_mujeres')


@login_required
def crear_responsable_mujeres(request):
    return _muj_crud(request, ResponsableMujeres, ResponsableMujeresForm, 'responsables', 'afiliados:lista_responsables_mujeres', 'afiliados:crear_responsable_mujeres', 'afiliados:editar_responsable_mujeres')


@login_required
def editar_responsable_mujeres(request, pk):
    return _muj_crud(request, ResponsableMujeres, ResponsableMujeresForm, 'responsables', 'afiliados:lista_responsables_mujeres', 'afiliados:crear_responsable_mujeres', 'afiliados:editar_responsable_mujeres', pk=pk)


@login_required
def detalle_responsable_mujeres(request, pk):
    denied = _require_mujeres(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/mujeres/detalle_generico.html', {'obj': get_object_or_404(ResponsableMujeres, pk=pk), 'titulo': 'Responsable mujeres'})


@login_required
def lista_reuniones_mujeres(request):
    return _muj_crud(request, ReunionMujeres, ReunionMujeresForm, 'reuniones', 'afiliados:lista_reuniones_mujeres', 'afiliados:crear_reunion_mujeres', 'afiliados:editar_reunion_mujeres')


@login_required
def crear_reunion_mujeres(request):
    return _muj_crud(request, ReunionMujeres, ReunionMujeresForm, 'reuniones', 'afiliados:lista_reuniones_mujeres', 'afiliados:crear_reunion_mujeres', 'afiliados:editar_reunion_mujeres')


@login_required
def editar_reunion_mujeres(request, pk):
    return _muj_crud(request, ReunionMujeres, ReunionMujeresForm, 'reuniones', 'afiliados:lista_reuniones_mujeres', 'afiliados:crear_reunion_mujeres', 'afiliados:editar_reunion_mujeres', pk=pk)


@login_required
def detalle_reunion_mujeres(request, pk):
    denied = _require_mujeres(request)
    if denied:
        return denied
    reunion = get_object_or_404(ReunionMujeres, pk=pk)
    return safe_render(request, 'afiliados/mujeres/detalle_reunion.html', {'reunion': reunion})


@login_required
def lista_incidencias_mujeres(request):
    return _muj_crud(request, IncidenciaMujeres, IncidenciaMujeresForm, 'incidencias', 'afiliados:lista_incidencias_mujeres', 'afiliados:crear_incidencia_mujeres', 'afiliados:editar_incidencia_mujeres')


@login_required
def crear_incidencia_mujeres(request):
    return _muj_crud(request, IncidenciaMujeres, IncidenciaMujeresForm, 'incidencias', 'afiliados:lista_incidencias_mujeres', 'afiliados:crear_incidencia_mujeres', 'afiliados:editar_incidencia_mujeres')


@login_required
def editar_incidencia_mujeres(request, pk):
    return _muj_crud(request, IncidenciaMujeres, IncidenciaMujeresForm, 'incidencias', 'afiliados:lista_incidencias_mujeres', 'afiliados:crear_incidencia_mujeres', 'afiliados:editar_incidencia_mujeres', pk=pk)


@login_required
def detalle_incidencia_mujeres(request, pk):
    denied = _require_mujeres(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/mujeres/detalle_generico.html', {'obj': get_object_or_404(IncidenciaMujeres, pk=pk), 'titulo': 'Incidencia mujeres'})


@login_required
def reporte_cobertura_mujeres(request):
    return panorama_municipal_mujeres(request)


@login_required
def reporte_coordinadores_mujeres(request):
    denied = _require_mujeres(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/mujeres/reporte_coordinadores.html', {'coordinadores': CoordinadoraMujeres.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha_creacion')})


@login_required
def reporte_reuniones_mujeres(request):
    denied = _require_mujeres(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/mujeres/reporte_reuniones.html', {'reuniones': ReunionMujeres.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha')})


@login_required
def reporte_crecimiento_mensual_mujeres(request):
    return panorama_municipal_mujeres(request)


@login_required
@require_GET
def mujeres_comunidad_lookup(request):
    denied = _require_mujeres(request)
    if denied:
        return denied
    comunidad_id = request.GET.get('comunidad_id')
    comunidad = Comunidad.objects.select_related('sector').filter(pk=comunidad_id).first()
    if not comunidad:
        return JsonResponse({'ok': False, 'error': 'Comunidad no encontrada.'}, status=404)

    centro = comunidad.sector.centros.first() if comunidad.sector else None
    return JsonResponse({'ok': True, 'data': {'comunidad': comunidad.nombre, 'sector': comunidad.sector.nombre if comunidad.sector else None, 'centro_votacion': centro.nombre if centro else None}})


def _es_usuario_logistica(user):
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=['Logistica', 'Administrador']).exists()


def _require_logistica(request):
    if not _es_usuario_logistica(request.user):
        return HttpResponseForbidden("No tiene permisos para acceder a Secretaría de Logística.")
    return None


@login_required
def dashboard_logistica(request):
    denied = _require_logistica(request)
    if denied:
        return denied

    registros = LogisticaIntegrante.objects.select_related('afiliado')
    context = {
        'total_integrantes': registros.count(),
        'total_pendientes': registros.filter(estado=LogisticaIntegrante.Estado.PENDIENTE).count(),
        'total_revisados': registros.filter(estado=LogisticaIntegrante.Estado.REVISADO).count(),
        'total_aprobados': registros.filter(estado=LogisticaIntegrante.Estado.APROBADO).count(),
        'ultimos_registros': registros.order_by('-fecha_creacion')[:10],
    }
    return safe_render(request, 'afiliados/logistica/dashboard.html', context)


@login_required
def lista_integrantes_logistica(request):
    denied = _require_logistica(request)
    if denied:
        return denied

    integrantes = LogisticaIntegrante.objects.select_related('afiliado', 'usuario_registro').order_by('-fecha_creacion')
    context = {
        'integrantes': integrantes,
        'total_integrantes': integrantes.count(),
        'total_pendientes': integrantes.filter(estado=LogisticaIntegrante.Estado.PENDIENTE).count(),
        'total_revisados': integrantes.filter(estado=LogisticaIntegrante.Estado.REVISADO).count(),
        'total_aprobados': integrantes.filter(estado=LogisticaIntegrante.Estado.APROBADO).count(),
    }
    return safe_render(request, 'afiliados/logistica/lista.html', context)


@login_required
def crear_integrante_logistica(request):
    denied = _require_logistica(request)
    if denied:
        return denied

    form = LogisticaIntegranteForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])

        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif LogisticaIntegrante.objects.filter(afiliado=afiliado).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de Logística.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_registro = request.user
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante agregado a Secretaría de Logística correctamente.')
            return redirect('afiliados:lista_integrantes_logistica')

    return safe_render(request, 'afiliados/logistica/form.html', {'form': form, 'es_edicion': False})


@login_required
def detalle_integrante_logistica(request, pk):
    denied = _require_logistica(request)
    if denied:
        return denied
    integrante = get_object_or_404(
        LogisticaIntegrante.objects.select_related('afiliado', 'usuario_registro', 'usuario_modificacion'),
        pk=pk,
    )
    return safe_render(request, 'afiliados/logistica/detalle.html', {'integrante': integrante})


@login_required
def editar_integrante_logistica(request, pk):
    denied = _require_logistica(request)
    if denied:
        return denied

    integrante = get_object_or_404(LogisticaIntegrante, pk=pk)
    form = LogisticaIntegranteForm(request.POST or None, instance=integrante)
    form.fields['dpi'].initial = integrante.afiliado.dpi

    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])
        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif LogisticaIntegrante.objects.filter(afiliado=afiliado).exclude(pk=integrante.pk).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de Logística.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante de Secretaría de Logística actualizado correctamente.')
            return redirect('afiliados:detalle_integrante_logistica', pk=integrante.pk)

    return safe_render(request, 'afiliados/logistica/form.html', {'form': form, 'es_edicion': True, 'integrante': integrante})


@login_required
@require_GET
def buscar_por_dpi_logistica(request):
    denied = _require_logistica(request)
    if denied:
        return denied

    afiliado, dpi_limpio, error_dpi = _afiliado_por_dpi(request.GET.get('dpi', ''))
    if error_dpi:
        return JsonResponse({'ok': False, 'error': error_dpi}, status=400)
    if not afiliado:
        return JsonResponse({'ok': True, 'exists': False, 'message': 'No existe afiliado con ese DPI.'})

    ya_registrado = LogisticaIntegrante.objects.filter(afiliado=afiliado).exists()
    return JsonResponse({
        'ok': True,
        'exists': True,
        'already_registered': ya_registrado,
        'data': {
            'dpi': dpi_limpio,
            'nombre_completo': afiliado.nombre_completo,
            'telefono': afiliado.telefono,
            'direccion': afiliado.direccion,
            'comunidad': afiliado.comunidad.nombre if afiliado.comunidad else None,
            'empadronado': afiliado.empadronado,
        }
    })


@login_required
@require_GET
def verificar_empadronamiento_logistica(request):
    denied = _require_logistica(request)
    if denied:
        return denied
    payload, status_code = _resultado_padron_local(request.GET.get("dpi", ""))
    return JsonResponse(payload, status=status_code)


@login_required
@require_POST
def cambiar_estado_logistica(request, pk):
    denied = _require_logistica(request)
    if denied:
        return denied
    integrante = get_object_or_404(LogisticaIntegrante, pk=pk)
    nuevo_estado = request.POST.get('estado', '').strip().upper()
    estados_validos = {choice[0] for choice in LogisticaIntegrante.Estado.choices}
    if nuevo_estado not in estados_validos:
        messages.error(request, 'Estado inválido.')
        return redirect('afiliados:detalle_integrante_logistica', pk=pk)
    integrante.estado = nuevo_estado
    integrante.usuario_modificacion = request.user
    integrante.save(update_fields=['estado', 'usuario_modificacion', 'fecha_actualizacion'])
    messages.success(request, 'Estado actualizado correctamente.')
    return redirect('afiliados:detalle_integrante_logistica', pk=pk)


@login_required
def panorama_municipal_logistica(request):
    denied = _require_logistica(request)
    if denied:
        return denied

    comunidad_qs = Comunidad.objects.select_related('sector')
    sector_qs = Sector.objects.all()
    centro_qs = CentroVotacion.objects.all()
    coordinadores = CoordinadorLogistica.objects.all()
    lideres = LiderLogistica.objects.all()
    estructuras = EstructuraLogistica.objects.all()
    reuniones = ReunionLogistica.objects.all()
    incidencias = IncidenciaLogistica.objects.filter(estado=EstadoRegistro.ACTIVO)
    responsables = ResponsableLogistica.objects.filter(estado=EstadoRegistro.ACTIVO)

    comunidad_ids_cubiertas = set(CoordinadorLogistica.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(LiderLogistica.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(EstructuraLogistica.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(ResponsableLogistica.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True))
    sector_ids_cubiertos = set(CoordinadorLogistica.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True)) | set(EstructuraLogistica.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True)) | set(ResponsableLogistica.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True))
    centro_ids_cubiertos = set(CoordinadorLogistica.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True)) | set(EstructuraLogistica.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True)) | set(ResponsableLogistica.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True))

    crecimiento, crecimiento_series = _construir_series_crecimiento({
        'coordinadores': coordinadores,
        'lideres': lideres,
        'estructuras': estructuras,
        'reuniones': reuniones,
    })

    resumen_estructuras_por_comunidad = comunidad_qs.annotate(estructuras_total=Count('estructuras_logistica', distinct=True)).order_by('nombre')
    resumen_miembros_por_estructura = EstructuraLogistica.objects.select_related('comunidad').annotate(miembros_total=Count('integrantes', distinct=True)).order_by('nombre')

    context = {
        'total_comunidades': comunidad_qs.count(),
        'comunidades_cubiertas': len(comunidad_ids_cubiertas),
        'comunidades_no_cubiertas': max(comunidad_qs.count() - len(comunidad_ids_cubiertas), 0),
        'total_sectores': sector_qs.count(),
        'sectores_con_responsable': responsables.exclude(sector__isnull=True).values('sector_id').distinct().count(),
        'sectores_sin_responsable': max(sector_qs.count() - responsables.exclude(sector__isnull=True).values('sector_id').distinct().count(), 0),
        'total_centros': centro_qs.count(),
        'centros_con_cobertura': len(centro_ids_cubiertos),
        'centros_sin_cobertura': max(centro_qs.count() - len(centro_ids_cubiertos), 0),
        'coordinadores_activos': coordinadores.filter(estado=EstadoRegistro.ACTIVO).count(),
        'coordinadores_inactivos': coordinadores.filter(estado=EstadoRegistro.INACTIVO).count(),
        'total_lideres': lideres.count(),
        'estructuras_activas': estructuras.filter(estado=EstadoRegistro.ACTIVO).count(),
        'reuniones_mes': reuniones.filter(fecha__month=timezone.now().month, fecha__year=timezone.now().year).count(),
        'incidencias_abiertas': incidencias.count(),
        'resumen_comunidad': comunidad_qs.annotate(total_sectores=Count('sector', distinct=True), reuniones_total=Count('reuniones_logistica', distinct=True), incidencias_abiertas=Count('incidencias_logistica', filter=Q(incidencias_logistica__estado=EstadoRegistro.ACTIVO), distinct=True)),
        'resumen_sector': sector_qs.annotate(lideres_total=Count('lideres_logistica', distinct=True), estructuras_total=Count('estructuras_logistica', distinct=True), incidencias_total=Count('incidencias_logistica', distinct=True)),
        'resumen_centro': centro_qs.annotate(incidencias_total=Count('incidencias_logistica', distinct=True), reuniones_total=Count('reuniones_logistica', distinct=True)),
        'resumen_estructuras_por_comunidad': resumen_estructuras_por_comunidad,
        'resumen_miembros_por_estructura': resumen_miembros_por_estructura,
        'crecimiento': json.dumps(crecimiento, cls=DjangoJSONEncoder),
        'charts_data': json.dumps({'cobertura': {'comunidades': [len(comunidad_ids_cubiertas), max(comunidad_qs.count() - len(comunidad_ids_cubiertas), 0)], 'sectores': [len(sector_ids_cubiertos), max(sector_qs.count() - len(sector_ids_cubiertos), 0)], 'centros': [len(centro_ids_cubiertos), max(centro_qs.count() - len(centro_ids_cubiertos), 0)]}, 'estructura': {'coordinadores': [coordinadores.filter(estado=EstadoRegistro.ACTIVO).count(), coordinadores.filter(estado=EstadoRegistro.INACTIVO).count()], 'lideres_total': lideres.count(), 'estructuras_activas': estructuras.filter(estado=EstadoRegistro.ACTIVO).count()}, 'crecimiento': crecimiento_series}, cls=DjangoJSONEncoder),
    }
    return safe_render(request, 'afiliados/logistica/panorama.html', context)


def _log_crud(request, model, form_class, template, list_name, create_name, edit_name, pk=None, extra_context=None):
    denied = _require_logistica(request)
    if denied:
        return denied
    instance = get_object_or_404(model, pk=pk) if pk else None
    form = form_class(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if not getattr(obj, 'usuario_creador_id', None):
            obj.usuario_creador = request.user
        obj.usuario_modificador = request.user
        obj.save()
        messages.success(request, 'Registro guardado correctamente.')
        return redirect(list_name)
    items_qs = model.objects.all().order_by('-id')
    if template == 'reuniones':
        items_qs = items_qs.select_related('comunidad', 'sector', 'centro_votacion')
    context = {
        'form': form,
        'items': items_qs[:200],
        'entity_label': template,
        'create_name': create_name,
        'edit_name': edit_name,
        'detail_name': {
            'coordinadores': 'afiliados:detalle_coordinador_logistica',
            'lideres': 'afiliados:detalle_lider_logistica',
            'estructuras': 'afiliados:detalle_estructura_logistica',
            'responsables': 'afiliados:detalle_responsable_logistica',
            'reuniones': 'afiliados:detalle_reunion_logistica',
            'incidencias': 'afiliados:detalle_incidencia_logistica',
        }.get(template),
        'comunidad_lookup_url': reverse('afiliados:logistica_comunidad_lookup'),
        'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_logistica'),
    }
    if template == 'reuniones':
        responsables_modal = []
        for responsable in ResponsableLogistica.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'RESPONSABLE', 'tipo_label': 'Responsable', 'id': responsable.id, 'nombre': responsable.nombre_completo, 'dpi': responsable.dpi, 'comunidad': responsable.comunidad.nombre if responsable.comunidad else ''})
        for coordinador in CoordinadorLogistica.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'COORDINADOR', 'tipo_label': 'Coordinador', 'id': coordinador.id, 'nombre': coordinador.nombre_completo, 'dpi': coordinador.dpi, 'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else ''})
        for lider in LiderLogistica.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'LIDER', 'tipo_label': 'Enlace logístico', 'id': lider.id, 'nombre': lider.nombre_completo, 'dpi': lider.dpi, 'comunidad': lider.comunidad.nombre if lider.comunidad else ''})
        context['responsables_modal'] = responsables_modal
    if template == 'estructuras':
        coordinadores_modal = []
        for coordinador in CoordinadorLogistica.objects.select_related('comunidad').order_by('nombre_completo'):
            coordinadores_modal.append({'id': coordinador.id, 'nombre': coordinador.nombre_completo, 'dpi': coordinador.dpi, 'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else '', 'estado': coordinador.get_estado_display()})
        context['coordinadores_modal'] = coordinadores_modal
        coord_id = form['coordinador_responsable'].value()
        coord = CoordinadorLogistica.objects.filter(pk=coord_id).first() if coord_id else None
        context['coordinador_responsable_display'] = coord.nombre_completo if coord else ''
    if extra_context:
        context.update(extra_context)
    return safe_render(request, f'afiliados/logistica/crud_{template}.html', context)


@login_required
def lista_estructura_territorial_logistica(request):
    denied = _require_logistica(request)
    if denied:
        return denied
    comunidades = Comunidad.objects.select_related('sector').order_by('nombre')
    sectores = Sector.objects.order_by('nombre')
    centros = CentroVotacion.objects.order_by('nombre')
    return safe_render(request, 'afiliados/logistica/estructura_territorial.html', {'comunidades': comunidades, 'sectores': sectores, 'centros': centros, 'total_comunidades': comunidades.count(), 'total_sectores': sectores.count(), 'total_centros': centros.count()})


@login_required
def lista_coordinadores_logistica(request):
    return _log_crud(request, CoordinadorLogistica, CoordinadorLogisticaForm, 'coordinadores', 'afiliados:lista_coordinadores_logistica', 'afiliados:crear_coordinador_logistica', 'afiliados:editar_coordinador_logistica')


@login_required
def crear_coordinador_logistica(request):
    return _log_crud(request, CoordinadorLogistica, CoordinadorLogisticaForm, 'coordinadores', 'afiliados:lista_coordinadores_logistica', 'afiliados:crear_coordinador_logistica', 'afiliados:editar_coordinador_logistica')


@login_required
def editar_coordinador_logistica(request, pk):
    return _log_crud(request, CoordinadorLogistica, CoordinadorLogisticaForm, 'coordinadores', 'afiliados:lista_coordinadores_logistica', 'afiliados:crear_coordinador_logistica', 'afiliados:editar_coordinador_logistica', pk=pk)


@login_required
def detalle_coordinador_logistica(request, pk):
    denied = _require_logistica(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/logistica/detalle_generico.html', {'obj': get_object_or_404(CoordinadorLogistica, pk=pk), 'titulo': 'Coordinador logístico'})


@login_required
def lista_lideres_logistica(request):
    return _log_crud(request, LiderLogistica, LiderLogisticaForm, 'lideres', 'afiliados:lista_lideres_logistica', 'afiliados:crear_lider_logistica', 'afiliados:editar_lider_logistica')


@login_required
def crear_lider_logistica(request):
    return _log_crud(request, LiderLogistica, LiderLogisticaForm, 'lideres', 'afiliados:lista_lideres_logistica', 'afiliados:crear_lider_logistica', 'afiliados:editar_lider_logistica')


@login_required
def editar_lider_logistica(request, pk):
    return _log_crud(request, LiderLogistica, LiderLogisticaForm, 'lideres', 'afiliados:lista_lideres_logistica', 'afiliados:crear_lider_logistica', 'afiliados:editar_lider_logistica', pk=pk)


@login_required
def detalle_lider_logistica(request, pk):
    denied = _require_logistica(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/logistica/detalle_generico.html', {'obj': get_object_or_404(LiderLogistica, pk=pk), 'titulo': 'Enlace logístico'})


@login_required
def lista_estructuras_logistica(request):
    return _log_crud(request, EstructuraLogistica, EstructuraLogisticaForm, 'estructuras', 'afiliados:lista_estructuras_logistica', 'afiliados:crear_estructura_logistica', 'afiliados:editar_estructura_logistica')


@login_required
def crear_estructura_logistica(request):
    return _log_crud(request, EstructuraLogistica, EstructuraLogisticaForm, 'estructuras', 'afiliados:lista_estructuras_logistica', 'afiliados:crear_estructura_logistica', 'afiliados:editar_estructura_logistica')


@login_required
def editar_estructura_logistica(request, pk):
    return _log_crud(request, EstructuraLogistica, EstructuraLogisticaForm, 'estructuras', 'afiliados:lista_estructuras_logistica', 'afiliados:crear_estructura_logistica', 'afiliados:editar_estructura_logistica', pk=pk)


@login_required
def detalle_estructura_logistica(request, pk):
    denied = _require_logistica(request)
    if denied:
        return denied
    estructura = get_object_or_404(EstructuraLogistica.objects.select_related('comunidad', 'sector', 'centro_votacion', 'coordinador_responsable'), pk=pk)
    if request.method == 'POST':
        dpi = re.sub(r"\D", "", request.POST.get('dpi', ''))
        verified_dpi = re.sub(r"\D", "", request.POST.get('verified_dpi', ''))
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        if not dpi or verified_dpi != dpi:
            msg = 'Primero debe verificar empadronamiento del DPI antes de confirmar agregado.'
            return JsonResponse({'ok': False, 'message': msg}, status=400) if is_ajax else redirect('afiliados:detalle_estructura_logistica', pk=estructura.pk)
        payload, status_code = _resultado_padron_local(dpi)
        if status_code != 200 or not payload.get('ok') or not payload.get('found'):
            msg = payload.get('error') or payload.get('message') or 'No aparece empadronado en padrón local.'
            return JsonResponse({'ok': False, 'message': msg}, status=400) if is_ajax else redirect('afiliados:detalle_estructura_logistica', pk=estructura.pk)
        persona = payload.get('data') or {}
        if EstructuraIntegranteLogistica.objects.filter(estructura=estructura, dpi=dpi).exists():
            msg = 'Este integrante ya está registrado en la estructura logística.'
            return JsonResponse({'ok': False, 'message': msg}, status=409) if is_ajax else redirect('afiliados:detalle_estructura_logistica', pk=estructura.pk)
        comunidad = Comunidad.objects.filter(nombre=persona.get('comunidad')).first() if persona.get('comunidad') else None
        integrante = EstructuraIntegranteLogistica.objects.create(estructura=estructura, dpi=dpi, nombre_completo=persona.get('nombre_completo') or f"Integrante {dpi}", comunidad=comunidad, usuario_registro=request.user)
        if is_ajax:
            return JsonResponse({'ok': True, 'message': 'Integrante agregado correctamente a la estructura logística.', 'integrante': {'id': integrante.id, 'dpi': integrante.dpi, 'nombre_completo': integrante.nombre_completo, 'estado': integrante.estado, 'fecha_registro': timezone.localtime(integrante.fecha_registro).strftime('%d/%m/%Y %H:%M')}, 'total_integrantes': estructura.integrantes.count()})
        messages.success(request, 'Integrante agregado correctamente a la estructura logística.')
        return redirect('afiliados:detalle_estructura_logistica', pk=estructura.pk)

    integrantes = estructura.integrantes.select_related('usuario_registro').all()
    return safe_render(request, 'afiliados/logistica/detalle_estructura.html', {'estructura': estructura, 'integrantes': integrantes, 'total_integrantes': integrantes.count(), 'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_logistica')})


@login_required
@require_POST
def eliminar_integrante_estructura_logistica(request, pk, integrante_id):
    denied = _require_logistica(request)
    if denied:
        return denied
    estructura = get_object_or_404(EstructuraLogistica, pk=pk)
    integrante = get_object_or_404(EstructuraIntegranteLogistica, pk=integrante_id, estructura=estructura)
    integrante.delete()
    messages.success(request, 'Integrante eliminado correctamente de la estructura logística.')
    return redirect('afiliados:detalle_estructura_logistica', pk=estructura.pk)


@login_required
def lista_responsables_logistica(request):
    return _log_crud(request, ResponsableLogistica, ResponsableLogisticaForm, 'responsables', 'afiliados:lista_responsables_logistica', 'afiliados:crear_responsable_logistica', 'afiliados:editar_responsable_logistica')


@login_required
def crear_responsable_logistica(request):
    return _log_crud(request, ResponsableLogistica, ResponsableLogisticaForm, 'responsables', 'afiliados:lista_responsables_logistica', 'afiliados:crear_responsable_logistica', 'afiliados:editar_responsable_logistica')


@login_required
def editar_responsable_logistica(request, pk):
    return _log_crud(request, ResponsableLogistica, ResponsableLogisticaForm, 'responsables', 'afiliados:lista_responsables_logistica', 'afiliados:crear_responsable_logistica', 'afiliados:editar_responsable_logistica', pk=pk)


@login_required
def detalle_responsable_logistica(request, pk):
    denied = _require_logistica(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/logistica/detalle_generico.html', {'obj': get_object_or_404(ResponsableLogistica, pk=pk), 'titulo': 'Responsable logístico'})


@login_required
def lista_reuniones_logistica(request):
    return _log_crud(request, ReunionLogistica, ReunionLogisticaForm, 'reuniones', 'afiliados:lista_reuniones_logistica', 'afiliados:crear_reunion_logistica', 'afiliados:editar_reunion_logistica')


@login_required
def crear_reunion_logistica(request):
    return _log_crud(request, ReunionLogistica, ReunionLogisticaForm, 'reuniones', 'afiliados:lista_reuniones_logistica', 'afiliados:crear_reunion_logistica', 'afiliados:editar_reunion_logistica')


@login_required
def editar_reunion_logistica(request, pk):
    return _log_crud(request, ReunionLogistica, ReunionLogisticaForm, 'reuniones', 'afiliados:lista_reuniones_logistica', 'afiliados:crear_reunion_logistica', 'afiliados:editar_reunion_logistica', pk=pk)


@login_required
def detalle_reunion_logistica(request, pk):
    denied = _require_logistica(request)
    if denied:
        return denied
    reunion = get_object_or_404(ReunionLogistica, pk=pk)
    return safe_render(request, 'afiliados/logistica/detalle_reunion.html', {'reunion': reunion})


@login_required
def lista_incidencias_logistica(request):
    return _log_crud(request, IncidenciaLogistica, IncidenciaLogisticaForm, 'incidencias', 'afiliados:lista_incidencias_logistica', 'afiliados:crear_incidencia_logistica', 'afiliados:editar_incidencia_logistica')


@login_required
def crear_incidencia_logistica(request):
    return _log_crud(request, IncidenciaLogistica, IncidenciaLogisticaForm, 'incidencias', 'afiliados:lista_incidencias_logistica', 'afiliados:crear_incidencia_logistica', 'afiliados:editar_incidencia_logistica')


@login_required
def editar_incidencia_logistica(request, pk):
    return _log_crud(request, IncidenciaLogistica, IncidenciaLogisticaForm, 'incidencias', 'afiliados:lista_incidencias_logistica', 'afiliados:crear_incidencia_logistica', 'afiliados:editar_incidencia_logistica', pk=pk)


@login_required
def detalle_incidencia_logistica(request, pk):
    denied = _require_logistica(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/logistica/detalle_generico.html', {'obj': get_object_or_404(IncidenciaLogistica, pk=pk), 'titulo': 'Incidencia logística'})


@login_required
def reporte_cobertura_logistica(request):
    return panorama_municipal_logistica(request)


@login_required
def reporte_coordinadores_logistica(request):
    denied = _require_logistica(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/logistica/reporte_coordinadores.html', {'coordinadores': CoordinadorLogistica.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha_creacion')})


@login_required
def reporte_reuniones_logistica(request):
    denied = _require_logistica(request)
    if denied:
        return denied
    return safe_render(request, 'afiliados/logistica/reporte_reuniones.html', {'reuniones': ReunionLogistica.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha')})


@login_required
def reporte_crecimiento_mensual_logistica(request):
    return panorama_municipal_logistica(request)


@login_required
@require_GET
def logistica_comunidad_lookup(request):
    denied = _require_logistica(request)
    if denied:
        return denied
    comunidad_id = request.GET.get('comunidad_id')
    comunidad = Comunidad.objects.select_related('sector').filter(pk=comunidad_id).first()
    if not comunidad:
        return JsonResponse({'ok': False, 'error': 'Comunidad no encontrada.'}, status=404)
    centro = comunidad.sector.centros.first() if comunidad.sector else None
    return JsonResponse({'ok': True, 'data': {'comunidad': comunidad.nombre, 'sector': comunidad.sector.nombre if comunidad.sector else None, 'centro_votacion': centro.nombre if centro else None}})


def _logistica_extra_crud(request, model, form_class, template_name, list_url, titulo, extra=None, pk=None):
    denied = _require_logistica(request)
    if denied:
        return denied
    instance = get_object_or_404(model, pk=pk) if pk else None
    form = form_class(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if not getattr(obj, 'usuario_creador_id', None):
            obj.usuario_creador = request.user
        obj.usuario_modificador = request.user
        obj.save()
        messages.success(request, 'Registro guardado correctamente.')
        return redirect(list_url)
    context = {
        'form': form,
        'items': model.objects.all().order_by('-id')[:200],
        'titulo': titulo,
        'list_url': list_url,
        'edit_url_name': extra.get('edit_url_name') if extra else None,
        'columns': extra.get('columns', []) if extra else [],
    }
    return safe_render(request, template_name, context)


@login_required
def lista_recursos_logisticos(request):
    return _logistica_extra_crud(request, RecursoLogistico, RecursoLogisticoForm, 'afiliados/logistica/crud_recurso_logistico.html', 'afiliados:lista_recursos_logisticos', 'Recursos logísticos', extra={'edit_url_name': 'afiliados:editar_recurso_logistico', 'columns': ['nombre', 'tipo_recurso', 'estado', 'observaciones']})


@login_required
def editar_recurso_logistico(request, pk):
    return _logistica_extra_crud(request, RecursoLogistico, RecursoLogisticoForm, 'afiliados/logistica/crud_recurso_logistico.html', 'afiliados:lista_recursos_logisticos', 'Recursos logísticos', extra={'edit_url_name': 'afiliados:editar_recurso_logistico', 'columns': ['nombre', 'tipo_recurso', 'estado', 'observaciones']}, pk=pk)


@login_required
def lista_asignaciones_logisticas(request):
    return _logistica_extra_crud(request, AsignacionLogistica, AsignacionLogisticaForm, 'afiliados/logistica/crud_asignacion_logistica.html', 'afiliados:lista_asignaciones_logisticas', 'Asignaciones logísticas', extra={'edit_url_name': 'afiliados:editar_asignacion_logistica', 'columns': ['recurso', 'comunidad', 'responsable', 'fecha', 'estado']})


@login_required
def editar_asignacion_logistica(request, pk):
    return _logistica_extra_crud(request, AsignacionLogistica, AsignacionLogisticaForm, 'afiliados/logistica/crud_asignacion_logistica.html', 'afiliados:lista_asignaciones_logisticas', 'Asignaciones logísticas', extra={'edit_url_name': 'afiliados:editar_asignacion_logistica', 'columns': ['recurso', 'comunidad', 'responsable', 'fecha', 'estado']}, pk=pk)


@login_required
def lista_solicitudes_logisticas(request):
    return _logistica_extra_crud(request, SolicitudLogistica, SolicitudLogisticaForm, 'afiliados/logistica/crud_solicitud_logistica.html', 'afiliados:lista_solicitudes_logisticas', 'Solicitudes logísticas', extra={'edit_url_name': 'afiliados:editar_solicitud_logistica', 'columns': ['titulo', 'comunidad', 'responsable', 'prioridad', 'estado']})


@login_required
def editar_solicitud_logistica(request, pk):
    return _logistica_extra_crud(request, SolicitudLogistica, SolicitudLogisticaForm, 'afiliados/logistica/crud_solicitud_logistica.html', 'afiliados:lista_solicitudes_logisticas', 'Solicitudes logísticas', extra={'edit_url_name': 'afiliados:editar_solicitud_logistica', 'columns': ['titulo', 'comunidad', 'responsable', 'prioridad', 'estado']}, pk=pk)


@login_required
def lista_entregas_logisticas(request):
    return _logistica_extra_crud(request, EntregaLogistica, EntregaLogisticaForm, 'afiliados/logistica/crud_entrega_logistica.html', 'afiliados:lista_entregas_logisticas', 'Control de entregas', extra={'edit_url_name': 'afiliados:editar_entrega_logistica', 'columns': ['fecha', 'comunidad', 'responsable', 'detalle', 'estado']})


@login_required
def editar_entrega_logistica(request, pk):
    return _logistica_extra_crud(request, EntregaLogistica, EntregaLogisticaForm, 'afiliados/logistica/crud_entrega_logistica.html', 'afiliados:lista_entregas_logisticas', 'Control de entregas', extra={'edit_url_name': 'afiliados:editar_entrega_logistica', 'columns': ['fecha', 'comunidad', 'responsable', 'detalle', 'estado']}, pk=pk)


@login_required
def lista_agenda_logistica(request):
    return _logistica_extra_crud(request, AgendaLogistica, AgendaLogisticaForm, 'afiliados/logistica/crud_agenda_logistica.html', 'afiliados:lista_agenda_logistica', 'Agenda logística', extra={'edit_url_name': 'afiliados:editar_agenda_logistica', 'columns': ['actividad', 'comunidad', 'responsables', 'fecha_programada', 'estado']})


@login_required
def editar_agenda_logistica(request, pk):
    return _logistica_extra_crud(request, AgendaLogistica, AgendaLogisticaForm, 'afiliados/logistica/crud_agenda_logistica.html', 'afiliados:lista_agenda_logistica', 'Agenda logística', extra={'edit_url_name': 'afiliados:editar_agenda_logistica', 'columns': ['actividad', 'comunidad', 'responsables', 'fecha_programada', 'estado']}, pk=pk)


def _es_usuario_comunicacion(user):
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=['Comunicacion', 'Administrador']).exists()


def _require_comunicacion(request):
    if not _es_usuario_comunicacion(request.user):
        return HttpResponseForbidden("No tiene permisos para acceder a Secretaría de Comunicación y Redes.")
    return None


@login_required
def dashboard_comunicacion(request):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    registros = ComunicacionIntegrante.objects.select_related('afiliado')
    context = {
        'total_integrantes': registros.count(),
        'total_pendientes': registros.filter(estado=ComunicacionIntegrante.Estado.PENDIENTE).count(),
        'total_revisados': registros.filter(estado=ComunicacionIntegrante.Estado.REVISADO).count(),
        'total_aprobados': registros.filter(estado=ComunicacionIntegrante.Estado.APROBADO).count(),
        'ultimos_registros': registros.order_by('-fecha_creacion')[:10],
    }
    return safe_render(request, 'afiliados/comunicacion/dashboard.html', context)


@login_required
def lista_integrantes_comunicacion(request):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    integrantes = ComunicacionIntegrante.objects.select_related('afiliado', 'usuario_registro').order_by('-fecha_creacion')
    context = {'integrantes': integrantes, 'total_integrantes': integrantes.count(), 'total_pendientes': integrantes.filter(estado=ComunicacionIntegrante.Estado.PENDIENTE).count(), 'total_revisados': integrantes.filter(estado=ComunicacionIntegrante.Estado.REVISADO).count(), 'total_aprobados': integrantes.filter(estado=ComunicacionIntegrante.Estado.APROBADO).count()}
    return safe_render(request, 'afiliados/comunicacion/lista.html', context)


@login_required
def crear_integrante_comunicacion(request):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    form = ComunicacionIntegranteForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])
        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif ComunicacionIntegrante.objects.filter(afiliado=afiliado).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de Comunicación y Redes.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_registro = request.user
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante agregado correctamente.')
            return redirect('afiliados:lista_integrantes_comunicacion')
    return safe_render(request, 'afiliados/comunicacion/form.html', {'form': form, 'es_edicion': False})


@login_required
def detalle_integrante_comunicacion(request, pk):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    integrante = get_object_or_404(ComunicacionIntegrante.objects.select_related('afiliado', 'usuario_registro', 'usuario_modificacion'), pk=pk)
    return safe_render(request, 'afiliados/comunicacion/detalle.html', {'integrante': integrante})


@login_required
def editar_integrante_comunicacion(request, pk):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    integrante = get_object_or_404(ComunicacionIntegrante, pk=pk)
    form = ComunicacionIntegranteForm(request.POST or None, instance=integrante)
    form.fields['dpi'].initial = integrante.afiliado.dpi
    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])
        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif ComunicacionIntegrante.objects.filter(afiliado=afiliado).exclude(pk=integrante.pk).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de Comunicación y Redes.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante actualizado correctamente.')
            return redirect('afiliados:detalle_integrante_comunicacion', pk=integrante.pk)
    return safe_render(request, 'afiliados/comunicacion/form.html', {'form': form, 'es_edicion': True, 'integrante': integrante})


@login_required
@require_GET
def buscar_por_dpi_comunicacion(request):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    afiliado, dpi_limpio, error_dpi = _afiliado_por_dpi(request.GET.get('dpi', ''))
    if error_dpi:
        return JsonResponse({'ok': False, 'error': error_dpi}, status=400)
    if not afiliado:
        return JsonResponse({'ok': True, 'exists': False, 'message': 'No existe afiliado con ese DPI.'})
    ya_registrado = ComunicacionIntegrante.objects.filter(afiliado=afiliado).exists()
    return JsonResponse({'ok': True, 'exists': True, 'already_registered': ya_registrado, 'data': {'dpi': dpi_limpio, 'nombre_completo': afiliado.nombre_completo, 'telefono': afiliado.telefono, 'direccion': afiliado.direccion, 'comunidad': afiliado.comunidad.nombre if afiliado.comunidad else None, 'empadronado': afiliado.empadronado}})


@login_required
@require_GET
def verificar_empadronamiento_comunicacion(request):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    payload, status_code = _resultado_padron_local(request.GET.get("dpi", ""))
    return JsonResponse(payload, status=status_code)


@login_required
@require_POST
def cambiar_estado_comunicacion(request, pk):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    integrante = get_object_or_404(ComunicacionIntegrante, pk=pk)
    nuevo_estado = request.POST.get('estado', '').strip().upper()
    estados_validos = {choice[0] for choice in ComunicacionIntegrante.Estado.choices}
    if nuevo_estado not in estados_validos:
        messages.error(request, 'Estado inválido.')
        return redirect('afiliados:detalle_integrante_comunicacion', pk=pk)
    integrante.estado = nuevo_estado
    integrante.usuario_modificacion = request.user
    integrante.save(update_fields=['estado', 'usuario_modificacion', 'fecha_actualizacion'])
    messages.success(request, 'Estado actualizado correctamente.')
    return redirect('afiliados:detalle_integrante_comunicacion', pk=pk)


@login_required
def panorama_municipal_comunicacion(request):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    comunidad_qs = Comunidad.objects.select_related('sector')
    sector_qs = Sector.objects.all()
    centro_qs = CentroVotacion.objects.all()
    coordinadores = CoordinadorComunicacion.objects.all()
    lideres = LiderComunicacion.objects.all()
    estructuras = EstructuraComunicacion.objects.all()
    reuniones = ReunionComunicacion.objects.all()
    incidencias = IncidenciaComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO)
    responsables = ResponsableComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO)
    comunidad_ids_cubiertas = set(CoordinadorComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(LiderComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(EstructuraComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True)) | set(ResponsableComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO, comunidad__isnull=False).values_list('comunidad_id', flat=True))
    sector_ids_cubiertos = set(CoordinadorComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True)) | set(EstructuraComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True)) | set(ResponsableComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO, sector__isnull=False).values_list('sector_id', flat=True))
    centro_ids_cubiertos = set(CoordinadorComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True)) | set(EstructuraComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True)) | set(ResponsableComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO, centro_votacion__isnull=False).values_list('centro_votacion_id', flat=True))
    crecimiento, crecimiento_series = _construir_series_crecimiento({
        'coordinadores': coordinadores,
        'lideres': lideres,
        'estructuras': estructuras,
        'reuniones': reuniones,
    })
    context = {'total_comunidades': comunidad_qs.count(), 'comunidades_cubiertas': len(comunidad_ids_cubiertas), 'comunidades_no_cubiertas': max(comunidad_qs.count() - len(comunidad_ids_cubiertas), 0), 'total_sectores': sector_qs.count(), 'sectores_con_responsable': responsables.exclude(sector__isnull=True).values('sector_id').distinct().count(), 'sectores_sin_responsable': max(sector_qs.count() - responsables.exclude(sector__isnull=True).values('sector_id').distinct().count(), 0), 'total_centros': centro_qs.count(), 'centros_con_cobertura': len(centro_ids_cubiertos), 'centros_sin_cobertura': max(centro_qs.count() - len(centro_ids_cubiertos), 0), 'coordinadores_activos': coordinadores.filter(estado=EstadoRegistro.ACTIVO).count(), 'coordinadores_inactivos': coordinadores.filter(estado=EstadoRegistro.INACTIVO).count(), 'total_lideres': lideres.count(), 'estructuras_activas': estructuras.filter(estado=EstadoRegistro.ACTIVO).count(), 'reuniones_mes': reuniones.filter(fecha__month=timezone.now().month, fecha__year=timezone.now().year).count(), 'incidencias_abiertas': incidencias.count(), 'resumen_comunidad': comunidad_qs.annotate(total_sectores=Count('sector', distinct=True), reuniones_total=Count('reuniones_comunicacion', distinct=True), incidencias_abiertas=Count('incidencias_comunicacion', filter=Q(incidencias_comunicacion__estado=EstadoRegistro.ACTIVO), distinct=True)), 'resumen_sector': sector_qs.annotate(lideres_total=Count('lideres_comunicacion', distinct=True), estructuras_total=Count('estructuras_comunicacion', distinct=True), incidencias_total=Count('incidencias_comunicacion', distinct=True)), 'resumen_centro': centro_qs.annotate(incidencias_total=Count('incidencias_comunicacion', distinct=True), reuniones_total=Count('reuniones_comunicacion', distinct=True)), 'crecimiento': json.dumps(crecimiento, cls=DjangoJSONEncoder), 'charts_data': json.dumps({'cobertura': {'comunidades': [len(comunidad_ids_cubiertas), max(comunidad_qs.count() - len(comunidad_ids_cubiertas), 0)], 'sectores': [len(sector_ids_cubiertos), max(sector_qs.count() - len(sector_ids_cubiertos), 0)], 'centros': [len(centro_ids_cubiertos), max(centro_qs.count() - len(centro_ids_cubiertos), 0)]}, 'estructura': {'coordinadores': [coordinadores.filter(estado=EstadoRegistro.ACTIVO).count(), coordinadores.filter(estado=EstadoRegistro.INACTIVO).count()], 'lideres_total': lideres.count(), 'estructuras_activas': estructuras.filter(estado=EstadoRegistro.ACTIVO).count()}, 'crecimiento': crecimiento_series}, cls=DjangoJSONEncoder)}
    return safe_render(request, 'afiliados/comunicacion/panorama.html', context)


def _com_crud(request, model, form_class, template, list_name, create_name, edit_name, pk=None):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    instance = get_object_or_404(model, pk=pk) if pk else None
    form = form_class(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if not getattr(obj, 'usuario_creador_id', None):
            obj.usuario_creador = request.user
        obj.usuario_modificador = request.user
        obj.save()
        messages.success(request, 'Registro guardado correctamente.')
        return redirect(list_name)
    context = {'form': form, 'items': model.objects.all().order_by('-id')[:200], 'entity_label': template, 'create_name': create_name, 'edit_name': edit_name, 'detail_name': {'coordinadores': 'afiliados:detalle_coordinador_comunicacion', 'lideres': 'afiliados:detalle_lider_comunicacion', 'estructuras': 'afiliados:detalle_estructura_comunicacion', 'responsables': 'afiliados:detalle_responsable_comunicacion', 'reuniones': 'afiliados:detalle_reunion_comunicacion', 'incidencias': 'afiliados:detalle_incidencia_comunicacion'}.get(template), 'comunidad_lookup_url': reverse('afiliados:comunicacion_comunidad_lookup'), 'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_comunicacion')}
    if template == 'reuniones':
        responsables_modal = []
        for responsable in ResponsableComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'RESPONSABLE', 'tipo_label': 'Responsable', 'id': responsable.id, 'nombre': responsable.nombre_completo, 'dpi': responsable.dpi, 'comunidad': responsable.comunidad.nombre if responsable.comunidad else ''})
        for coordinador in CoordinadorComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'COORDINADOR', 'tipo_label': 'Coordinador', 'id': coordinador.id, 'nombre': coordinador.nombre_completo, 'dpi': coordinador.dpi, 'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else ''})
        for lider in LiderComunicacion.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'LIDER', 'tipo_label': 'Enlace de comunicación', 'id': lider.id, 'nombre': lider.nombre_completo, 'dpi': lider.dpi, 'comunidad': lider.comunidad.nombre if lider.comunidad else ''})
        context['responsables_modal'] = responsables_modal
    if template == 'estructuras':
        coordinadores_modal = []
        for coordinador in CoordinadorComunicacion.objects.select_related('comunidad').order_by('nombre_completo'):
            coordinadores_modal.append({'id': coordinador.id, 'nombre': coordinador.nombre_completo, 'dpi': coordinador.dpi, 'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else '', 'estado': coordinador.get_estado_display()})
        context['coordinadores_modal'] = coordinadores_modal
        coord_id = form['coordinador_responsable'].value()
        coord = CoordinadorComunicacion.objects.filter(pk=coord_id).first() if coord_id else None
        context['coordinador_responsable_display'] = coord.nombre_completo if coord else ''
    return safe_render(request, f'afiliados/comunicacion/crud_{template}.html', context)


@login_required
def lista_estructura_territorial_comunicacion(request):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    comunidades = Comunidad.objects.select_related('sector').order_by('nombre')
    sectores = Sector.objects.order_by('nombre')
    centros = CentroVotacion.objects.order_by('nombre')
    return safe_render(request, 'afiliados/comunicacion/estructura_territorial.html', {'comunidades': comunidades, 'sectores': sectores, 'centros': centros, 'total_comunidades': comunidades.count(), 'total_sectores': sectores.count(), 'total_centros': centros.count()})


@login_required
def lista_coordinadores_comunicacion(request): return _com_crud(request, CoordinadorComunicacion, CoordinadorComunicacionForm, 'coordinadores', 'afiliados:lista_coordinadores_comunicacion', 'afiliados:crear_coordinador_comunicacion', 'afiliados:editar_coordinador_comunicacion')
@login_required
def crear_coordinador_comunicacion(request): return _com_crud(request, CoordinadorComunicacion, CoordinadorComunicacionForm, 'coordinadores', 'afiliados:lista_coordinadores_comunicacion', 'afiliados:crear_coordinador_comunicacion', 'afiliados:editar_coordinador_comunicacion')
@login_required
def editar_coordinador_comunicacion(request, pk): return _com_crud(request, CoordinadorComunicacion, CoordinadorComunicacionForm, 'coordinadores', 'afiliados:lista_coordinadores_comunicacion', 'afiliados:crear_coordinador_comunicacion', 'afiliados:editar_coordinador_comunicacion', pk=pk)
@login_required
def detalle_coordinador_comunicacion(request, pk):
    denied = _require_comunicacion(request)
    if denied: return denied
    return safe_render(request, 'afiliados/comunicacion/detalle_generico.html', {'obj': get_object_or_404(CoordinadorComunicacion, pk=pk), 'titulo': 'Coordinador de comunicación'})
@login_required
def lista_lideres_comunicacion(request): return _com_crud(request, LiderComunicacion, LiderComunicacionForm, 'lideres', 'afiliados:lista_lideres_comunicacion', 'afiliados:crear_lider_comunicacion', 'afiliados:editar_lider_comunicacion')
@login_required
def crear_lider_comunicacion(request): return _com_crud(request, LiderComunicacion, LiderComunicacionForm, 'lideres', 'afiliados:lista_lideres_comunicacion', 'afiliados:crear_lider_comunicacion', 'afiliados:editar_lider_comunicacion')
@login_required
def editar_lider_comunicacion(request, pk): return _com_crud(request, LiderComunicacion, LiderComunicacionForm, 'lideres', 'afiliados:lista_lideres_comunicacion', 'afiliados:crear_lider_comunicacion', 'afiliados:editar_lider_comunicacion', pk=pk)
@login_required
def detalle_lider_comunicacion(request, pk):
    denied = _require_comunicacion(request)
    if denied: return denied
    return safe_render(request, 'afiliados/comunicacion/detalle_generico.html', {'obj': get_object_or_404(LiderComunicacion, pk=pk), 'titulo': 'Enlace de comunicación'})
@login_required
def lista_estructuras_comunicacion(request): return _com_crud(request, EstructuraComunicacion, EstructuraComunicacionForm, 'estructuras', 'afiliados:lista_estructuras_comunicacion', 'afiliados:crear_estructura_comunicacion', 'afiliados:editar_estructura_comunicacion')
@login_required
def crear_estructura_comunicacion(request): return _com_crud(request, EstructuraComunicacion, EstructuraComunicacionForm, 'estructuras', 'afiliados:lista_estructuras_comunicacion', 'afiliados:crear_estructura_comunicacion', 'afiliados:editar_estructura_comunicacion')
@login_required
def editar_estructura_comunicacion(request, pk): return _com_crud(request, EstructuraComunicacion, EstructuraComunicacionForm, 'estructuras', 'afiliados:lista_estructuras_comunicacion', 'afiliados:crear_estructura_comunicacion', 'afiliados:editar_estructura_comunicacion', pk=pk)


@login_required
def detalle_estructura_comunicacion(request, pk):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    estructura = get_object_or_404(EstructuraComunicacion.objects.select_related('comunidad', 'sector', 'centro_votacion', 'coordinador_responsable'), pk=pk)
    if request.method == 'POST':
        dpi = re.sub(r"\D", "", request.POST.get('dpi', ''))
        verified_dpi = re.sub(r"\D", "", request.POST.get('verified_dpi', ''))
        if not dpi or verified_dpi != dpi:
            messages.error(request, 'Primero debe verificar empadronamiento del DPI antes de confirmar agregado.')
            return redirect('afiliados:detalle_estructura_comunicacion', pk=estructura.pk)
        payload, status_code = _resultado_padron_local(dpi)
        if status_code != 200 or not payload.get('ok') or not payload.get('found'):
            messages.error(request, payload.get('error') or payload.get('message') or 'No aparece empadronado en padrón local.')
            return redirect('afiliados:detalle_estructura_comunicacion', pk=estructura.pk)
        persona = payload.get('data') or {}
        if EstructuraIntegranteComunicacion.objects.filter(estructura=estructura, dpi=dpi).exists():
            messages.warning(request, 'Este integrante ya está registrado en la estructura de comunicación.')
            return redirect('afiliados:detalle_estructura_comunicacion', pk=estructura.pk)
        comunidad = Comunidad.objects.filter(nombre=persona.get('comunidad')).first() if persona.get('comunidad') else None
        EstructuraIntegranteComunicacion.objects.create(estructura=estructura, dpi=dpi, nombre_completo=persona.get('nombre_completo') or f"Integrante {dpi}", comunidad=comunidad, usuario_registro=request.user)
        messages.success(request, 'Integrante agregado correctamente a la estructura de comunicación.')
        return redirect('afiliados:detalle_estructura_comunicacion', pk=estructura.pk)
    integrantes = estructura.integrantes.select_related('usuario_registro').all()
    return safe_render(request, 'afiliados/comunicacion/detalle_estructura.html', {'estructura': estructura, 'integrantes': integrantes, 'total_integrantes': integrantes.count(), 'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_comunicacion')})


@login_required
@require_POST
def eliminar_integrante_estructura_comunicacion(request, pk, integrante_id):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    estructura = get_object_or_404(EstructuraComunicacion, pk=pk)
    integrante = get_object_or_404(EstructuraIntegranteComunicacion, pk=integrante_id, estructura=estructura)
    integrante.delete()
    messages.success(request, 'Integrante eliminado correctamente de la estructura de comunicación.')
    return redirect('afiliados:detalle_estructura_comunicacion', pk=estructura.pk)

@login_required
def lista_responsables_comunicacion(request): return _com_crud(request, ResponsableComunicacion, ResponsableComunicacionForm, 'responsables', 'afiliados:lista_responsables_comunicacion', 'afiliados:crear_responsable_comunicacion', 'afiliados:editar_responsable_comunicacion')
@login_required
def crear_responsable_comunicacion(request): return _com_crud(request, ResponsableComunicacion, ResponsableComunicacionForm, 'responsables', 'afiliados:lista_responsables_comunicacion', 'afiliados:crear_responsable_comunicacion', 'afiliados:editar_responsable_comunicacion')
@login_required
def editar_responsable_comunicacion(request, pk): return _com_crud(request, ResponsableComunicacion, ResponsableComunicacionForm, 'responsables', 'afiliados:lista_responsables_comunicacion', 'afiliados:crear_responsable_comunicacion', 'afiliados:editar_responsable_comunicacion', pk=pk)
@login_required
def detalle_responsable_comunicacion(request, pk):
    denied = _require_comunicacion(request)
    if denied: return denied
    return safe_render(request, 'afiliados/comunicacion/detalle_generico.html', {'obj': get_object_or_404(ResponsableComunicacion, pk=pk), 'titulo': 'Responsable de comunicación'})
@login_required
def lista_reuniones_comunicacion(request): return _com_crud(request, ReunionComunicacion, ReunionComunicacionForm, 'reuniones', 'afiliados:lista_reuniones_comunicacion', 'afiliados:crear_reunion_comunicacion', 'afiliados:editar_reunion_comunicacion')
@login_required
def crear_reunion_comunicacion(request): return _com_crud(request, ReunionComunicacion, ReunionComunicacionForm, 'reuniones', 'afiliados:lista_reuniones_comunicacion', 'afiliados:crear_reunion_comunicacion', 'afiliados:editar_reunion_comunicacion')
@login_required
def editar_reunion_comunicacion(request, pk): return _com_crud(request, ReunionComunicacion, ReunionComunicacionForm, 'reuniones', 'afiliados:lista_reuniones_comunicacion', 'afiliados:crear_reunion_comunicacion', 'afiliados:editar_reunion_comunicacion', pk=pk)
@login_required
def detalle_reunion_comunicacion(request, pk):
    denied = _require_comunicacion(request)
    if denied: return denied
    return safe_render(request, 'afiliados/comunicacion/detalle_reunion.html', {'reunion': get_object_or_404(ReunionComunicacion, pk=pk)})
@login_required
def lista_incidencias_comunicacion(request): return _com_crud(request, IncidenciaComunicacion, IncidenciaComunicacionForm, 'incidencias', 'afiliados:lista_incidencias_comunicacion', 'afiliados:crear_incidencia_comunicacion', 'afiliados:editar_incidencia_comunicacion')
@login_required
def crear_incidencia_comunicacion(request): return _com_crud(request, IncidenciaComunicacion, IncidenciaComunicacionForm, 'incidencias', 'afiliados:lista_incidencias_comunicacion', 'afiliados:crear_incidencia_comunicacion', 'afiliados:editar_incidencia_comunicacion')
@login_required
def editar_incidencia_comunicacion(request, pk): return _com_crud(request, IncidenciaComunicacion, IncidenciaComunicacionForm, 'incidencias', 'afiliados:lista_incidencias_comunicacion', 'afiliados:crear_incidencia_comunicacion', 'afiliados:editar_incidencia_comunicacion', pk=pk)
@login_required
def detalle_incidencia_comunicacion(request, pk):
    denied = _require_comunicacion(request)
    if denied: return denied
    return safe_render(request, 'afiliados/comunicacion/detalle_generico.html', {'obj': get_object_or_404(IncidenciaComunicacion, pk=pk), 'titulo': 'Incidencia de comunicación'})
@login_required
def reporte_cobertura_comunicacion(request): return panorama_municipal_comunicacion(request)
@login_required
def reporte_coordinadores_comunicacion(request):
    denied = _require_comunicacion(request)
    if denied: return denied
    return safe_render(request, 'afiliados/comunicacion/reporte_coordinadores.html', {'coordinadores': CoordinadorComunicacion.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha_creacion')})
@login_required
def reporte_reuniones_comunicacion(request):
    denied = _require_comunicacion(request)
    if denied: return denied
    return safe_render(request, 'afiliados/comunicacion/reporte_reuniones.html', {'reuniones': ReunionComunicacion.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha')})
@login_required
def reporte_crecimiento_mensual_comunicacion(request): return panorama_municipal_comunicacion(request)


@login_required
@require_GET
def comunicacion_comunidad_lookup(request):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    comunidad = Comunidad.objects.select_related('sector').filter(pk=request.GET.get('comunidad_id')).first()
    if not comunidad:
        return JsonResponse({'ok': False, 'error': 'Comunidad no encontrada.'}, status=404)
    centro = comunidad.sector.centros.first() if comunidad.sector else None
    return JsonResponse({'ok': True, 'data': {'comunidad': comunidad.nombre, 'sector': comunidad.sector.nombre if comunidad.sector else None, 'centro_votacion': centro.nombre if centro else None}})


def _com_extra_crud(request, model, form_class, template_name, list_url, pk=None):
    denied = _require_comunicacion(request)
    if denied:
        return denied
    instance = get_object_or_404(model, pk=pk) if pk else None
    form = form_class(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if not getattr(obj, 'usuario_creador_id', None):
            obj.usuario_creador = request.user
        obj.usuario_modificador = request.user
        obj.save()
        messages.success(request, 'Registro guardado correctamente.')
        return redirect(list_url)
    return safe_render(request, template_name, {'form': form, 'items': model.objects.all().order_by('-id')[:200]})


@login_required
def lista_agenda_publicaciones(request): return _com_extra_crud(request, AgendaPublicacion, AgendaPublicacionForm, 'afiliados/comunicacion/crud_agenda_publicaciones.html', 'afiliados:lista_agenda_publicaciones')
@login_required
def editar_agenda_publicaciones(request, pk): return _com_extra_crud(request, AgendaPublicacion, AgendaPublicacionForm, 'afiliados/comunicacion/crud_agenda_publicaciones.html', 'afiliados:lista_agenda_publicaciones', pk=pk)
@login_required
def lista_solicitudes_contenido(request): return _com_extra_crud(request, SolicitudContenido, SolicitudContenidoForm, 'afiliados/comunicacion/crud_solicitudes_contenido.html', 'afiliados:lista_solicitudes_contenido')
@login_required
def editar_solicitudes_contenido(request, pk): return _com_extra_crud(request, SolicitudContenido, SolicitudContenidoForm, 'afiliados/comunicacion/crud_solicitudes_contenido.html', 'afiliados:lista_solicitudes_contenido', pk=pk)
@login_required
def lista_cobertura_actividades(request): return _com_extra_crud(request, CoberturaActividad, CoberturaActividadForm, 'afiliados/comunicacion/crud_cobertura_actividades.html', 'afiliados:lista_cobertura_actividades')
@login_required
def editar_cobertura_actividades(request, pk): return _com_extra_crud(request, CoberturaActividad, CoberturaActividadForm, 'afiliados/comunicacion/crud_cobertura_actividades.html', 'afiliados:lista_cobertura_actividades', pk=pk)
@login_required
def lista_banco_medios(request): return _com_extra_crud(request, BancoMedios, BancoMediosForm, 'afiliados/comunicacion/crud_banco_medios.html', 'afiliados:lista_banco_medios')
@login_required
def editar_banco_medios(request, pk): return _com_extra_crud(request, BancoMedios, BancoMediosForm, 'afiliados/comunicacion/crud_banco_medios.html', 'afiliados:lista_banco_medios', pk=pk)
@login_required
def lista_campanas_comunicacion(request): return _com_extra_crud(request, CampanaComunicacion, CampanaComunicacionForm, 'afiliados/comunicacion/crud_campanas_comunicacion.html', 'afiliados:lista_campanas_comunicacion')
@login_required
def editar_campanas_comunicacion(request, pk): return _com_extra_crud(request, CampanaComunicacion, CampanaComunicacionForm, 'afiliados/comunicacion/crud_campanas_comunicacion.html', 'afiliados:lista_campanas_comunicacion', pk=pk)
@login_required
def lista_monitoreo_redes(request): return _com_extra_crud(request, MonitoreoRed, MonitoreoRedForm, 'afiliados/comunicacion/crud_monitoreo_redes.html', 'afiliados:lista_monitoreo_redes')
@login_required
def editar_monitoreo_redes(request, pk): return _com_extra_crud(request, MonitoreoRed, MonitoreoRedForm, 'afiliados/comunicacion/crud_monitoreo_redes.html', 'afiliados:lista_monitoreo_redes', pk=pk)



def _es_usuario_plan_hormiga(user):
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=['PlanHormiga', 'Administrador']).exists()


def _require_plan_hormiga(request):
    if not _es_usuario_plan_hormiga(request.user):
        return HttpResponseForbidden("No tiene permisos para acceder a Secretaría de Plan Hormiga.")
    return None


@login_required
def dashboard_plan_hormiga(request):
    denied = _require_plan_hormiga(request)
    if denied:
        return denied
    registros = PlanHormigaIntegrante.objects.select_related('afiliado')
    context = {
        'total_integrantes': registros.count(),
        'total_pendientes': registros.filter(estado=PlanHormigaIntegrante.Estado.PENDIENTE).count(),
        'total_revisados': registros.filter(estado=PlanHormigaIntegrante.Estado.REVISADO).count(),
        'total_aprobados': registros.filter(estado=PlanHormigaIntegrante.Estado.APROBADO).count(),
        'ultimos_registros': registros.order_by('-fecha_creacion')[:10],
    }
    return safe_render(request, 'afiliados/plan_hormiga/dashboard.html', context)


@login_required
def lista_integrantes_plan_hormiga(request):
    denied = _require_plan_hormiga(request)
    if denied:
        return denied
    integrantes = PlanHormigaIntegrante.objects.select_related('afiliado', 'usuario_registro').order_by('-fecha_creacion')
    context = {'integrantes': integrantes, 'total_integrantes': integrantes.count(), 'total_pendientes': integrantes.filter(estado=PlanHormigaIntegrante.Estado.PENDIENTE).count(), 'total_revisados': integrantes.filter(estado=PlanHormigaIntegrante.Estado.REVISADO).count(), 'total_aprobados': integrantes.filter(estado=PlanHormigaIntegrante.Estado.APROBADO).count()}
    return safe_render(request, 'afiliados/plan_hormiga/lista.html', context)


@login_required
def crear_integrante_plan_hormiga(request):
    denied = _require_plan_hormiga(request)
    if denied:
        return denied
    form = PlanHormigaIntegranteForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])
        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif PlanHormigaIntegrante.objects.filter(afiliado=afiliado).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de Plan Hormiga.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_registro = request.user
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante agregado correctamente.')
            return redirect('afiliados:lista_integrantes_plan_hormiga')
    return safe_render(request, 'afiliados/plan_hormiga/form.html', {'form': form, 'es_edicion': False})


@login_required
def detalle_integrante_plan_hormiga(request, pk):
    denied = _require_plan_hormiga(request)
    if denied:
        return denied
    integrante = get_object_or_404(PlanHormigaIntegrante.objects.select_related('afiliado', 'usuario_registro', 'usuario_modificacion'), pk=pk)
    return safe_render(request, 'afiliados/plan_hormiga/detalle.html', {'integrante': integrante})


@login_required
def editar_integrante_plan_hormiga(request, pk):
    denied = _require_plan_hormiga(request)
    if denied:
        return denied
    integrante = get_object_or_404(PlanHormigaIntegrante, pk=pk)
    form = PlanHormigaIntegranteForm(request.POST or None, instance=integrante)
    form.fields['dpi'].initial = integrante.afiliado.dpi
    if request.method == 'POST' and form.is_valid():
        afiliado, _dpi, error_dpi = _afiliado_por_dpi(form.cleaned_data['dpi'])
        if error_dpi:
            form.add_error('dpi', error_dpi)
        elif not afiliado:
            form.add_error('dpi', 'No existe un afiliado registrado con este DPI.')
        elif PlanHormigaIntegrante.objects.filter(afiliado=afiliado).exclude(pk=integrante.pk).exists():
            form.add_error('dpi', 'Este afiliado ya está registrado en Secretaría de Plan Hormiga.')
        else:
            integrante = form.save(commit=False)
            integrante.afiliado = afiliado
            integrante.usuario_modificacion = request.user
            integrante.save()
            messages.success(request, 'Integrante actualizado correctamente.')
            return redirect('afiliados:detalle_integrante_plan_hormiga', pk=integrante.pk)
    return safe_render(request, 'afiliados/plan_hormiga/form.html', {'form': form, 'es_edicion': True, 'integrante': integrante})

@login_required
@require_GET
def buscar_por_dpi_plan_hormiga(request):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    afiliado, dpi_limpio, error_dpi = _afiliado_por_dpi(request.GET.get('dpi', ''))
    if error_dpi: return JsonResponse({'ok': False, 'error': error_dpi}, status=400)
    if not afiliado: return JsonResponse({'ok': True, 'exists': False, 'message': 'No existe afiliado con ese DPI.'})
    ya_registrado = PlanHormigaIntegrante.objects.filter(afiliado=afiliado).exists()
    return JsonResponse({'ok': True, 'exists': True, 'already_registered': ya_registrado, 'data': {'dpi': dpi_limpio, 'nombre_completo': afiliado.nombre_completo, 'telefono': afiliado.telefono, 'direccion': afiliado.direccion, 'comunidad': afiliado.comunidad.nombre if afiliado.comunidad else None, 'empadronado': afiliado.empadronado}})

@login_required
@require_POST
def cambiar_estado_plan_hormiga(request, pk):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    integrante = get_object_or_404(PlanHormigaIntegrante, pk=pk)
    nuevo_estado = request.POST.get('estado', '').upper().strip()
    estados_validos = {choice[0] for choice in PlanHormigaIntegrante.Estado.choices}
    if nuevo_estado not in estados_validos:
        messages.error(request, 'Estado inválido.')
        return redirect('afiliados:detalle_integrante_plan_hormiga', pk=pk)
    integrante.estado = nuevo_estado
    integrante.usuario_modificacion = request.user
    integrante.save(update_fields=['estado', 'usuario_modificacion', 'fecha_actualizacion'])
    messages.success(request, 'Estado actualizado correctamente.')
    return redirect('afiliados:detalle_integrante_plan_hormiga', pk=pk)

@login_required
@require_GET
def verificar_empadronamiento_plan_hormiga(request):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    dpi = re.sub(r"\D", "", request.GET.get('dpi', ''))
    payload, status_code = _resultado_padron_local(dpi)
    return JsonResponse(payload, status=status_code)

@login_required
def panorama_municipal_plan_hormiga(request):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    return safe_render(request, 'afiliados/plan_hormiga/panorama.html', {'coordinadores': CoordinadorPlanHormiga.objects.count(), 'enlaces': EnlaceTerritorialPlanHormiga.objects.count(), 'estructuras': EstructuraPlanHormiga.objects.count(), 'reuniones': ReunionPlanHormiga.objects.count(), 'incidencias': IncidenciaPlanHormiga.objects.count()})


def _plan_hormiga_crud(request, model, form_class, template, list_name, create_name, edit_name, pk=None):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    instance = get_object_or_404(model, pk=pk) if pk else None
    form = form_class(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if hasattr(obj, 'usuario_creador_id') and not getattr(obj, 'usuario_creador_id', None):
            obj.usuario_creador = request.user
        if hasattr(obj, 'usuario_modificador_id'):
            obj.usuario_modificador = request.user
        obj.save()
        messages.success(request, 'Registro guardado correctamente.')
        return redirect(list_name)
    context = {'form': form, 'items': model.objects.all().order_by('-id')[:200], 'entity_label': template, 'create_name': create_name, 'edit_name': edit_name, 'detail_name': {'coordinadores': 'afiliados:detalle_coordinador_plan_hormiga', 'enlaces': 'afiliados:detalle_enlace_plan_hormiga', 'estructuras': 'afiliados:detalle_estructura_plan_hormiga', 'responsables': 'afiliados:detalle_responsable_plan_hormiga', 'reuniones': 'afiliados:detalle_reunion_plan_hormiga', 'incidencias': 'afiliados:detalle_incidencia_plan_hormiga'}.get(template), 'comunidad_lookup_url': reverse('afiliados:plan_hormiga_comunidad_lookup'), 'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_plan_hormiga')}
    if template == 'reuniones':
        responsables_modal = []
        for responsable in ResponsableZonaPlanHormiga.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'RESPONSABLE', 'tipo_label': 'Responsable de zona', 'id': responsable.id, 'nombre': responsable.nombre_completo, 'dpi': responsable.dpi, 'comunidad': responsable.comunidad.nombre if responsable.comunidad else ''})
        for coordinador in CoordinadorPlanHormiga.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'COORDINADOR', 'tipo_label': 'Coordinador', 'id': coordinador.id, 'nombre': coordinador.nombre_completo, 'dpi': coordinador.dpi, 'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else ''})
        for enlace in EnlaceTerritorialPlanHormiga.objects.filter(estado=EstadoRegistro.ACTIVO).select_related('comunidad').order_by('nombre_completo'):
            responsables_modal.append({'tipo': 'ENLACE', 'tipo_label': 'Enlace territorial', 'id': enlace.id, 'nombre': enlace.nombre_completo, 'dpi': enlace.dpi, 'comunidad': enlace.comunidad.nombre if enlace.comunidad else ''})
        context['responsables_modal'] = responsables_modal
    if template == 'estructuras':
        coordinadores_modal = []
        for coordinador in CoordinadorPlanHormiga.objects.select_related('comunidad').order_by('nombre_completo'):
            coordinadores_modal.append({'id': coordinador.id, 'nombre': coordinador.nombre_completo, 'dpi': coordinador.dpi, 'comunidad': coordinador.comunidad.nombre if coordinador.comunidad else '', 'estado': coordinador.get_estado_display()})
        context['coordinadores_modal'] = coordinadores_modal
    return safe_render(request, f'afiliados/plan_hormiga/crud_{template}.html', context)

@login_required
def lista_estructura_territorial_plan_hormiga(request):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    comunidades = Comunidad.objects.select_related('sector').order_by('nombre')
    sectores = Sector.objects.order_by('nombre')
    centros = CentroVotacion.objects.order_by('nombre')
    return safe_render(request, 'afiliados/plan_hormiga/estructura_territorial.html', {'comunidades': comunidades, 'sectores': sectores, 'centros': centros, 'total_comunidades': comunidades.count(), 'total_sectores': sectores.count(), 'total_centros': centros.count()})

@login_required
def lista_coordinadores_plan_hormiga(request): return _plan_hormiga_crud(request, CoordinadorPlanHormiga, CoordinadorPlanHormigaForm, 'coordinadores', 'afiliados:lista_coordinadores_plan_hormiga', 'afiliados:crear_coordinador_plan_hormiga', 'afiliados:editar_coordinador_plan_hormiga')
@login_required
def crear_coordinador_plan_hormiga(request): return _plan_hormiga_crud(request, CoordinadorPlanHormiga, CoordinadorPlanHormigaForm, 'coordinadores', 'afiliados:lista_coordinadores_plan_hormiga', 'afiliados:crear_coordinador_plan_hormiga', 'afiliados:editar_coordinador_plan_hormiga')
@login_required
def editar_coordinador_plan_hormiga(request, pk): return _plan_hormiga_crud(request, CoordinadorPlanHormiga, CoordinadorPlanHormigaForm, 'coordinadores', 'afiliados:lista_coordinadores_plan_hormiga', 'afiliados:crear_coordinador_plan_hormiga', 'afiliados:editar_coordinador_plan_hormiga', pk=pk)
@login_required
def detalle_coordinador_plan_hormiga(request, pk):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    return safe_render(request, 'afiliados/plan_hormiga/detalle_generico.html', {'obj': get_object_or_404(CoordinadorPlanHormiga, pk=pk), 'titulo': 'Coordinador de Plan Hormiga'})
@login_required
def lista_enlaces_plan_hormiga(request): return _plan_hormiga_crud(request, EnlaceTerritorialPlanHormiga, EnlaceTerritorialPlanHormigaForm, 'enlaces', 'afiliados:lista_enlaces_plan_hormiga', 'afiliados:crear_enlace_plan_hormiga', 'afiliados:editar_enlace_plan_hormiga')
@login_required
def crear_enlace_plan_hormiga(request): return _plan_hormiga_crud(request, EnlaceTerritorialPlanHormiga, EnlaceTerritorialPlanHormigaForm, 'enlaces', 'afiliados:lista_enlaces_plan_hormiga', 'afiliados:crear_enlace_plan_hormiga', 'afiliados:editar_enlace_plan_hormiga')
@login_required
def editar_enlace_plan_hormiga(request, pk): return _plan_hormiga_crud(request, EnlaceTerritorialPlanHormiga, EnlaceTerritorialPlanHormigaForm, 'enlaces', 'afiliados:lista_enlaces_plan_hormiga', 'afiliados:crear_enlace_plan_hormiga', 'afiliados:editar_enlace_plan_hormiga', pk=pk)
@login_required
def detalle_enlace_plan_hormiga(request, pk):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    return safe_render(request, 'afiliados/plan_hormiga/detalle_generico.html', {'obj': get_object_or_404(EnlaceTerritorialPlanHormiga, pk=pk), 'titulo': 'Enlace territorial'})
@login_required
def lista_estructuras_plan_hormiga(request): return _plan_hormiga_crud(request, EstructuraPlanHormiga, EstructuraPlanHormigaForm, 'estructuras', 'afiliados:lista_estructuras_plan_hormiga', 'afiliados:crear_estructura_plan_hormiga', 'afiliados:editar_estructura_plan_hormiga')
@login_required
def crear_estructura_plan_hormiga(request): return _plan_hormiga_crud(request, EstructuraPlanHormiga, EstructuraPlanHormigaForm, 'estructuras', 'afiliados:lista_estructuras_plan_hormiga', 'afiliados:crear_estructura_plan_hormiga', 'afiliados:editar_estructura_plan_hormiga')
@login_required
def editar_estructura_plan_hormiga(request, pk): return _plan_hormiga_crud(request, EstructuraPlanHormiga, EstructuraPlanHormigaForm, 'estructuras', 'afiliados:lista_estructuras_plan_hormiga', 'afiliados:crear_estructura_plan_hormiga', 'afiliados:editar_estructura_plan_hormiga', pk=pk)
@login_required
def detalle_estructura_plan_hormiga(request, pk):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    estructura = get_object_or_404(EstructuraPlanHormiga.objects.select_related('comunidad', 'sector', 'centro_votacion', 'coordinador_responsable'), pk=pk)
    integrantes = estructura.integrantes.select_related('usuario_registro').all()
    return safe_render(request, 'afiliados/plan_hormiga/detalle_estructura.html', {'estructura': estructura, 'integrantes': integrantes, 'total_integrantes': integrantes.count(), 'empadronamiento_url': reverse('afiliados:verificar_empadronamiento_plan_hormiga')})
@login_required
@require_POST
def eliminar_integrante_estructura_plan_hormiga(request, pk, integrante_id):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    estructura = get_object_or_404(EstructuraPlanHormiga, pk=pk)
    integrante = get_object_or_404(EstructuraIntegrantePlanHormiga, pk=integrante_id, estructura=estructura)
    integrante.delete()
    messages.success(request, 'Integrante eliminado correctamente de la estructura de Plan Hormiga.')
    return redirect('afiliados:detalle_estructura_plan_hormiga', pk=estructura.pk)
@login_required
def lista_responsables_plan_hormiga(request): return _plan_hormiga_crud(request, ResponsableZonaPlanHormiga, ResponsableZonaPlanHormigaForm, 'responsables', 'afiliados:lista_responsables_plan_hormiga', 'afiliados:crear_responsable_plan_hormiga', 'afiliados:editar_responsable_plan_hormiga')
@login_required
def crear_responsable_plan_hormiga(request): return _plan_hormiga_crud(request, ResponsableZonaPlanHormiga, ResponsableZonaPlanHormigaForm, 'responsables', 'afiliados:lista_responsables_plan_hormiga', 'afiliados:crear_responsable_plan_hormiga', 'afiliados:editar_responsable_plan_hormiga')
@login_required
def editar_responsable_plan_hormiga(request, pk): return _plan_hormiga_crud(request, ResponsableZonaPlanHormiga, ResponsableZonaPlanHormigaForm, 'responsables', 'afiliados:lista_responsables_plan_hormiga', 'afiliados:crear_responsable_plan_hormiga', 'afiliados:editar_responsable_plan_hormiga', pk=pk)
@login_required
def detalle_responsable_plan_hormiga(request, pk):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    return safe_render(request, 'afiliados/plan_hormiga/detalle_generico.html', {'obj': get_object_or_404(ResponsableZonaPlanHormiga, pk=pk), 'titulo': 'Responsable de zona'})
@login_required
def lista_reuniones_plan_hormiga(request): return _plan_hormiga_crud(request, ReunionPlanHormiga, ReunionPlanHormigaForm, 'reuniones', 'afiliados:lista_reuniones_plan_hormiga', 'afiliados:crear_reunion_plan_hormiga', 'afiliados:editar_reunion_plan_hormiga')
@login_required
def crear_reunion_plan_hormiga(request): return _plan_hormiga_crud(request, ReunionPlanHormiga, ReunionPlanHormigaForm, 'reuniones', 'afiliados:lista_reuniones_plan_hormiga', 'afiliados:crear_reunion_plan_hormiga', 'afiliados:editar_reunion_plan_hormiga')
@login_required
def editar_reunion_plan_hormiga(request, pk): return _plan_hormiga_crud(request, ReunionPlanHormiga, ReunionPlanHormigaForm, 'reuniones', 'afiliados:lista_reuniones_plan_hormiga', 'afiliados:crear_reunion_plan_hormiga', 'afiliados:editar_reunion_plan_hormiga', pk=pk)
@login_required
def detalle_reunion_plan_hormiga(request, pk):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    return safe_render(request, 'afiliados/plan_hormiga/detalle_reunion.html', {'reunion': get_object_or_404(ReunionPlanHormiga, pk=pk)})
@login_required
def lista_incidencias_plan_hormiga(request): return _plan_hormiga_crud(request, IncidenciaPlanHormiga, IncidenciaPlanHormigaForm, 'incidencias', 'afiliados:lista_incidencias_plan_hormiga', 'afiliados:crear_incidencia_plan_hormiga', 'afiliados:editar_incidencia_plan_hormiga')
@login_required
def crear_incidencia_plan_hormiga(request): return _plan_hormiga_crud(request, IncidenciaPlanHormiga, IncidenciaPlanHormigaForm, 'incidencias', 'afiliados:lista_incidencias_plan_hormiga', 'afiliados:crear_incidencia_plan_hormiga', 'afiliados:editar_incidencia_plan_hormiga')
@login_required
def editar_incidencia_plan_hormiga(request, pk): return _plan_hormiga_crud(request, IncidenciaPlanHormiga, IncidenciaPlanHormigaForm, 'incidencias', 'afiliados:lista_incidencias_plan_hormiga', 'afiliados:crear_incidencia_plan_hormiga', 'afiliados:editar_incidencia_plan_hormiga', pk=pk)
@login_required
def detalle_incidencia_plan_hormiga(request, pk):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    return safe_render(request, 'afiliados/plan_hormiga/detalle_generico.html', {'obj': get_object_or_404(IncidenciaPlanHormiga, pk=pk), 'titulo': 'Incidencia territorial'})
@login_required
def reporte_cobertura_plan_hormiga(request): return panorama_municipal_plan_hormiga(request)
@login_required
def reporte_coordinadores_plan_hormiga(request):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    return safe_render(request, 'afiliados/plan_hormiga/reporte_coordinadores.html', {'coordinadores': CoordinadorPlanHormiga.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha_creacion')})
@login_required
def reporte_reuniones_plan_hormiga(request):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    return safe_render(request, 'afiliados/plan_hormiga/reporte_reuniones.html', {'reuniones': ReunionPlanHormiga.objects.select_related('comunidad', 'sector', 'centro_votacion').order_by('-fecha')})
@login_required
def reporte_crecimiento_mensual_plan_hormiga(request): return panorama_municipal_plan_hormiga(request)

@login_required
@require_GET
def plan_hormiga_comunidad_lookup(request):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    comunidad = Comunidad.objects.select_related('sector').filter(pk=request.GET.get('comunidad_id')).first()
    if not comunidad: return JsonResponse({'ok': False, 'error': 'Comunidad no encontrada.'}, status=404)
    centro = comunidad.sector.centros.first() if comunidad.sector else None
    return JsonResponse({'ok': True, 'data': {'comunidad': comunidad.nombre, 'sector': comunidad.sector.nombre if comunidad.sector else None, 'centro_votacion': centro.nombre if centro else None}})


def _plan_hormiga_extra_crud(request, model, form_class, template_name, list_url, pk=None):
    denied = _require_plan_hormiga(request)
    if denied: return denied
    instance = get_object_or_404(model, pk=pk) if pk else None
    form = form_class(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if not getattr(obj, 'usuario_creador_id', None): obj.usuario_creador = request.user
        obj.usuario_modificador = request.user
        obj.save()
        messages.success(request, 'Registro guardado correctamente.')
        return redirect(list_url)
    return safe_render(request, template_name, {'form': form, 'items': model.objects.all().order_by('-id')[:200]})

@login_required
def lista_visitas_plan_hormiga(request): return _plan_hormiga_extra_crud(request, VisitaPlanHormiga, VisitaPlanHormigaForm, 'afiliados/plan_hormiga/crud_visitas.html', 'afiliados:lista_visitas_plan_hormiga')
@login_required
def editar_visitas_plan_hormiga(request, pk): return _plan_hormiga_extra_crud(request, VisitaPlanHormiga, VisitaPlanHormigaForm, 'afiliados/plan_hormiga/crud_visitas.html', 'afiliados:lista_visitas_plan_hormiga', pk=pk)
@login_required
def lista_seguimientos_plan_hormiga(request): return _plan_hormiga_extra_crud(request, SeguimientoContactoPlanHormiga, SeguimientoContactoPlanHormigaForm, 'afiliados/plan_hormiga/crud_seguimientos.html', 'afiliados:lista_seguimientos_plan_hormiga')
@login_required
def editar_seguimientos_plan_hormiga(request, pk): return _plan_hormiga_extra_crud(request, SeguimientoContactoPlanHormiga, SeguimientoContactoPlanHormigaForm, 'afiliados/plan_hormiga/crud_seguimientos.html', 'afiliados:lista_seguimientos_plan_hormiga', pk=pk)
@login_required
def lista_compromisos_plan_hormiga(request): return _plan_hormiga_extra_crud(request, CompromisoTerritorialPlanHormiga, CompromisoTerritorialPlanHormigaForm, 'afiliados/plan_hormiga/crud_compromisos.html', 'afiliados:lista_compromisos_plan_hormiga')
@login_required
def editar_compromisos_plan_hormiga(request, pk): return _plan_hormiga_extra_crud(request, CompromisoTerritorialPlanHormiga, CompromisoTerritorialPlanHormigaForm, 'afiliados/plan_hormiga/crud_compromisos.html', 'afiliados:lista_compromisos_plan_hormiga', pk=pk)
@login_required
def lista_puntos_visitados_plan_hormiga(request): return _plan_hormiga_extra_crud(request, PuntoVisitadoPlanHormiga, PuntoVisitadoPlanHormigaForm, 'afiliados/plan_hormiga/crud_puntos_visitados.html', 'afiliados:lista_puntos_visitados_plan_hormiga')
@login_required
def editar_puntos_visitados_plan_hormiga(request, pk): return _plan_hormiga_extra_crud(request, PuntoVisitadoPlanHormiga, PuntoVisitadoPlanHormigaForm, 'afiliados/plan_hormiga/crud_puntos_visitados.html', 'afiliados:lista_puntos_visitados_plan_hormiga', pk=pk)
@login_required
def lista_activaciones_plan_hormiga(request): return _plan_hormiga_extra_crud(request, ActivacionTerritorialPlanHormiga, ActivacionTerritorialPlanHormigaForm, 'afiliados/plan_hormiga/crud_activaciones.html', 'afiliados:lista_activaciones_plan_hormiga')
@login_required
def editar_activaciones_plan_hormiga(request, pk): return _plan_hormiga_extra_crud(request, ActivacionTerritorialPlanHormiga, ActivacionTerritorialPlanHormigaForm, 'afiliados/plan_hormiga/crud_activaciones.html', 'afiliados:lista_activaciones_plan_hormiga', pk=pk)
@login_required
def lista_coberturas_visitas_plan_hormiga(request):
    denied = _require_plan_hormiga(request)
    if denied:
        return denied
    form = CoberturaVisitaPlanHormigaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if not getattr(obj, 'usuario_creador_id', None):
            obj.usuario_creador = request.user
        obj.usuario_modificador = request.user
        obj.save()
        messages.success(request, 'Registro de cobertura guardado correctamente.')
        return redirect('afiliados:lista_coberturas_visitas_plan_hormiga')
    items = CoberturaVisitaPlanHormiga.objects.all().order_by('-fecha_corte', '-id')[:200]
    resumen = items.aggregate(
        comunidades_atendidas=Sum('comunidades_atendidas'),
        comunidades_pendientes=Sum('comunidades_pendientes'),
        sectores_visitados=Sum('sectores_visitados'),
        sectores_pendientes=Sum('sectores_pendientes'),
    )
    resumen = {k: (v or 0) for k, v in resumen.items()}
    return safe_render(request, 'afiliados/plan_hormiga/crud_cobertura_visitas.html', {'form': form, 'items': items, 'resumen': resumen})


@login_required
def editar_coberturas_visitas_plan_hormiga(request, pk):
    denied = _require_plan_hormiga(request)
    if denied:
        return denied
    instance = get_object_or_404(CoberturaVisitaPlanHormiga, pk=pk)
    form = CoberturaVisitaPlanHormigaForm(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if not getattr(obj, 'usuario_creador_id', None):
            obj.usuario_creador = request.user
        obj.usuario_modificador = request.user
        obj.save()
        messages.success(request, 'Registro de cobertura actualizado correctamente.')
        return redirect('afiliados:lista_coberturas_visitas_plan_hormiga')
    items = CoberturaVisitaPlanHormiga.objects.all().order_by('-fecha_corte', '-id')[:200]
    resumen = items.aggregate(
        comunidades_atendidas=Sum('comunidades_atendidas'),
        comunidades_pendientes=Sum('comunidades_pendientes'),
        sectores_visitados=Sum('sectores_visitados'),
        sectores_pendientes=Sum('sectores_pendientes'),
    )
    resumen = {k: (v or 0) for k, v in resumen.items()}
    return safe_render(request, 'afiliados/plan_hormiga/crud_cobertura_visitas.html', {'form': form, 'items': items, 'resumen': resumen})
